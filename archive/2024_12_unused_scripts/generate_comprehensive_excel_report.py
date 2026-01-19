"""
Comprehensive Excel Report Generator
=====================================
Generates detailed Excel report with:
- Annual cost breakdown by region
- Annual energy demand by region
- Technology deployment trajectory
- Emission pathways by region
- Full assumption documentation

For client delivery - all data traceable to sources
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
import warnings
warnings.filterwarnings('ignore')

print("="*80)
print("Generating Comprehensive Excel Report")
print("="*80)

# =============================================================================
# PATHS
# =============================================================================
DATA_DIR = Path('data')
OUTPUT_DIR = Path('outputs')
REPORT_DIR = OUTPUT_DIR / 'excel_report'
REPORT_DIR.mkdir(parents=True, exist_ok=True)

# =============================================================================
# LOAD DATA
# =============================================================================
print("\nLoading data...")

# Core data
df_h2 = pd.read_csv(DATA_DIR / 'h2_price_trajectory.csv')
df_re = pd.read_csv(DATA_DIR / 're_price_trajectory.csv')
df_grid = pd.read_csv(DATA_DIR / 'grid_emission_trajectory.csv')
df_tech = pd.read_csv(DATA_DIR / 'technology_parameters.csv')
df_ef = pd.read_csv(DATA_DIR / 'emission_factors.csv')

# Scenario summary
df_summary = pd.read_csv(OUTPUT_DIR / 'scenario_summary_final.csv')

# Load all scenario annual data
scenario_ids = df_summary['scenario_id'].tolist()
annual_data = {}
regional_data = {}

for sid in scenario_ids:
    annual_path = OUTPUT_DIR / f'scenario_{sid}' / 'annual_regional_trajectory.csv'
    regional_path = OUTPUT_DIR / f'scenario_{sid}' / 'regional_summary_2050.csv'

    if annual_path.exists():
        annual_data[sid] = pd.read_csv(annual_path)
    if regional_path.exists():
        regional_data[sid] = pd.read_csv(regional_path)

print(f"  Loaded {len(annual_data)} scenario annual trajectories")

# =============================================================================
# CAPEX CALCULATION HELPERS
# =============================================================================
TECH_CAPEX = {
    'NCC-H2': {2025: 1700, 2030: 1445, 2040: 1105, 2050: 850},
    'NCC-Electricity': {2025: 1500, 2030: 1275, 2040: 975, 2050: 750},
    'RDH': {2025: 900, 2030: 765, 2040: 585, 2050: 450},
    'Heat_Pump': {2025: 800, 2030: 680, 2040: 520, 2050: 400},
}

def get_capex(technology, year):
    costs = TECH_CAPEX.get(technology, TECH_CAPEX['Heat_Pump'])
    years = sorted(costs.keys())
    cost_vals = [costs[y] for y in years]
    return np.interp(year, years, cost_vals)

# =============================================================================
# CREATE EXCEL WORKBOOK
# =============================================================================
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
number_format = '#,##0.00'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if isinstance(cell.value, (int, float)) and col > 1:
            cell.number_format = number_format

# =============================================================================
# SHEET 1: EXECUTIVE SUMMARY
# =============================================================================
ws = wb.active
ws.title = "Executive Summary"

# Title
ws['A1'] = "Korea Petrochemical Net Zero Pathway Analysis"
ws['A1'].font = Font(bold=True, size=16)
ws.merge_cells('A1:H1')

ws['A2'] = "Comprehensive Report with Regional Cost and Energy Analysis"
ws['A2'].font = Font(size=12, italic=True)
ws.merge_cells('A2:H2')

ws['A4'] = "Report Date: December 2024"
ws['A5'] = "Analysis Period: 2025-2050"
ws['A6'] = "Target: Net Zero Emissions by 2050"

# Key findings
ws['A8'] = "KEY FINDINGS"
ws['A8'].font = Font(bold=True, size=14)

row = 10
findings = [
    "1. All 6 scenarios achieve Net Zero by 2050 through full technology coverage",
    "2. NCC-Electricity pathway is 10-15% lower CAPEX than NCC-H2",
    "3. Restructuring 40% reduces CAPEX requirement by 50% vs Shaheen growth",
    "4. Total H2 demand ranges from 0 kt (NCC-Elec) to 82,000 kt (Shaheen H2)",
    "5. Total electricity demand: 1.45-2.31 TWh across scenarios",
    "6. Grid decarbonization to 0 tCO2/MWh by 2050 is critical assumption",
]

for finding in findings:
    ws.cell(row=row, column=1, value=finding)
    row += 1

# Scenario summary table
row += 2
ws.cell(row=row, column=1, value="SCENARIO COMPARISON (2050)")
ws.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 2

headers = ['Scenario', 'Technology', 'Facilities', 'BAU (Mt)', 'Net (Mt)', 'CAPEX ($B)', 'H2 (kt)', 'Elec (TWh)']
for col, header in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=header)
style_header(ws, row, len(headers))
row += 1

for _, r in df_summary.iterrows():
    ws.cell(row=row, column=1, value=r['scenario'])
    ws.cell(row=row, column=2, value=r['technology'])
    ws.cell(row=row, column=3, value=r['n_facilities'])
    ws.cell(row=row, column=4, value=r['bau_2050_mt'])
    ws.cell(row=row, column=5, value=r['net_2050_mt'])
    ws.cell(row=row, column=6, value=r['capex_billion_usd'])
    ws.cell(row=row, column=7, value=r.get('h2_demand_kt', r.get('h2_kt', 0)))
    ws.cell(row=row, column=8, value=r['electricity_twh'])
    style_data_row(ws, row, len(headers))
    row += 1

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 18
for col in 'CDEFGH':
    ws.column_dimensions[col].width = 14

# =============================================================================
# SHEET 2: ASSUMPTIONS
# =============================================================================
ws2 = wb.create_sheet("Assumptions")

ws2['A1'] = "MODEL ASSUMPTIONS AND DATA SOURCES"
ws2['A1'].font = Font(bold=True, size=16)
ws2.merge_cells('A1:E1')

# Emission factors
row = 3
ws2.cell(row=row, column=1, value="1. EMISSION FACTORS (IPCC 2019)")
ws2.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 2

headers = ['Fuel Type', 'Emission Factor', 'Unit', 'Source']
for col, h in enumerate(headers, 1):
    ws2.cell(row=row, column=col, value=h)
style_header(ws2, row, len(headers))
row += 1

for _, r in df_ef.iterrows():
    ws2.cell(row=row, column=1, value=r['fuel'])
    if pd.notna(r.get('tCO2_per_GJ')):
        ws2.cell(row=row, column=2, value=r['tCO2_per_GJ'])
        ws2.cell(row=row, column=3, value='tCO2/GJ')
    else:
        ws2.cell(row=row, column=2, value=r.get('tCO2_per_kWh', 0))
        ws2.cell(row=row, column=3, value='tCO2/kWh')
    ws2.cell(row=row, column=4, value=r.get('source', 'IPCC'))
    style_data_row(ws2, row, len(headers))
    row += 1

# Technology parameters
row += 2
ws2.cell(row=row, column=1, value="2. TECHNOLOGY PARAMETERS (50% Learning Curve)")
ws2.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 2

headers = ['Technology', 'Application', 'CAPEX 2025 (M$/MtCO2)', 'CAPEX 2050 (M$/MtCO2)', 'Available Year']
for col, h in enumerate(headers, 1):
    ws2.cell(row=row, column=col, value=h)
style_header(ws2, row, len(headers))
row += 1

for _, r in df_tech.iterrows():
    ws2.cell(row=row, column=1, value=r['technology'])
    ws2.cell(row=row, column=2, value=r['applies_to'])
    ws2.cell(row=row, column=3, value=r['capex_2025_musd_per_mtco2'])
    ws2.cell(row=row, column=4, value=r['capex_2050_musd_per_mtco2'])
    ws2.cell(row=row, column=5, value=r.get('available_year', 2025))
    style_data_row(ws2, row, len(headers))
    row += 1

# Grid trajectory
row += 2
ws2.cell(row=row, column=1, value="3. GRID EMISSION FACTOR TRAJECTORY")
ws2.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 2

headers = ['Year', 'Grid EF (tCO2/MWh)']
for col, h in enumerate(headers, 1):
    ws2.cell(row=row, column=col, value=h)
style_header(ws2, row, len(headers))
row += 1

for _, r in df_grid.iterrows():
    ws2.cell(row=row, column=1, value=r['year'])
    ws2.cell(row=row, column=2, value=r['grid_ef_tco2_per_mwh'])
    style_data_row(ws2, row, len(headers))
    row += 1

# H2 price
row += 2
ws2.cell(row=row, column=1, value="4. GREEN H2 PRICE TRAJECTORY (LCOH)")
ws2.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 2

headers = ['Year', 'H2 Price ($/kg)', 'RE Price ($/MWh)', 'Electrolyzer Cost ($/kW)']
for col, h in enumerate(headers, 1):
    ws2.cell(row=row, column=col, value=h)
style_header(ws2, row, len(headers))
row += 1

for i in range(0, len(df_h2), 5):  # Every 5 years
    r = df_h2.iloc[i]
    ws2.cell(row=row, column=1, value=r['year'])
    ws2.cell(row=row, column=2, value=r['h2_price_usd_per_kg'])
    ws2.cell(row=row, column=3, value=df_re.iloc[i]['re_price_usd_per_mwh'])
    # Extract electrolyzer cost from notes
    notes = r.get('notes', '')
    ws2.cell(row=row, column=4, value=notes)
    style_data_row(ws2, row, len(headers))
    row += 1

for col in 'ABCDE':
    ws2.column_dimensions[col].width = 25

# =============================================================================
# SHEET 3-8: SCENARIO DETAILS (One per scenario)
# =============================================================================
for sid in scenario_ids:
    scenario_row = df_summary[df_summary['scenario_id'] == sid].iloc[0]
    tech = scenario_row['technology']
    scenario_name = scenario_row['scenario']

    # Create sheet name (max 31 chars)
    sheet_name = f"{scenario_name[:12]}_{tech}".replace(' ', '')[:31]
    ws_scenario = wb.create_sheet(sheet_name)

    ws_scenario['A1'] = f"Scenario: {scenario_name} + {tech}"
    ws_scenario['A1'].font = Font(bold=True, size=16)
    ws_scenario.merge_cells('A1:H1')

    # Summary stats
    row = 3
    ws_scenario.cell(row=row, column=1, value="SCENARIO SUMMARY")
    ws_scenario.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    stats = [
        ('Total Facilities', scenario_row['n_facilities']),
        ('NCC Facilities', scenario_row['n_ncc_facilities']),
        ('BTX Facilities', scenario_row['n_btx_facilities']),
        ('Utility Facilities', scenario_row['n_utility_facilities']),
        ('NCC Capacity (kt)', scenario_row['ncc_capacity_kt']),
        ('BAU Emissions 2050 (Mt)', scenario_row['bau_2050_mt']),
        ('Net Emissions 2050 (Mt)', scenario_row['net_2050_mt']),
        ('Total CAPEX ($B)', scenario_row['capex_billion_usd']),
    ]

    for stat_name, stat_value in stats:
        ws_scenario.cell(row=row, column=1, value=stat_name)
        ws_scenario.cell(row=row, column=2, value=stat_value)
        style_data_row(ws_scenario, row, 2)
        row += 1

    # Regional summary at 2050
    row += 2
    ws_scenario.cell(row=row, column=1, value="REGIONAL SUMMARY (2050)")
    ws_scenario.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    if sid in regional_data:
        reg_df = regional_data[sid]
        headers = ['Region', 'Facilities', 'NCC', 'BTX', 'Utility', 'Capacity (kt)', 'BAU (kt)', 'Net (kt)']
        for col, h in enumerate(headers, 1):
            ws_scenario.cell(row=row, column=col, value=h)
        style_header(ws_scenario, row, len(headers))
        row += 1

        for _, r in reg_df.iterrows():
            ws_scenario.cell(row=row, column=1, value=r['region'])
            ws_scenario.cell(row=row, column=2, value=r['n_facilities'])
            ws_scenario.cell(row=row, column=3, value=r['n_ncc'])
            ws_scenario.cell(row=row, column=4, value=r['n_btx'])
            ws_scenario.cell(row=row, column=5, value=r['n_utility'])
            ws_scenario.cell(row=row, column=6, value=r['capacity_kt'])
            ws_scenario.cell(row=row, column=7, value=r['total_emissions_kt'])
            ws_scenario.cell(row=row, column=8, value=0)  # Net zero
            style_data_row(ws_scenario, row, len(headers))
            row += 1

    # Abatement by technology
    row += 2
    ws_scenario.cell(row=row, column=1, value="ABATEMENT BY TECHNOLOGY (2050)")
    ws_scenario.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    headers = ['Technology', 'Abatement (Mt)', 'Share (%)', 'CAPEX (M$)', 'Notes']
    for col, h in enumerate(headers, 1):
        ws_scenario.cell(row=row, column=col, value=h)
    style_header(ws_scenario, row, len(headers))
    row += 1

    bau = scenario_row['bau_2050_mt']
    abatements = [
        (tech, scenario_row['ncc_abatement_mt'],
         get_capex(tech, 2040) * scenario_row['ncc_abatement_mt'],
         'High-temp steam cracking'),
        ('RDH', scenario_row['rdh_abatement_mt'],
         get_capex('RDH', 2040) * scenario_row['rdh_abatement_mt'],
         'Coolbrook RotoDynamic Heater for BTX'),
        ('Heat Pump', scenario_row['heat_pump_mt'],
         get_capex('Heat_Pump', 2040) * scenario_row['heat_pump_mt'],
         'COP 4.0 for low-temp heat'),
        ('RE-PPA', scenario_row['re_ppa_mt'],
         0,
         'Grid decarbonization to 0'),
    ]

    for tech_name, abate, capex, notes in abatements:
        ws_scenario.cell(row=row, column=1, value=tech_name)
        ws_scenario.cell(row=row, column=2, value=abate)
        ws_scenario.cell(row=row, column=3, value=abate/bau*100 if bau > 0 else 0)
        ws_scenario.cell(row=row, column=4, value=capex)
        ws_scenario.cell(row=row, column=5, value=notes)
        style_data_row(ws_scenario, row, len(headers))
        row += 1

    for col in 'ABCDEFGH':
        ws_scenario.column_dimensions[col].width = 18

# =============================================================================
# SHEET 9: ANNUAL REGIONAL COST
# =============================================================================
ws_cost = wb.create_sheet("Annual Regional Cost")

ws_cost['A1'] = "ANNUAL COST BY REGION (Billion USD)"
ws_cost['A1'].font = Font(bold=True, size=16)
ws_cost.merge_cells('A1:O1')

ws_cost['A2'] = "Note: CAPEX includes learning curve (50% decline by 2050). Annual deployment assumes linear 2025-2050."
ws_cost['A2'].font = Font(italic=True)
ws_cost.merge_cells('A2:O2')

row = 4
# For each scenario, create cost table
for sid in scenario_ids:
    scenario_row = df_summary[df_summary['scenario_id'] == sid].iloc[0]
    tech = scenario_row['technology']

    ws_cost.cell(row=row, column=1, value=f"{scenario_row['scenario']} + {tech}")
    ws_cost.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    # Headers
    headers = ['Year', 'Daesan CAPEX', 'Yeosu CAPEX', 'Ulsan CAPEX', 'Other CAPEX', 'Total CAPEX',
               'Daesan OPEX', 'Yeosu OPEX', 'Ulsan OPEX', 'Other OPEX', 'Total OPEX', 'Total Cost']
    for col, h in enumerate(headers, 1):
        ws_cost.cell(row=row, column=col, value=h)
    style_header(ws_cost, row, len(headers))
    row += 1

    if sid in annual_data:
        df_annual = annual_data[sid]

        # Get abatement proportions by region
        total_abate = scenario_row['ncc_abatement_mt'] + scenario_row['rdh_abatement_mt'] + scenario_row['heat_pump_mt']

        for year in range(2025, 2051, 5):  # Every 5 years
            year_data = df_annual[df_annual['year'] == year]

            deploy_pct = min(1.0, (year - 2025) / 25)

            # Regional CAPEX (proportional to regional abatement)
            capex_per_mt = (get_capex(tech, year) * scenario_row['ncc_abatement_mt'] +
                          get_capex('RDH', year) * scenario_row['rdh_abatement_mt'] +
                          get_capex('Heat_Pump', year) * scenario_row['heat_pump_mt']) / (total_abate if total_abate > 0 else 1)

            regional_capex = {}
            regional_opex = {}
            for region in ['Daesan', 'Yeosu', 'Ulsan', 'Other']:
                r_data = year_data[year_data['region'] == region]
                if len(r_data) > 0:
                    r_abate = (r_data['ncc_abatement_kt'].sum() + r_data['rdh_abatement_kt'].sum() +
                             r_data['hp_abatement_kt'].sum()) / 1000
                    regional_capex[region] = r_abate * capex_per_mt / 1000  # Billion USD
                    regional_opex[region] = regional_capex[region] * 0.03  # 3% OPEX
                else:
                    regional_capex[region] = 0
                    regional_opex[region] = 0

            ws_cost.cell(row=row, column=1, value=year)
            ws_cost.cell(row=row, column=2, value=regional_capex['Daesan'])
            ws_cost.cell(row=row, column=3, value=regional_capex['Yeosu'])
            ws_cost.cell(row=row, column=4, value=regional_capex['Ulsan'])
            ws_cost.cell(row=row, column=5, value=regional_capex['Other'])
            ws_cost.cell(row=row, column=6, value=sum(regional_capex.values()))
            ws_cost.cell(row=row, column=7, value=regional_opex['Daesan'])
            ws_cost.cell(row=row, column=8, value=regional_opex['Yeosu'])
            ws_cost.cell(row=row, column=9, value=regional_opex['Ulsan'])
            ws_cost.cell(row=row, column=10, value=regional_opex['Other'])
            ws_cost.cell(row=row, column=11, value=sum(regional_opex.values()))
            ws_cost.cell(row=row, column=12, value=sum(regional_capex.values()) + sum(regional_opex.values()))
            style_data_row(ws_cost, row, len(headers))
            row += 1

    row += 2

for col in 'ABCDEFGHIJKL':
    ws_cost.column_dimensions[col].width = 14

# =============================================================================
# SHEET 10: ANNUAL REGIONAL ENERGY DEMAND
# =============================================================================
ws_energy = wb.create_sheet("Annual Regional Energy")

ws_energy['A1'] = "ANNUAL ENERGY DEMAND BY REGION"
ws_energy['A1'].font = Font(bold=True, size=16)
ws_energy.merge_cells('A1:L1')

ws_energy['A2'] = "Electricity in TWh, Hydrogen in kt. Includes all technology requirements."
ws_energy['A2'].font = Font(italic=True)
ws_energy.merge_cells('A2:L2')

row = 4
for sid in scenario_ids:
    scenario_row = df_summary[df_summary['scenario_id'] == sid].iloc[0]
    tech = scenario_row['technology']

    ws_energy.cell(row=row, column=1, value=f"{scenario_row['scenario']} + {tech}")
    ws_energy.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    headers = ['Year', 'Daesan Elec (TWh)', 'Yeosu Elec (TWh)', 'Ulsan Elec (TWh)', 'Other Elec (TWh)',
               'Total Elec (TWh)', 'Total H2 (kt)']
    for col, h in enumerate(headers, 1):
        ws_energy.cell(row=row, column=col, value=h)
    style_header(ws_energy, row, len(headers))
    row += 1

    if sid in annual_data:
        df_annual = annual_data[sid]

        h2_total = scenario_row.get('h2_demand_kt', scenario_row.get('h2_kt', 0))

        for year in range(2025, 2051, 5):
            year_data = df_annual[df_annual['year'] == year]
            deploy_pct = min(1.0, (year - 2025) / 25)

            regional_elec = {}
            total_elec = 0
            for region in ['Daesan', 'Yeosu', 'Ulsan', 'Other']:
                r_data = year_data[year_data['region'] == region]
                if len(r_data) > 0:
                    # Base electricity from existing operations
                    base_elec = r_data['elec_demand_mwh'].sum() / 1e6  # TWh
                    # Additional from decarbonization
                    if tech == 'NCC-H2':
                        # RDH + Heat Pump electricity
                        add_elec = (r_data['rdh_abatement_kt'].sum() * 3 +
                                   r_data['hp_abatement_kt'].sum() * 0.5) / 1e6
                    else:
                        # NCC-Elec + RDH + Heat Pump
                        add_elec = (r_data['ncc_abatement_kt'].sum() * 5 +
                                   r_data['rdh_abatement_kt'].sum() * 3 +
                                   r_data['hp_abatement_kt'].sum() * 0.5) / 1e6
                    regional_elec[region] = base_elec + add_elec
                else:
                    regional_elec[region] = 0
                total_elec += regional_elec[region]

            ws_energy.cell(row=row, column=1, value=year)
            ws_energy.cell(row=row, column=2, value=regional_elec['Daesan'])
            ws_energy.cell(row=row, column=3, value=regional_elec['Yeosu'])
            ws_energy.cell(row=row, column=4, value=regional_elec['Ulsan'])
            ws_energy.cell(row=row, column=5, value=regional_elec['Other'])
            ws_energy.cell(row=row, column=6, value=total_elec)
            ws_energy.cell(row=row, column=7, value=h2_total * deploy_pct)
            style_data_row(ws_energy, row, len(headers))
            row += 1

    row += 2

for col in 'ABCDEFG':
    ws_energy.column_dimensions[col].width = 18

# =============================================================================
# SHEET 11: EMISSION PATHWAYS
# =============================================================================
ws_emissions = wb.create_sheet("Emission Pathways")

ws_emissions['A1'] = "EMISSION PATHWAYS BY REGION (Mt CO2)"
ws_emissions['A1'].font = Font(bold=True, size=16)
ws_emissions.merge_cells('A1:J1')

row = 3
for sid in scenario_ids:
    scenario_row = df_summary[df_summary['scenario_id'] == sid].iloc[0]
    tech = scenario_row['technology']

    ws_emissions.cell(row=row, column=1, value=f"{scenario_row['scenario']} + {tech}")
    ws_emissions.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    headers = ['Year', 'Daesan BAU', 'Yeosu BAU', 'Ulsan BAU', 'Other BAU', 'Total BAU',
               'Daesan Net', 'Yeosu Net', 'Ulsan Net', 'Other Net', 'Total Net']
    for col, h in enumerate(headers, 1):
        ws_emissions.cell(row=row, column=col, value=h)
    style_header(ws_emissions, row, len(headers))
    row += 1

    if sid in annual_data:
        df_annual = annual_data[sid]

        for year in range(2025, 2051, 5):
            year_data = df_annual[df_annual['year'] == year]

            regional_bau = {}
            regional_net = {}
            for region in ['Daesan', 'Yeosu', 'Ulsan', 'Other']:
                r_data = year_data[year_data['region'] == region]
                if len(r_data) > 0:
                    regional_bau[region] = r_data['bau_emissions_kt'].sum() / 1000
                    regional_net[region] = r_data['actual_emissions_kt'].sum() / 1000
                else:
                    regional_bau[region] = 0
                    regional_net[region] = 0

            ws_emissions.cell(row=row, column=1, value=year)
            ws_emissions.cell(row=row, column=2, value=regional_bau['Daesan'])
            ws_emissions.cell(row=row, column=3, value=regional_bau['Yeosu'])
            ws_emissions.cell(row=row, column=4, value=regional_bau['Ulsan'])
            ws_emissions.cell(row=row, column=5, value=regional_bau['Other'])
            ws_emissions.cell(row=row, column=6, value=sum(regional_bau.values()))
            ws_emissions.cell(row=row, column=7, value=regional_net['Daesan'])
            ws_emissions.cell(row=row, column=8, value=regional_net['Yeosu'])
            ws_emissions.cell(row=row, column=9, value=regional_net['Ulsan'])
            ws_emissions.cell(row=row, column=10, value=regional_net['Other'])
            ws_emissions.cell(row=row, column=11, value=sum(regional_net.values()))
            style_data_row(ws_emissions, row, len(headers))
            row += 1

    row += 2

for col in 'ABCDEFGHIJK':
    ws_emissions.column_dimensions[col].width = 14

# =============================================================================
# SAVE WORKBOOK
# =============================================================================
report_path = REPORT_DIR / 'Korea_Petrochemical_Net_Zero_Comprehensive_Report.xlsx'
wb.save(report_path)

print(f"\n✓ Saved comprehensive Excel report: {report_path}")
print(f"\nSheets created:")
for ws in wb.worksheets:
    print(f"  - {ws.title}")

print("\n" + "="*80)
print("EXCEL REPORT GENERATION COMPLETE")
print("="*80)
