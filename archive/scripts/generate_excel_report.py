"""
Generate Comprehensive Excel Report for Korea Petrochemical Net Zero Transition
===============================================================================
Author: PLANiT Institute
Date: 2025

This script generates a professional Excel report with clear documentation of:
- All assumptions and their sources
- Emission calculation methodologies
- Technology parameters and costs
- Energy demand projections
- Optimization results
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference, AreaChart
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# =============================================================================
# PATHS
# =============================================================================
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
STREAMLIT_DATA_DIR = BASE_DIR / "streamlit_data"
OUTPUT_DIR = BASE_DIR / "reports"
OUTPUT_DIR.mkdir(exist_ok=True)

# =============================================================================
# STYLES
# =============================================================================
def create_styles(wb):
    """Create named styles for the workbook"""

    # Header style (dark blue)
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, color='FFFFFF', size=11)
    header_style.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_style.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wb.add_named_style(header_style)

    # Sub-header style (light blue)
    subheader_style = NamedStyle(name='subheader_style')
    subheader_style.font = Font(bold=True, color='1F4E79', size=10)
    subheader_style.fill = PatternFill(start_color='D6E3F8', end_color='D6E3F8', fill_type='solid')
    subheader_style.alignment = Alignment(horizontal='left', vertical='center')
    subheader_style.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wb.add_named_style(subheader_style)

    # Data cell style
    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='right', vertical='center')
    data_style.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wb.add_named_style(data_style)

    # Title style
    title_style = NamedStyle(name='title_style')
    title_style.font = Font(bold=True, color='1F4E79', size=14)
    title_style.alignment = Alignment(horizontal='left', vertical='center')
    wb.add_named_style(title_style)

    # Source style (italic, gray)
    source_style = NamedStyle(name='source_style')
    source_style.font = Font(italic=True, color='666666', size=9)
    source_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    wb.add_named_style(source_style)

    return wb


def write_section_header(ws, row, col, title, width=5):
    """Write a section header spanning multiple columns"""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+width-1)
    cell = ws.cell(row=row, column=col, value=title)
    cell.style = 'title_style'
    return row + 1


def write_table(ws, start_row, start_col, df, include_index=True):
    """Write a DataFrame as a formatted table"""
    if include_index:
        df = df.reset_index()

    # Write headers
    for c_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.style = 'header_style'

    # Write data
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row+1):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.style = 'data_style'

            # Format numbers
            if isinstance(value, float):
                if abs(value) >= 1000:
                    cell.number_format = '#,##0'
                elif abs(value) >= 1:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.0000'

    return start_row + len(df) + 2


def add_source_note(ws, row, col, text):
    """Add a source/reference note"""
    cell = ws.cell(row=row, column=col, value=text)
    cell.style = 'source_style'
    return row + 1


# =============================================================================
# LOAD DATA
# =============================================================================
def load_all_data():
    """Load all necessary data files"""
    data = {}

    # Trajectories
    data['bau_traj'] = pd.read_csv(STREAMLIT_DATA_DIR / 'bau_trajectory_2025_2050.csv')
    data['opt_traj'] = pd.read_csv(STREAMLIT_DATA_DIR / 'optimization_trajectory.csv')
    data['macc_annual'] = pd.read_csv(STREAMLIT_DATA_DIR / 'macc_annual_2025_2050.csv')

    # Input data
    data['facilities'] = pd.read_csv(DATA_DIR / 'facility_database_with_regions.csv')
    data['energy_intensities'] = pd.read_csv(DATA_DIR / 'energy_intensities.csv')
    data['emission_factors'] = pd.read_csv(DATA_DIR / 'emission_factors.csv')
    data['tech_params'] = pd.read_csv(DATA_DIR / 'technology_parameters.csv')
    data['grid_ef'] = pd.read_csv(DATA_DIR / 'grid_emission_trajectory.csv')
    data['grid_price'] = pd.read_csv(DATA_DIR / 'grid_price_trajectory.csv')
    data['h2_price'] = pd.read_csv(DATA_DIR / 'h2_price_trajectory.csv')
    data['re_price'] = pd.read_csv(DATA_DIR / 're_price_trajectory.csv')
    data['demand_growth'] = pd.read_csv(DATA_DIR / 'demand_growth_trajectory.csv')

    return data


# =============================================================================
# SHEET 1: EXECUTIVE SUMMARY
# =============================================================================
def create_executive_summary(wb, data):
    """Create Executive Summary sheet"""
    ws = wb.create_sheet("1. Executive Summary")

    row = 1

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Korea Petrochemical Industry Net Zero Transition Analysis"
    ws['A1'].font = Font(bold=True, size=16, color='1F4E79')
    row = 3

    # Study Info
    ws['A3'] = "Study Information"
    ws['A3'].font = Font(bold=True, size=12, color='1F4E79')

    info_data = [
        ("Author", "PLANiT Institute"),
        ("Publication", "Carbon Neutrality (Springer Nature), 2025"),
        ("Scope", "248 petrochemical facilities across Korea"),
        ("Baseline Year", "2025"),
        ("Target Year", "2050 (Net Zero)"),
        ("Baseline Emissions", f"{data['bau_traj']['total_emissions_mt'].iloc[0]:.2f} MtCO2/year"),
    ]

    row = 5
    for label, value in info_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Key Results
    ws.cell(row=row, column=1, value="Key Results (2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt_2050 = data['opt_traj'][data['opt_traj']['year'] == 2050].iloc[0]
    bau_2050 = data['bau_traj'][data['bau_traj']['year'] == 2050].iloc[0]

    results = [
        ("BAU Emissions (2050)", f"{bau_2050['total_emissions_mt']:.2f} MtCO2"),
        ("Net Zero Achieved", f"{opt_2050['actual_emissions_mt']:.4f} MtCO2"),
        ("Total CAPEX Investment", f"${data['opt_traj']['cumulative_capex_musd'].max()/1000:.1f} Billion"),
        ("Electricity Demand (2050)", f"{opt_2050['electricity_consumption_increase_twh']:.1f} TWh"),
        ("Key Technology", "NCC-Electricity (Electric Cracker)"),
    ]

    for label, value in results:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Technology Deployment Summary
    ws.cell(row=row, column=1, value="Technology Deployment (2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    tech_summary = [
        ("NCC-Electricity", f"{opt_2050['ncc_elec_mt']:.1f} MtCO2", "Electric naphtha cracking"),
        ("Heat Pump", f"{opt_2050['heat_pump_mt']:.1f} MtCO2", "Industrial heat pumps (<165C)"),
        ("RE PPA", f"{opt_2050['re_ppa_mt']:.1f} MtCO2", "Renewable power purchase"),
    ]

    ws.cell(row=row, column=1, value="Technology").style = 'header_style'
    ws.cell(row=row, column=2, value="Abatement").style = 'header_style'
    ws.cell(row=row, column=3, value="Description").style = 'header_style'
    row += 1

    for tech, abate, desc in tech_summary:
        ws.cell(row=row, column=1, value=tech)
        ws.cell(row=row, column=2, value=abate)
        ws.cell(row=row, column=3, value=desc)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40

    return wb


# =============================================================================
# SHEET 2: ASSUMPTIONS & REFERENCES
# =============================================================================
def create_assumptions_sheet(wb, data):
    """Create Assumptions & References sheet"""
    ws = wb.create_sheet("2. Assumptions & References")

    row = 1

    # Title
    ws['A1'] = "Model Assumptions and Data Sources"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    row = 3

    # Section 1: Emission Factors
    ws.cell(row=row, column=1, value="1. Emission Factors").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ef_data = [
        ("Fuel", "Emission Factor", "Unit", "Source"),
        ("Naphtha", "0.0542", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("LNG", "0.0561", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("Fuel Gas", "0.050", "tCO2/GJ", "API Compendium 2021 (Mixed refinery gas)"),
        ("Byproduct Gas", "0.048", "tCO2/GJ", "API Compendium 2021 (Higher H2 content)"),
        ("LPG", "0.0631", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("Fuel Oil", "0.0773", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("Diesel", "0.0741", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("Green H2", "0.0", "tCO2/kg", "Zero-emission (electrolysis)"),
    ]

    for i, (fuel, ef, unit, source) in enumerate(ef_data):
        if i == 0:
            for j, val in enumerate([fuel, ef, unit, source], 1):
                ws.cell(row=row, column=j, value=val).style = 'header_style'
        else:
            ws.cell(row=row, column=1, value=fuel)
            ws.cell(row=row, column=2, value=ef)
            ws.cell(row=row, column=3, value=unit)
            ws.cell(row=row, column=4, value=source)
        row += 1

    row += 2

    # Section 2: Grid Emission Factor Trajectory
    ws.cell(row=row, column=1, value="2. Grid Emission Factor Trajectory").font = Font(bold=True, size=12, color='1F4E79')
    row += 1
    ws.cell(row=row, column=1, value="Source: Korea 10th Basic Plan for Electricity Supply and Demand (2023)").font = Font(italic=True, size=9)
    row += 2

    grid_ef_summary = data['grid_ef'][data['grid_ef']['year'].isin([2025, 2030, 2035, 2040, 2045, 2050])]

    ws.cell(row=row, column=1, value="Year").style = 'header_style'
    ws.cell(row=row, column=2, value="Grid EF (tCO2/MWh)").style = 'header_style'
    ws.cell(row=row, column=3, value="Reduction vs 2025").style = 'header_style'
    row += 1

    base_ef = 0.436
    for _, r in grid_ef_summary.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(r['grid_ef_tco2_per_mwh'], 3))
        ws.cell(row=row, column=3, value=f"{(1 - r['grid_ef_tco2_per_mwh']/base_ef)*100:.0f}%")
        row += 1

    row += 2

    # Section 3: Technology Parameters
    ws.cell(row=row, column=1, value="3. Technology Parameters").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    tech_data = [
        ("Technology", "Key Parameter", "Value", "Source", "TRL", "Available"),
        ("Heat Pump", "COP", "4.0", "Kosmadakis et al. 2020", "9", "2025"),
        ("Heat Pump", "CAPEX (2025)", "$800/tCO2", "McKinsey 2024", "", ""),
        ("Heat Pump", "CAPEX (2050)", "$400/tCO2", "Learning curve (50% reduction)", "", ""),
        ("NCC-Electricity", "Electricity", "5.0 MWh/ton C2H4", "BASF/SABIC/Linde Pilot 2024", "8", "2030"),
        ("NCC-Electricity", "CAPEX (2025)", "$1,500/t-C2H4/yr", "Toribio-Ramirez et al. 2025", "", ""),
        ("NCC-Electricity", "CAPEX (2050)", "$900/t-C2H4/yr", "Learning curve (40% reduction)", "", ""),
        ("NCC-H2", "H2 Consumption", "0.2 ton/ton C2H4", "Lummus Tech 2023 (Energy Only)", "7", "2030"),
        ("NCC-H2", "CAPEX (2025)", "$1,700/t-C2H4/yr", "Thunder Said Energy 2023", "", ""),
        ("NCC-H2", "CAPEX (2050)", "$780/t-C2H4/yr", "Learning curve (54% reduction)", "", ""),
        ("RDH", "Efficiency", "93%", "Coolbrook 2024", "8", "2026"),
        ("RDH", "Temperature", "Up to 1,700C", "Coolbrook 2024", "", ""),
        ("RE PPA", "Price (2025)", "$129/MWh", "PLANiT Institute", "-", "2025"),
        ("RE PPA", "Price (2050)", "$191/MWh", "PLANiT Institute", "", ""),
    ]

    for i, vals in enumerate(tech_data):
        if i == 0:
            for j, val in enumerate(vals, 1):
                ws.cell(row=row, column=j, value=val).style = 'header_style'
        else:
            for j, val in enumerate(vals, 1):
                ws.cell(row=row, column=j, value=val)
        row += 1

    row += 2

    # Section 4: Price Trajectories
    ws.cell(row=row, column=1, value="4. Price Trajectories").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    years = [2025, 2030, 2035, 2040, 2045, 2050]

    ws.cell(row=row, column=1, value="Year").style = 'header_style'
    ws.cell(row=row, column=2, value="Grid Price ($/MWh)").style = 'header_style'
    ws.cell(row=row, column=3, value="RE PPA ($/MWh)").style = 'header_style'
    ws.cell(row=row, column=4, value="H2 Price ($/kg)").style = 'header_style'
    row += 1

    for year in years:
        grid_p = data['grid_price'][data['grid_price']['year'] == year]['grid_price_usd_per_mwh'].values[0]
        re_p = data['re_price'][data['re_price']['year'] == year]['re_price_usd_per_mwh'].values[0]
        h2_p = data['h2_price'][data['h2_price']['year'] == year]['h2_price_usd_per_kg'].values[0]

        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=round(grid_p, 2))
        ws.cell(row=row, column=3, value=round(re_p, 2))
        ws.cell(row=row, column=4, value=round(h2_p, 2))
        row += 1

    ws.cell(row=row, column=1, value="Source: PLANiT Institute assumptions based on literature review").font = Font(italic=True, size=9)
    row += 2

    # Section 5: Demand Growth
    ws.cell(row=row, column=1, value="5. Demand Growth Assumptions").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    growth_data = [
        ("Period", "Annual Growth Rate", "Operating Rate", "Rationale"),
        ("2025-2030", "1.5%", "70%", "Baseline growth; 70% ops (Chinese overcapacity)"),
        ("2031-2035", "1.3%", "70%", "Moderate growth"),
        ("2036-2040", "1.0%", "70%", "Market maturation"),
        ("2041-2050", "0.5-0.8%", "70%", "Slow growth, mature market"),
    ]

    for i, vals in enumerate(growth_data):
        if i == 0:
            for j, val in enumerate(vals, 1):
                ws.cell(row=row, column=j, value=val).style = 'header_style'
        else:
            for j, val in enumerate(vals, 1):
                ws.cell(row=row, column=j, value=val)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    return wb


# =============================================================================
# SHEET 3: FACILITY DATABASE
# =============================================================================
def create_facility_sheet(wb, data):
    """Create Facility Database sheet"""
    ws = wb.create_sheet("3. Facility Database")

    ws['A1'] = "Korea Petrochemical Facility Database (248 Facilities)"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Source: Korea Petrochemical Industry Association (KPIA) 2023"
    ws['A2'].font = Font(italic=True, size=9)

    # Summary statistics
    df_fac = data['facilities']

    ws['A4'] = "Summary Statistics"
    ws['A4'].font = Font(bold=True, size=12)

    summary = [
        ("Total Facilities", len(df_fac)),
        ("Total Capacity", f"{df_fac['capacity_kt'].sum():,.0f} kt/year"),
        ("Locations", ", ".join(df_fac['location'].unique()[:6])),
        ("Products", df_fac['product'].nunique()),
        ("Companies", df_fac['company'].nunique()),
    ]

    row = 6
    for label, value in summary:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Regional Summary
    ws.cell(row=row, column=1, value="Capacity by Region").font = Font(bold=True, size=12)
    row += 2

    regional = df_fac.groupby('location')['capacity_kt'].sum().sort_values(ascending=False).reset_index()
    regional['share_pct'] = regional['capacity_kt'] / regional['capacity_kt'].sum() * 100

    ws.cell(row=row, column=1, value="Location").style = 'header_style'
    ws.cell(row=row, column=2, value="Capacity (kt)").style = 'header_style'
    ws.cell(row=row, column=3, value="Share (%)").style = 'header_style'
    row += 1

    for _, r in regional.head(10).iterrows():
        ws.cell(row=row, column=1, value=r['location'])
        ws.cell(row=row, column=2, value=f"{r['capacity_kt']:,.0f}")
        ws.cell(row=row, column=3, value=f"{r['share_pct']:.1f}%")
        row += 1

    row += 2

    # Product Summary
    ws.cell(row=row, column=1, value="Capacity by Product").font = Font(bold=True, size=12)
    row += 2

    product = df_fac.groupby('product')['capacity_kt'].sum().sort_values(ascending=False).reset_index()

    ws.cell(row=row, column=1, value="Product").style = 'header_style'
    ws.cell(row=row, column=2, value="Capacity (kt)").style = 'header_style'
    row += 1

    for _, r in product.head(15).iterrows():
        ws.cell(row=row, column=1, value=r['product'])
        ws.cell(row=row, column=2, value=f"{r['capacity_kt']:,.0f}")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    return wb


# =============================================================================
# SHEET 4: EMISSION METHODOLOGY
# =============================================================================
def create_emission_methodology_sheet(wb, data):
    """Create Emission Calculation Methodology sheet"""
    ws = wb.create_sheet("4. Emission Methodology")

    ws['A1'] = "Emission Calculation Methodology"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Formula
    ws.cell(row=row, column=1, value="1. Baseline Emission Calculation").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="Formula:")
    ws.cell(row=row, column=2, value="Total_Emissions = Sum (Fuel_Consumption_i x EF_i)").font = Font(italic=True)
    row += 2

    ws.cell(row=row, column=1, value="Where:")
    row += 1
    ws.cell(row=row, column=1, value="  - Fuel_Consumption = Capacity (kt) x Energy_Intensity (GJ/kt or kWh/kt)")
    row += 1
    ws.cell(row=row, column=1, value="  - EF_i = Emission factor for fuel type i (from IPCC/API)")
    row += 1
    ws.cell(row=row, column=1, value="  - Electricity emissions use grid EF trajectory (Korea 10th Power Plan)")
    row += 2

    # Energy Intensities
    ws.cell(row=row, column=1, value="2. Energy Intensities by Product").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    # Get unique product intensities
    intensity_cols = ['product', 'Naphtha_GJ_per_tonne', 'Electricity_kWh_per_tonne',
                      'LNG_GJ_per_tonne', 'Fuel_Gas_GJ_per_tonne']

    intensities = data['energy_intensities'][intensity_cols].drop_duplicates('product').head(10)

    headers = ["Product", "Naphtha (GJ/t)", "Electricity (kWh/t)", "LNG (GJ/t)", "Fuel Gas (GJ/t)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in intensities.iterrows():
        ws.cell(row=row, column=1, value=r['product'])
        ws.cell(row=row, column=2, value=round(r['Naphtha_GJ_per_tonne'], 2) if r['Naphtha_GJ_per_tonne'] > 0 else '-')
        ws.cell(row=row, column=3, value=round(r['Electricity_kWh_per_tonne'], 2))
        ws.cell(row=row, column=4, value=round(r['LNG_GJ_per_tonne'], 2) if r['LNG_GJ_per_tonne'] > 0 else '-')
        ws.cell(row=row, column=5, value=round(r['Fuel_Gas_GJ_per_tonne'], 2) if r['Fuel_Gas_GJ_per_tonne'] > 0 else '-')
        row += 1

    row += 2

    # BAU Projection
    ws.cell(row=row, column=1, value="3. BAU Trajectory Projection").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="Formula:")
    row += 1
    ws.cell(row=row, column=1, value="BAU_Emissions(year) = [Fossil x Capacity_Multiplier x Op_Rate]").font = Font(italic=True)
    row += 1
    ws.cell(row=row, column=1, value="                    + [Elec x Capacity_Multiplier x Op_Rate x Grid_EF(year)/Grid_EF(2025)]").font = Font(italic=True)
    row += 2

    ws.cell(row=row, column=1, value="Key Features:")
    row += 1
    ws.cell(row=row, column=1, value="  - Grid decarbonization: Electricity emissions decline with grid EF improvement")
    row += 1
    ws.cell(row=row, column=1, value="  - Demand growth: Production scales with capacity multiplier (1.0 -> 1.288 by 2050)")
    row += 1
    ws.cell(row=row, column=1, value="  - Operating rate: 70% (reflects 2023 market crisis)")
    row += 2

    # Baseline Results
    ws.cell(row=row, column=1, value="4. Baseline Results (2025, 70% Operating Rate)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    base_results = [
        ("Total Emissions", "46.34 MtCO2/year"),
        ("Fossil Fuel Emissions", "40.47 MtCO2 (87%)"),
        ("Electricity Emissions", "5.86 MtCO2 (13%)"),
    ]

    for label, value in base_results:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    return wb


# =============================================================================
# SHEET 5: BAU TRAJECTORY
# =============================================================================
def create_bau_trajectory_sheet(wb, data):
    """Create BAU Trajectory sheet"""
    ws = wb.create_sheet("5. BAU Trajectory")

    ws['A1'] = "Business-As-Usual (BAU) Emission Trajectory (2025-2050)"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Scenario: No technology deployment; only grid decarbonization benefit"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Write trajectory data
    bau = data['bau_traj'][['year', 'fossil_emissions_mt', 'electricity_emissions_mt',
                            'total_emissions_mt', 'grid_ef_tco2_per_mwh', 'total_capacity_kt']]

    headers = ["Year", "Fossil Emissions (Mt)", "Electricity Emissions (Mt)",
               "Total Emissions (Mt)", "Grid EF (tCO2/MWh)", "Total Capacity (kt)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in bau.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(r['fossil_emissions_mt'], 2))
        ws.cell(row=row, column=3, value=round(r['electricity_emissions_mt'], 2))
        ws.cell(row=row, column=4, value=round(r['total_emissions_mt'], 2))
        ws.cell(row=row, column=5, value=round(r['grid_ef_tco2_per_mwh'], 3))
        ws.cell(row=row, column=6, value=f"{r['total_capacity_kt']:,.0f}")
        row += 1

    # Add chart
    chart = LineChart()
    chart.title = "BAU Emission Trajectory"
    chart.y_axis.title = "Emissions (MtCO2)"
    chart.x_axis.title = "Year"
    chart.style = 10

    data_ref = Reference(ws, min_col=4, min_row=4, max_row=4+len(bau), max_col=4)
    cats = Reference(ws, min_col=1, min_row=5, max_row=4+len(bau))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 15

    ws.add_chart(chart, "H4")

    # Column widths
    for col in 'ABCDEF':
        ws.column_dimensions[col].width = 22

    return wb


# =============================================================================
# SHEET 6: TECHNOLOGY MACC
# =============================================================================
def create_macc_sheet(wb, data):
    """Create Technology MACC sheet"""
    ws = wb.create_sheet("6. Technology MACC")

    ws['A1'] = "Marginal Abatement Cost Curve (MACC) by Technology"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # MACC Formula
    ws.cell(row=row, column=1, value="MACC Formula:").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="MACC ($/tCO2) = [CAPEX_Ann + OPEX_Ann + Fuel_Cost_Diff] / Abatement_Potential").font = Font(italic=True)
    row += 2

    # 2030 MACC Summary
    ws.cell(row=row, column=1, value="MACC Summary (2030)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    macc_2030 = data['macc_annual'][data['macc_annual']['year'] == 2030]

    headers = ["Technology", "Abatement (Mt)", "CAPEX ($/tCO2)", "OPEX ($/tCO2)",
               "Fuel Cost ($/tCO2)", "Total MACC ($/tCO2)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in macc_2030.iterrows():
        ws.cell(row=row, column=1, value=r['technology'])
        ws.cell(row=row, column=2, value=round(r['abatement_potential_mtco2'], 2))
        ws.cell(row=row, column=3, value=round(r['capex_ann_usd_per_tco2'], 0))
        ws.cell(row=row, column=4, value=round(r['opex_ann_usd_per_tco2'], 0))
        ws.cell(row=row, column=5, value=round(r['fuel_cost_diff_usd_per_tco2'], 0))
        ws.cell(row=row, column=6, value=round(r['total_cost_usd_per_tco2'], 0))
        row += 1

    row += 2

    # 2050 MACC Summary
    ws.cell(row=row, column=1, value="MACC Summary (2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    macc_2050 = data['macc_annual'][data['macc_annual']['year'] == 2050]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in macc_2050.iterrows():
        ws.cell(row=row, column=1, value=r['technology'])
        ws.cell(row=row, column=2, value=round(r['abatement_potential_mtco2'], 2))
        ws.cell(row=row, column=3, value=round(r['capex_ann_usd_per_tco2'], 0))
        ws.cell(row=row, column=4, value=round(r['opex_ann_usd_per_tco2'], 0))
        ws.cell(row=row, column=5, value=round(r['fuel_cost_diff_usd_per_tco2'], 0))
        ws.cell(row=row, column=6, value=round(r['total_cost_usd_per_tco2'], 0))
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col in 'BCDEF':
        ws.column_dimensions[col].width = 18

    return wb


# =============================================================================
# SHEET 7: OPTIMIZATION RESULTS
# =============================================================================
def create_optimization_sheet(wb, data):
    """Create Optimization Results sheet"""
    ws = wb.create_sheet("7. Optimization Results")

    ws['A1'] = "Net Zero Transition Pathway: Optimization Results"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Scenario: Cost-Effective Technology Deployment to Net Zero by 2050"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Key Milestones
    ws.cell(row=row, column=1, value="Key Milestones").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt = data['opt_traj']
    milestones = opt[opt['year'].isin([2025, 2030, 2035, 2040, 2045, 2050])]

    headers = ["Year", "BAU (Mt)", "Heat Pump (Mt)", "NCC-Elec (Mt)", "RE PPA (Mt)",
               "Total Abated (Mt)", "Actual Emissions (Mt)", "Cumulative CAPEX ($M)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in milestones.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(r['bau_mt'], 2))
        ws.cell(row=row, column=3, value=round(r['heat_pump_mt'], 2))
        ws.cell(row=row, column=4, value=round(r['ncc_elec_mt'], 2))
        ws.cell(row=row, column=5, value=round(r['re_ppa_mt'], 2))
        ws.cell(row=row, column=6, value=round(r['total_deployed_mt'], 2))
        ws.cell(row=row, column=7, value=round(r['actual_emissions_mt'], 4))
        ws.cell(row=row, column=8, value=f"{r['cumulative_capex_musd']:,.0f}")
        row += 1

    row += 2

    # Annual Trajectory
    ws.cell(row=row, column=1, value="Annual Trajectory").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    headers = ["Year", "BAU (Mt)", "Actual (Mt)", "Elec Demand (TWh)", "Cumulative CAPEX ($M)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in opt.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(r['bau_mt'], 2))
        ws.cell(row=row, column=3, value=round(r['actual_emissions_mt'], 4))
        ws.cell(row=row, column=4, value=round(r['electricity_consumption_increase_twh'], 1))
        ws.cell(row=row, column=5, value=f"{r['cumulative_capex_musd']:,.0f}")
        row += 1

    # Column widths
    for col in 'ABCDEFGH':
        ws.column_dimensions[col].width = 18

    return wb


# =============================================================================
# SHEET 8: ENERGY DEMAND
# =============================================================================
def create_energy_demand_sheet(wb, data):
    """Create Energy Demand sheet"""
    ws = wb.create_sheet("8. Energy Demand")

    ws['A1'] = "Energy Demand Analysis"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Electricity Demand
    ws.cell(row=row, column=1, value="1. Electricity Demand Trajectory").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt = data['opt_traj']

    headers = ["Year", "Baseline Elec (TWh)", "New Elec from Tech (TWh)", "Total Elec (TWh)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    # Baseline electricity is approximately 7.2 TWh/year
    baseline_elec = 7.2

    for _, r in opt[opt['year'].isin([2025, 2030, 2035, 2040, 2045, 2050])].iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(baseline_elec * (1 + (r['year']-2025)*0.01), 1))
        ws.cell(row=row, column=3, value=round(r['electricity_consumption_increase_twh'], 1))
        ws.cell(row=row, column=4, value=round(baseline_elec * (1 + (r['year']-2025)*0.01) + r['electricity_consumption_increase_twh'], 1))
        row += 1

    row += 2

    # Policy Context
    ws.cell(row=row, column=1, value="2. Policy Context").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    policy_data = [
        ("Korea 2036 RE Target", "97 TWh", "10th Basic Plan for Electricity"),
        ("Korea 2050 H2 Target", "27.3 Mt", "Hydrogen Economy Roadmap"),
        ("Model 2050 Elec Demand", f"{opt[opt['year']==2050]['electricity_consumption_increase_twh'].values[0]:.0f} TWh", "NCC-Electricity scenario"),
    ]

    for label, value, source in policy_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=source).font = Font(italic=True, size=9)
        row += 1

    row += 2

    # Feasibility Note
    ws.cell(row=row, column=1, value="3. Feasibility Assessment").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="The NCC-Electricity pathway requires significant grid expansion.")
    row += 1
    ws.cell(row=row, column=1, value=f"2050 additional electricity demand: {opt[opt['year']==2050]['electricity_consumption_increase_twh'].values[0]:.0f} TWh")
    row += 1
    ws.cell(row=row, column=1, value="This represents approximately 40% of Korea's current total electricity generation.")

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20

    return wb


# =============================================================================
# SHEET 9: COST ANALYSIS
# =============================================================================
def create_cost_analysis_sheet(wb, data):
    """Create Cost Analysis sheet"""
    ws = wb.create_sheet("9. Cost Analysis")

    ws['A1'] = "Investment and Cost Analysis"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Cost Summary
    ws.cell(row=row, column=1, value="1. Total Investment Summary").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt = data['opt_traj']
    total_capex = opt['cumulative_capex_musd'].max() / 1000  # Billion

    summary = [
        ("Total CAPEX (2025-2050)", f"${total_capex:.1f} Billion"),
        ("Average Annual Investment", f"${total_capex/26:.2f} Billion/year"),
        ("Peak Investment Period", "2035-2040"),
    ]

    for label, value in summary:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Annual CAPEX
    ws.cell(row=row, column=1, value="2. Annual CAPEX Trajectory").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt_copy = opt.copy()
    opt_copy['annual_capex'] = opt_copy['cumulative_capex_musd'].diff().fillna(0)

    headers = ["Year", "Annual CAPEX ($M)", "Cumulative CAPEX ($M)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in opt_copy.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=f"{r['annual_capex']:,.0f}")
        ws.cell(row=row, column=3, value=f"{r['cumulative_capex_musd']:,.0f}")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25

    return wb


# =============================================================================
# SHEET 10: REFERENCES
# =============================================================================
def create_references_sheet(wb):
    """Create References sheet"""
    ws = wb.create_sheet("10. References")

    ws['A1'] = "References and Data Sources"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Academic References
    ws.cell(row=row, column=1, value="Academic and Industry References").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    references = [
        ("Emission Factors", "", ""),
        ("", "IPCC (2019)", "2019 Refinement to the 2006 IPCC Guidelines for National Greenhouse Gas Inventories"),
        ("", "API (2021)", "Compendium of Greenhouse Gas Emissions Methodologies for the Oil and Natural Gas Industry"),
        ("", "", ""),
        ("Technology - NCC Electric", "", ""),
        ("", "BASF/SABIC/Linde (2024)", "Commercial pilot plant data (6 MW, 4 ton/hr naphtha)"),
        ("", "Toribio-Ramirez et al. (2025)", "Electric cracker CAPEX analysis"),
        ("", "Tijani et al. (2022)", "Electric cracking technical review"),
        ("", "", ""),
        ("Technology - NCC H2", "", ""),
        ("", "Lummus Tech (2023)", "Engineering case study for 1,000 kt/yr hydrogen cracker"),
        ("", "Thunder Said Energy (2023)", "Hydrogen cracker economics"),
        ("", "ExxonMobil", "98% H2 operation validation"),
        ("", "", ""),
        ("Technology - Heat Pump", "", ""),
        ("", "Kosmadakis et al. (2020)", "Heat pump for industrial processes (COP 4.0)"),
        ("", "McKinsey (2024)", "Industrial heat pump cost analysis"),
        ("", "", ""),
        ("Technology - RDH", "", ""),
        ("", "Coolbrook (2024)", "RotoDynamic Heater for industrial heating (TRL 8)"),
        ("", "", ""),
        ("Grid & Policy", "", ""),
        ("", "Korea 10th Basic Plan (2023)", "Electricity Supply and Demand Plan"),
        ("", "KPIA (2023)", "Korea Petrochemical Industry Association capacity database"),
        ("", "", ""),
        ("Price Assumptions", "", ""),
        ("", "IRENA (2024)", "Global renewable energy cost review"),
        ("", "PLANiT Institute (2025)", "Custom H2 and RE price trajectories"),
    ]

    for cat, ref, desc in references:
        if cat:
            ws.cell(row=row, column=1, value=cat).font = Font(bold=True)
        if ref:
            ws.cell(row=row, column=2, value=ref).font = Font(italic=True)
            ws.cell(row=row, column=3, value=desc)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 70

    return wb


# =============================================================================
# MAIN
# =============================================================================
def generate_report():
    """Generate the complete Excel report"""
    print("Loading data...")
    data = load_all_data()

    print("Creating workbook...")
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create styles
    wb = create_styles(wb)

    # Create sheets
    print("Creating Executive Summary...")
    wb = create_executive_summary(wb, data)

    print("Creating Assumptions & References...")
    wb = create_assumptions_sheet(wb, data)

    print("Creating Facility Database...")
    wb = create_facility_sheet(wb, data)

    print("Creating Emission Methodology...")
    wb = create_emission_methodology_sheet(wb, data)

    print("Creating BAU Trajectory...")
    wb = create_bau_trajectory_sheet(wb, data)

    print("Creating Technology MACC...")
    wb = create_macc_sheet(wb, data)

    print("Creating Optimization Results...")
    wb = create_optimization_sheet(wb, data)

    print("Creating Energy Demand...")
    wb = create_energy_demand_sheet(wb, data)

    print("Creating Cost Analysis...")
    wb = create_cost_analysis_sheet(wb, data)

    print("Creating References...")
    wb = create_references_sheet(wb)

    print("Creating Regional Analysis...")
    wb = create_regional_analysis_sheet(wb, data)

    print("Creating Operating Rate Analysis...")
    wb = create_operating_rate_sheet(wb, data)

    # Save
    output_path = OUTPUT_DIR / "Korea_Petrochemical_NetZero_Report.xlsx"
    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")

    return output_path


# =============================================================================
# SHEET 11: REGIONAL ANALYSIS (NEW)
# =============================================================================
def create_regional_analysis_sheet(wb, data):
    """Create comprehensive Regional Analysis sheet"""
    ws = wb.create_sheet("11. Regional Analysis")
    
    ws['A1'] = "Regional Analysis: Korea Petrochemical Complexes"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Facility-level breakdown by location with emissions and investment allocation"
    ws['A2'].font = Font(italic=True, size=9)
    
    row = 4
    
    # Load and process facility data
    fac = data['facilities']
    energy = data['energy_intensities']
    
    # Merge to get emissions data
    merged = pd.merge(fac, energy[['product', 'process', 'company', 'location', 
                                   'Naphtha_GJ_per_tonne', 'Electricity_kWh_per_tonne',
                                   'LNG_GJ_per_tonne', 'Fuel_Gas_GJ_per_tonne']], 
                      on=['product', 'process', 'company', 'location'], how='left')
    
    # Emission factors
    EF_NAPHTHA = 0.0542
    EF_LNG = 0.0561
    EF_FUEL_GAS = 0.050
    GRID_EF_2025 = 0.436
    
    # Calculate emissions
    merged['emissions_fossil_kt'] = (
        merged['capacity_kt'] * merged['Naphtha_GJ_per_tonne'].fillna(0) * EF_NAPHTHA +
        merged['capacity_kt'] * merged['LNG_GJ_per_tonne'].fillna(0) * EF_LNG +
        merged['capacity_kt'] * merged['Fuel_Gas_GJ_per_tonne'].fillna(0) * EF_FUEL_GAS
    )
    merged['emissions_elec_kt'] = merged['capacity_kt'] * merged['Electricity_kWh_per_tonne'].fillna(0) / 1000 * GRID_EF_2025
    merged['emissions_total_kt'] = merged['emissions_fossil_kt'] + merged['emissions_elec_kt']
    
    # Section 1: Regional Summary
    ws.cell(row=row, column=1, value="1. Regional Capacity and Emissions Summary").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    # Aggregate by region
    regional = merged.groupby('location').agg({
        'capacity_kt': 'sum',
        'emissions_fossil_kt': 'sum',
        'emissions_elec_kt': 'sum',
        'emissions_total_kt': 'sum',
        'product': 'count'
    }).rename(columns={'product': 'n_facilities'}).reset_index()
    
    regional['share_pct'] = regional['capacity_kt'] / regional['capacity_kt'].sum() * 100
    regional['emissions_mt'] = regional['emissions_total_kt'] / 1000
    regional = regional.sort_values('capacity_kt', ascending=False)
    
    # Calculate investment allocation (proportional to emissions)
    total_investment = data['opt_traj']['cumulative_capex_musd'].max() / 1000  # Billion
    total_emissions = regional['emissions_mt'].sum()
    regional['investment_b'] = (regional['emissions_mt'] / total_emissions) * total_investment
    
    headers = ["Region", "Facilities", "Capacity (kt)", "Emissions (Mt)", "Share (%)", "Est. Investment ($B)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1
    
    for _, r in regional.iterrows():
        ws.cell(row=row, column=1, value=r['location'])
        ws.cell(row=row, column=2, value=int(r['n_facilities']))
        ws.cell(row=row, column=3, value=f"{r['capacity_kt']:,.0f}")
        ws.cell(row=row, column=4, value=round(r['emissions_mt'], 2))
        ws.cell(row=row, column=5, value=f"{r['share_pct']:.1f}%")
        ws.cell(row=row, column=6, value=f"${r['investment_b']:.2f}")
        row += 1
    
    row += 2
    
    # Section 2: Top 4 Complexes Detailed Analysis
    ws.cell(row=row, column=1, value="2. Major Complex Analysis (Top 4)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    top_regions = ['Yeosu', 'Daesan', 'Ulsan', 'Onsan']
    
    for region in top_regions:
        loc_data = merged[merged['location'] == region]
        total_cap = loc_data['capacity_kt'].sum()
        total_emit = loc_data['emissions_total_kt'].sum() / 1000
        n_fac = len(loc_data)
        
        ws.cell(row=row, column=1, value=f"{region} Complex").font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        ws.cell(row=row, column=1, value=f"Facilities: {n_fac} | Capacity: {total_cap:,.0f} kt | Emissions: {total_emit:.2f} Mt")
        row += 2
        
        # Product mix
        ws.cell(row=row, column=1, value="Product").style = 'header_style'
        ws.cell(row=row, column=2, value="Capacity (kt)").style = 'header_style'
        ws.cell(row=row, column=3, value="Share (%)").style = 'header_style'
        row += 1
        
        product_mix = loc_data.groupby('product')['capacity_kt'].sum().sort_values(ascending=False).head(5)
        for prod, cap in product_mix.items():
            ws.cell(row=row, column=1, value=prod)
            ws.cell(row=row, column=2, value=f"{cap:,.0f}")
            ws.cell(row=row, column=3, value=f"{cap/total_cap*100:.1f}%")
            row += 1
        
        row += 2
    
    # Section 3: Regional Transition Implications
    ws.cell(row=row, column=1, value="3. Regional Transition Implications").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    implications = [
        ("Yeosu Complex", "Largest NCC cluster (5.4 Mt ethylene)", "Priority for NCC-Electricity deployment", "$8.7B investment"),
        ("Daesan Complex", "High P-X and polymer production", "Heat pump + NCC-Electricity focus", "$7.2B investment"),
        ("Ulsan Complex", "Diverse product mix, older facilities", "Phased technology deployment", "$3.2B investment"),
        ("Onsan Complex", "Concentrated NCC operations", "Efficient technology clustering", "$2.1B investment"),
    ]
    
    ws.cell(row=row, column=1, value="Region").style = 'header_style'
    ws.cell(row=row, column=2, value="Characteristics").style = 'header_style'
    ws.cell(row=row, column=3, value="Transition Strategy").style = 'header_style'
    ws.cell(row=row, column=4, value="Est. Investment").style = 'header_style'
    row += 1
    
    for region, char, strategy, invest in implications:
        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value=char)
        ws.cell(row=row, column=3, value=strategy)
        ws.cell(row=row, column=4, value=invest)
        row += 1
    
    row += 2
    
    # Section 4: Regional Grid Infrastructure Requirements
    ws.cell(row=row, column=1, value="4. Regional Grid Infrastructure Requirements (2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    # Estimate electricity demand by region proportional to NCC capacity
    ncc_by_region = merged[merged['process'] == 'Naphtha Cracker'].groupby('location')['capacity_kt'].sum()
    total_ncc = ncc_by_region.sum()
    total_elec_demand = data['opt_traj'][data['opt_traj']['year'] == 2050]['electricity_consumption_increase_twh'].values[0]
    
    ws.cell(row=row, column=1, value="Region").style = 'header_style'
    ws.cell(row=row, column=2, value="NCC Capacity (kt)").style = 'header_style'
    ws.cell(row=row, column=3, value="Est. Elec Demand (TWh)").style = 'header_style'
    ws.cell(row=row, column=4, value="Grid Expansion Need").style = 'header_style'
    row += 1
    
    for region in top_regions:
        if region in ncc_by_region.index:
            ncc_cap = ncc_by_region[region]
            elec_demand = (ncc_cap / total_ncc) * total_elec_demand
            grid_note = "Major upgrade required" if elec_demand > 50 else "Moderate upgrade" if elec_demand > 20 else "Minor upgrade"
            
            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value=f"{ncc_cap:,.0f}")
            ws.cell(row=row, column=3, value=f"{elec_demand:.1f}")
            ws.cell(row=row, column=4, value=grid_note)
            row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    
    return wb


# =============================================================================
# SHEET 12: OPERATING RATE ANALYSIS (NEW)
# =============================================================================
def create_operating_rate_sheet(wb, data):
    """Create Operating Rate Analysis sheet"""
    ws = wb.create_sheet("12. Operating Rate Analysis")
    
    ws['A1'] = "Operating Rate Analysis and Model Assumptions"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    
    row = 3
    
    # Section 1: Operating Rate Definition
    ws.cell(row=row, column=1, value="1. Operating Rate Definition").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    definitions = [
        ("Operating Rate", "Actual production / Installed capacity", "70%"),
        ("Capacity Multiplier", "Installed capacity growth from 2025 baseline", "1.0 -> 1.288"),
        ("Effective Multiplier", "Capacity Multiplier x Operating Rate", "Production scaling factor"),
    ]
    
    ws.cell(row=row, column=1, value="Parameter").style = 'header_style'
    ws.cell(row=row, column=2, value="Definition").style = 'header_style'
    ws.cell(row=row, column=3, value="Value").style = 'header_style'
    row += 1
    
    for param, defn, val in definitions:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=defn)
        ws.cell(row=row, column=3, value=val)
        row += 1
    
    row += 2
    
    # Section 2: Operating Rate in Model
    ws.cell(row=row, column=1, value="2. How Operating Rate is Used in the Model").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    ws.cell(row=row, column=1, value="Formula:")
    row += 1
    ws.cell(row=row, column=1, value="Effective_Multiplier = Capacity_Multiplier x Operating_Rate").font = Font(italic=True)
    row += 1
    ws.cell(row=row, column=1, value="Actual_Emissions = Baseline_Emissions x Effective_Multiplier").font = Font(italic=True)
    row += 2
    
    ws.cell(row=row, column=1, value="Implementation in Code (modules/baseline.py, modules/macc.py):")
    row += 1
    ws.cell(row=row, column=1, value="  1. Read operating_rate_pct from demand_growth_trajectory.csv")
    row += 1
    ws.cell(row=row, column=1, value="  2. Calculate effective_multiplier = capacity_multiplier * (operating_rate_pct / 100)")
    row += 1
    ws.cell(row=row, column=1, value="  3. Scale all emissions by effective_multiplier")
    row += 2
    
    # Section 3: Operating Rate Trajectory
    ws.cell(row=row, column=1, value="3. Operating Rate Trajectory (2025-2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    demand = data['demand_growth']
    
    ws.cell(row=row, column=1, value="Year").style = 'header_style'
    ws.cell(row=row, column=2, value="Annual Growth (%)").style = 'header_style'
    ws.cell(row=row, column=3, value="Capacity Multiplier").style = 'header_style'
    ws.cell(row=row, column=4, value="Operating Rate (%)").style = 'header_style'
    ws.cell(row=row, column=5, value="Effective Multiplier").style = 'header_style'
    row += 1
    
    for _, r in demand.iterrows():
        op_rate = r['operating_rate_pct']
        eff_mult = r['cumulative_capacity_multiplier'] * (op_rate / 100)
        
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=f"{r['annual_growth_rate_pct']:.1f}%")
        ws.cell(row=row, column=3, value=round(r['cumulative_capacity_multiplier'], 3))
        ws.cell(row=row, column=4, value=f"{op_rate:.0f}%")
        ws.cell(row=row, column=5, value=round(eff_mult, 3))
        row += 1
    
    row += 2
    
    # Section 4: Impact of Operating Rate
    ws.cell(row=row, column=1, value="4. Impact of Operating Rate on Results").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    impacts = [
        ("Baseline Emissions (2025)", "At 100% op rate: 66.2 MtCO2", "At 70% op rate: 46.3 MtCO2"),
        ("BAU Emissions (2050)", "At 100% op rate: ~76 MtCO2", "At 70% op rate: 53.3 MtCO2"),
        ("Total Investment Need", "At 100% op rate: ~$53B", "At 70% op rate: $37.3B"),
        ("Technology Sizing", "Scales with actual production", "30% smaller than nameplate"),
    ]
    
    ws.cell(row=row, column=1, value="Metric").style = 'header_style'
    ws.cell(row=row, column=2, value="100% Operating Rate").style = 'header_style'
    ws.cell(row=row, column=3, value="70% Operating Rate (Model)").style = 'header_style'
    row += 1
    
    for metric, val_100, val_70 in impacts:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=val_100)
        ws.cell(row=row, column=3, value=val_70)
        row += 1
    
    row += 2
    
    # Section 5: Rationale
    ws.cell(row=row, column=1, value="5. Rationale for 70% Operating Rate").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    rationale = [
        "- Korean petrochemical industry experienced significant overcapacity crisis in 2023-2024",
        "- Chinese petrochemical expansion led to global supply glut",
        "- KPIA data shows actual operating rates dropped to 65-75% in 2023",
        "- 70% represents conservative assumption for 2025-2050 period",
        "- Model can be adjusted for sensitivity analysis (80%, 90%, 100% scenarios)",
    ]
    
    for text in rationale:
        ws.cell(row=row, column=1, value=text)
        row += 1
    
    row += 2
    
    # Section 6: Sensitivity Note
    ws.cell(row=row, column=1, value="6. Sensitivity Analysis").font = Font(bold=True, size=12, color='1F4E79')
    row += 2
    
    ws.cell(row=row, column=1, value="Operating rate significantly impacts all model outputs:")
    row += 1
    ws.cell(row=row, column=1, value="  - Higher operating rate = higher emissions = higher abatement need = higher investment")
    row += 1
    ws.cell(row=row, column=1, value="  - Lower operating rate = lower emissions = lower technology deployment = lower cost")
    row += 1
    ws.cell(row=row, column=1, value="  - To run sensitivity, modify 'operating_rate_pct' in data/demand_growth_trajectory.csv")
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    return wb


if __name__ == "__main__":
    generate_report()

