"""
Client-Ready Excel Report Generator
Comprehensive analysis with NCC-Electricity vs NCC-H2 comparison
All assumptions with literature references
Updated: Dec 2024
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, BarChart, Reference

OUTPUT_DIR = Path('outputs/client_report')
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

DATA_DIR = Path('data')
SCENARIO_DIR = Path('outputs/new_scenarios/cost_effective')
BAU_DIR = Path('outputs/new_scenarios/bau')

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=12, color="2E75B6")
REF_FONT = Font(italic=True, size=9, color="666666")
HIGHLIGHT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))


def apply_header_style(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def load_all_data():
    """Load all necessary data"""
    data = {}
    
    # Load trajectories
    try:
        data['bau_traj'] = pd.read_csv(BAU_DIR / 'module_01/bau_trajectory_2025_2050.csv')
        data['opt_traj'] = pd.read_csv(SCENARIO_DIR / 'module_03/optimization_trajectory.csv')
        data['macc_annual'] = pd.read_csv(SCENARIO_DIR / 'module_02/macc_annual_2025_2050.csv')
    except Exception as e:
        print(f"Warning: Could not load scenario data: {e}")
        data['bau_traj'] = None
        data['opt_traj'] = None
        data['macc_annual'] = None
    
    # Load static data
    data['facilities'] = pd.read_csv(DATA_DIR / "facility_database_with_regions.csv")
    data['tech_params'] = pd.read_csv(DATA_DIR / "technology_parameters.csv")
    data['emission_factors'] = pd.read_csv(DATA_DIR / "emission_factors.csv")
    data['re_prices'] = pd.read_csv(DATA_DIR / "re_price_trajectory.csv")
    data['h2_prices'] = pd.read_csv(DATA_DIR / "h2_price_trajectory.csv")
    data['baseline'] = pd.read_csv(SCENARIO_DIR / 'module_01/baseline_2025_detailed.csv')
    
    return data


def create_assumptions_sheet(wb, data):
    """Create comprehensive assumptions sheet with all references"""
    ws = wb.create_sheet("Assumptions", 0)
    
    ws['B2'] = "Model Assumptions and Data Sources"
    ws['B2'].font = TITLE_FONT
    ws['B3'] = "All costs in USD (Real 2024)"
    ws['B3'].font = Font(italic=True, color="7F7F7F")
    
    row = 5
    
    # 1. Technology CAPEX
    ws.cell(row=row, column=2, value="1. TECHNOLOGY CAPEX ($/tCO2 or $/t-C2H4)").font = SUBTITLE_FONT
    row += 1
    
    headers = ["Technology", "2025", "2030", "2040", "2050", "Unit", "Source"]
    apply_header_style(ws, row, 2, 8)
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2+i, value=h)
    row += 1
    
    capex_data = [
        ("NCC-Electricity", 1500, 1350, 1050, 900, "$/t-C2H4/yr", "Toribio-Ramirez et al. 2025, BASF/SABIC/Linde 2024"),
        ("NCC-H2", 1700, 1300, 935, 780, "$/t-C2H4/yr", "Thunder Said Energy 2023, Lummus Tech 2023"),
        ("Heat Pump", 800, 640, 480, 400, "$/tCO2", "McKinsey 2024, Kosmadakis et al. 2020"),
        ("RDH", 900, 720, 540, 450, "$/tCO2", "Coolbrook 2024"),
    ]
    
    for tech, c25, c30, c40, c50, unit, source in capex_data:
        ws.cell(row=row, column=2, value=tech).border = THIN_BORDER
        ws.cell(row=row, column=3, value=c25).border = THIN_BORDER
        ws.cell(row=row, column=4, value=c30).border = THIN_BORDER
        ws.cell(row=row, column=5, value=c40).border = THIN_BORDER
        ws.cell(row=row, column=6, value=c50).border = THIN_BORDER
        ws.cell(row=row, column=7, value=unit).border = THIN_BORDER
        ws.cell(row=row, column=8, value=source).font = REF_FONT
        ws.cell(row=row, column=8).border = THIN_BORDER
        row += 1
    
    row += 1
    
    # 2. Emission Factors
    ws.cell(row=row, column=2, value="2. EMISSION FACTORS").font = SUBTITLE_FONT
    row += 1
    
    headers = ["Fuel Type", "EF (tCO2/GJ)", "Source"]
    apply_header_style(ws, row, 2, 4)
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2+i, value=h)
    row += 1
    
    for _, ef_row in data['emission_factors'].iterrows():
        if pd.notna(ef_row.get('tCO2_per_GJ')):
            ws.cell(row=row, column=2, value=ef_row['fuel']).border = THIN_BORDER
            ws.cell(row=row, column=3, value=ef_row['tCO2_per_GJ']).border = THIN_BORDER
            ws.cell(row=row, column=4, value=ef_row['source']).font = REF_FONT
            ws.cell(row=row, column=4).border = THIN_BORDER
            row += 1
    
    row += 1
    
    # 3. Energy Intensity
    ws.cell(row=row, column=2, value="3. ENERGY INTENSITY").font = SUBTITLE_FONT
    row += 1
    
    headers = ["Process", "Parameter", "Value", "Unit", "Source"]
    apply_header_style(ws, row, 2, 6)
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2+i, value=h)
    row += 1
    
    intensity_data = [
        ("Naphtha Cracker", "Heat Input", 11.0, "GJ/t-C2H4", "Industry benchmark"),
        ("Naphtha Cracker", "Electricity", 5.0, "MWh/t-C2H4", "BASF/SABIC/Linde 2024 Demo"),
        ("Naphtha Cracker", "H2 Consumption", 0.2, "t-H2/t-C2H4", "Lummus Tech 2023"),
        ("BTX Plant", "Heat Input", 8.5, "GJ/t-BTX", "Industry benchmark"),
        ("Heat Pump", "COP", 4.0, "-", "Kosmadakis et al. 2020"),
        ("RDH", "Efficiency", 93, "%", "Coolbrook 2024"),
    ]
    
    for process, param, value, unit, source in intensity_data:
        ws.cell(row=row, column=2, value=process).border = THIN_BORDER
        ws.cell(row=row, column=3, value=param).border = THIN_BORDER
        ws.cell(row=row, column=4, value=value).border = THIN_BORDER
        ws.cell(row=row, column=5, value=unit).border = THIN_BORDER
        ws.cell(row=row, column=6, value=source).font = REF_FONT
        ws.cell(row=row, column=6).border = THIN_BORDER
        row += 1
    
    row += 1
    
    # 4. Electricity Price Trajectory
    ws.cell(row=row, column=2, value="4. ELECTRICITY PRICE TRAJECTORY (RE-PPA)").font = SUBTITLE_FONT
    row += 1
    
    years = [2025, 2030, 2035, 2040, 2045, 2050]
    headers = ["Year"] + [str(y) for y in years]
    apply_header_style(ws, row, 2, 2+len(years))
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2+i, value=h)
    row += 1
    
    ws.cell(row=row, column=2, value="RE Price ($/MWh)").border = THIN_BORDER
    for i, y in enumerate(years):
        price = data['re_prices'][data['re_prices']['year'] == y]['re_price_usd_per_mwh'].values[0]
        ws.cell(row=row, column=3+i, value=round(price, 1)).border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=2, value="Source: PLANiT Institute (2025)").font = REF_FONT
    row += 2
    
    # 5. Hydrogen Price Trajectory (LCOH)
    ws.cell(row=row, column=2, value="5. HYDROGEN PRICE TRAJECTORY (LCOH)").font = SUBTITLE_FONT
    row += 1
    
    headers = ["Year"] + [str(y) for y in years]
    apply_header_style(ws, row, 2, 2+len(years))
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2+i, value=h)
    row += 1
    
    ws.cell(row=row, column=2, value="H2 Price ($/kg)").border = THIN_BORDER
    for i, y in enumerate(years):
        price = data['h2_prices'][data['h2_prices']['year'] == y]['h2_price_usd_per_kg'].values[0]
        ws.cell(row=row, column=3+i, value=round(price, 2)).border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=2, value="Source: PLANiT Institute LCOH Model (2025)").font = REF_FONT
    row += 2
    
    # 6. LCOH Assumptions
    ws.cell(row=row, column=2, value="6. LCOH CALCULATION ASSUMPTIONS").font = SUBTITLE_FONT
    row += 1
    
    lcoh_params = [
        ("Electrolyzer Efficiency", "70%", "HHV basis"),
        ("Electrolyzer CAPEX (2025)", "$1,200/kW", "IEA 2024"),
        ("Electrolyzer CAPEX (2050)", "$300/kW", "Learning rate projection"),
        ("Capacity Factor", "90%", "Dedicated RE"),
        ("Stack Lifetime", "10 years", "PEM technology"),
        ("Discount Rate", "8%", "WACC for industrial projects"),
        ("System Lifetime", "20 years", ""),
    ]
    
    headers = ["Parameter", "Value", "Source/Notes"]
    apply_header_style(ws, row, 2, 4)
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2+i, value=h)
    row += 1
    
    for param, value, source in lcoh_params:
        ws.cell(row=row, column=2, value=param).border = THIN_BORDER
        ws.cell(row=row, column=3, value=value).border = THIN_BORDER
        ws.cell(row=row, column=4, value=source).font = REF_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1
    
    # Column widths
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 45
    
    return ws


def create_scenario_comparison(wb, data):
    """Create NCC-Electricity vs NCC-H2 comparison sheet"""
    ws = wb.create_sheet("Scenario_Comparison", 1)
    
    ws['B2'] = "Scenario Comparison: NCC-Electricity vs NCC-H2"
    ws['B2'].font = TITLE_FONT
    
    # Calculate costs for both scenarios
    facilities = data['facilities']
    ncc_capacity_kt = facilities[facilities['process'] == 'Naphtha Cracker']['capacity_kt'].sum()
    
    years = [2030, 2035, 2040, 2045, 2050]
    
    # NCC-Electricity scenario
    elec_mwh_per_ton = 5.0
    elec_demand_mwh = ncc_capacity_kt * 1000 * elec_mwh_per_ton * 0.7  # 70% utilization
    
    # NCC-H2 scenario
    h2_ton_per_ton = 0.2
    h2_demand_ton = ncc_capacity_kt * 1000 * h2_ton_per_ton * 0.7
    
    row = 5
    ws.cell(row=row, column=2, value="Key Metrics by Scenario").font = SUBTITLE_FONT
    row += 1
    
    headers = ["Metric", "NCC-Electricity", "NCC-H2", "Difference"]
    apply_header_style(ws, row, 2, 5)
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2+i, value=h)
    row += 1
    
    # Get 2050 prices
    re_2050 = data['re_prices'][data['re_prices']['year'] == 2050]['re_price_usd_per_mwh'].values[0]
    h2_2050 = data['h2_prices'][data['h2_prices']['year'] == 2050]['h2_price_usd_per_kg'].values[0]
    
    # Annual energy cost (2050)
    elec_cost = elec_demand_mwh * re_2050 / 1e9  # $B
    h2_cost = h2_demand_ton * h2_2050 * 1000 / 1e9  # $B (H2 in $/kg * 1000 kg/ton)
    
    metrics = [
        ("NCC Capacity (kt)", ncc_capacity_kt, ncc_capacity_kt, 0),
        ("Energy Consumption", f"{elec_demand_mwh/1e6:.1f} TWh", f"{h2_demand_ton/1e6:.2f} Mt H2", "-"),
        ("Energy Price (2050)", f"${re_2050:.0f}/MWh", f"${h2_2050:.2f}/kg", "-"),
        ("Annual Energy Cost (2050)", f"${elec_cost:.2f}B", f"${h2_cost:.2f}B", f"${h2_cost-elec_cost:+.2f}B"),
        ("CAPEX ($/t-C2H4)", "$900", "$780", "-$120"),
        ("Energy Efficiency", "95%", "85%", "-10%"),
        ("Infrastructure", "Grid upgrade", "H2 pipeline", "-"),
    ]
    
    for metric, elec, h2, diff in metrics:
        ws.cell(row=row, column=2, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=3, value=elec).border = THIN_BORDER
        ws.cell(row=row, column=4, value=h2).border = THIN_BORDER
        ws.cell(row=row, column=5, value=diff).border = THIN_BORDER
        row += 1
    
    row += 2
    
    # Annual trajectory comparison
    ws.cell(row=row, column=2, value="Annual Cost Trajectory Comparison").font = SUBTITLE_FONT
    row += 1
    
    headers = ["Year", "RE Price ($/MWh)", "H2 Price ($/kg)", "Elec Cost ($B)", "H2 Cost ($B)", "Savings (+H2)"]
    apply_header_style(ws, row, 2, 7)
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2+i, value=h)
    row += 1
    
    for y in years:
        re_p = data['re_prices'][data['re_prices']['year'] == y]['re_price_usd_per_mwh'].values[0]
        h2_p = data['h2_prices'][data['h2_prices']['year'] == y]['h2_price_usd_per_kg'].values[0]
        
        elec_c = elec_demand_mwh * re_p / 1e9
        h2_c = h2_demand_ton * h2_p * 1000 / 1e9
        
        ws.cell(row=row, column=2, value=y).border = THIN_BORDER
        ws.cell(row=row, column=3, value=round(re_p, 1)).border = THIN_BORDER
        ws.cell(row=row, column=4, value=round(h2_p, 2)).border = THIN_BORDER
        ws.cell(row=row, column=5, value=round(elec_c, 2)).border = THIN_BORDER
        ws.cell(row=row, column=6, value=round(h2_c, 2)).border = THIN_BORDER
        ws.cell(row=row, column=7, value=round(elec_c - h2_c, 2)).border = THIN_BORDER
        row += 1
    
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    
    return ws


def create_facility_transition(wb, data):
    """Create facility transition schedule"""
    ws = wb.create_sheet("Facility_Transition", 2)
    
    ws['B2'] = "Facility-Level Transition Schedule"
    ws['B2'].font = TITLE_FONT
    
    df_fac = data['facilities'].copy()
    df_base = data['baseline'].copy()
    
    # Create keys for matching
    df_base['key'] = df_base['product'] + '|' + df_base['company'] + '|' + df_base['location']
    df_fac['key'] = df_fac['product'] + '|' + df_fac['company'] + '|' + df_fac['location']
    
    emissions_map = df_base.groupby('key')['total_emissions_kt'].sum().to_dict()
    df_fac['total_emissions_kt'] = df_fac['key'].map(emissions_map).fillna(0)
    
    def assign_tech(row):
        if row['process'] == 'Naphtha Cracker':
            return 'NCC-Electricity', 2030, 2040
        elif row['process'] == 'BTX Plant':
            return 'RDH', 2030, 2045
        else:
            return 'Heat Pump', 2025, 2035
    
    df_fac[['technology', 'start_year', 'complete_year']] = df_fac.apply(
        lambda x: pd.Series(assign_tech(x)), axis=1)
    
    df_fac = df_fac.reset_index(drop=True)
    df_fac['facility_id'] = ['FAC-' + str(i+1).zfill(3) for i in df_fac.index]
    
    capex_rates = {'NCC-Electricity': 900, 'Heat Pump': 640, 'RDH': 720}
    df_fac['investment_musd'] = df_fac.apply(
        lambda x: x['total_emissions_kt'] * capex_rates.get(x['technology'], 800) / 1000, axis=1)
    
    df_fac['elec_demand_gwh'] = df_fac.apply(lambda x: 
        x['capacity_kt'] * 5 if x['technology'] == 'NCC-Electricity' else
        x['capacity_kt'] * 2 if x['technology'] == 'Heat Pump' else
        x['capacity_kt'] * 3, axis=1)
    
    # Headers
    headers = ["ID", "Company", "Location", "Process", "Product", "Capacity (kt)", 
               "Emissions (kt)", "Technology", "Start", "Complete", "Invest (M$)", "Elec (GWh)"]
    apply_header_style(ws, 4, 2, 13)
    for i, h in enumerate(headers):
        ws.cell(row=4, column=2+i, value=h)
    
    df_fac = df_fac.sort_values(['location', 'total_emissions_kt'], ascending=[True, False])
    
    for idx, row in df_fac.iterrows():
        r = 5 + idx
        ws.cell(row=r, column=2, value=row['facility_id']).border = THIN_BORDER
        ws.cell(row=r, column=3, value=row['company']).border = THIN_BORDER
        ws.cell(row=r, column=4, value=row['location']).border = THIN_BORDER
        ws.cell(row=r, column=5, value=row['process']).border = THIN_BORDER
        ws.cell(row=r, column=6, value=row['product']).border = THIN_BORDER
        ws.cell(row=r, column=7, value=round(row['capacity_kt'], 1)).border = THIN_BORDER
        ws.cell(row=r, column=8, value=round(row['total_emissions_kt'], 1)).border = THIN_BORDER
        ws.cell(row=r, column=9, value=row['technology']).border = THIN_BORDER
        ws.cell(row=r, column=10, value=row['start_year']).border = THIN_BORDER
        ws.cell(row=r, column=11, value=row['complete_year']).border = THIN_BORDER
        ws.cell(row=r, column=12, value=round(row['investment_musd'], 1)).border = THIN_BORDER
        ws.cell(row=r, column=13, value=round(row['elec_demand_gwh'], 1)).border = THIN_BORDER

    ws.auto_filter.ref = f"B4:M{4+len(df_fac)}"
    ws.freeze_panes = 'B5'
    
    for col in range(2, 14):
        ws.column_dimensions[chr(64+col)].width = 12
    ws.column_dimensions['C'].width = 20
    
    return ws


def create_regional_summary(wb, data):
    """Create regional cost and energy summary"""
    ws = wb.create_sheet("Regional_Summary", 3)
    
    ws['B2'] = "Regional Investment and Energy Demand"
    ws['B2'].font = TITLE_FONT
    
    df_fac = data['facilities'].copy()
    df_base = data['baseline'].copy()
    
    df_base['key'] = df_base['product'] + '|' + df_base['company'] + '|' + df_base['location']
    df_fac['key'] = df_fac['product'] + '|' + df_fac['company'] + '|' + df_fac['location']
    
    emissions_map = df_base.groupby('key')['total_emissions_kt'].sum().to_dict()
    df_fac['total_emissions_kt'] = df_fac['key'].map(emissions_map).fillna(0)
    
    regional = df_fac.groupby('location').agg({
        'product': 'count',
        'capacity_kt': 'sum',
        'total_emissions_kt': 'sum'
    }).rename(columns={'product': 'num_facilities'})
    
    regional['investment_musd'] = regional['total_emissions_kt'] * 850 / 1000
    regional['elec_demand_twh'] = regional['capacity_kt'] * 5 / 1000
    regional = regional.sort_values('total_emissions_kt', ascending=False)
    
    headers = ["Region", "Facilities", "Capacity (kt)", "Emissions (kt)", "Investment (M$)", "Elec (TWh)"]
    apply_header_style(ws, 4, 2, 7)
    for i, h in enumerate(headers):
        ws.cell(row=4, column=2+i, value=h)
    
    row = 5
    for region, rd in regional.iterrows():
        ws.cell(row=row, column=2, value=region).border = THIN_BORDER
        ws.cell(row=row, column=3, value=int(rd['num_facilities'])).border = THIN_BORDER
        ws.cell(row=row, column=4, value=round(rd['capacity_kt'], 0)).border = THIN_BORDER
        ws.cell(row=row, column=5, value=round(rd['total_emissions_kt'], 0)).border = THIN_BORDER
        ws.cell(row=row, column=6, value=round(rd['investment_musd'], 1)).border = THIN_BORDER
        ws.cell(row=row, column=7, value=round(rd['elec_demand_twh'], 1)).border = THIN_BORDER
        row += 1
    
    # Totals
    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=int(regional['num_facilities'].sum())).font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(regional['capacity_kt'].sum(), 0)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(regional['total_emissions_kt'].sum(), 0)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(regional['investment_musd'].sum(), 1)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(regional['elec_demand_twh'].sum(), 1)).font = Font(bold=True)
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = THIN_BORDER
    
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15
    
    return ws


def main():
    print("Loading data...")
    data = load_all_data()
    
    print("Creating workbook...")
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    print("Creating Assumptions sheet...")
    create_assumptions_sheet(wb, data)
    
    print("Creating Scenario Comparison...")
    create_scenario_comparison(wb, data)
    
    print("Creating Facility Transition...")
    create_facility_transition(wb, data)
    
    print("Creating Regional Summary...")
    create_regional_summary(wb, data)

    output_path = OUTPUT_DIR / 'Korea_Petrochemical_NetZero_Report.xlsx'
    wb.save(output_path)
    print(f"\n✓ Report generated: {output_path}")
    print("  - Assumptions with all references")
    print("  - NCC-Electricity vs NCC-H2 comparison")
    print("  - 248 facility transition schedule")
    print("  - Regional investment and energy demand")


if __name__ == "__main__":
    main()
