#!/usr/bin/env python3
"""
Update Excel MACC Matrix with Detailed CAPEX and Fuel Costs
Modify the existing Korean_Petrochemical_MACC_Model_Calibrated_52Mt.xlsx
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

class ExcelMACCUpdater:
    def __init__(self):
        """Initialize with existing Excel file"""
        self.excel_path = "../data/Korean_Petrochemical_MACC_Model_Calibrated_52Mt.xlsx"
        self.output_path = "../data/Korean_Petrochemical_MACC_Model_Updated_Detailed.xlsx"

        # Load existing workbook
        self.load_existing_data()
        self.define_updated_macc_technologies()

    def load_existing_data(self):
        """Load existing Excel data"""
        print("📊 Loading existing Excel data...")

        # Load existing sheets
        self.facilities_df = pd.read_excel(self.excel_path, sheet_name='source_Original')
        self.ci_df = pd.read_excel(self.excel_path, sheet_name='CI_Corrected', index_col=0)
        self.ci2_df = pd.read_excel(self.excel_path, sheet_name='CI2_Corrected', index_col=0)

        # Try to load existing MACC matrix if it exists
        try:
            self.existing_macc_df = pd.read_excel(self.excel_path, sheet_name='MACC_Matrix')
            print(f"   Found existing MACC matrix: {self.existing_macc_df.shape}")
        except:
            print("   No existing MACC matrix found, will create new one")
            self.existing_macc_df = None

        print(f"   Facilities: {len(self.facilities_df)}")
        print(f"   CI data: {self.ci_df.shape}")
        print(f"   CI2 data: {self.ci2_df.shape}")

    def define_updated_macc_technologies(self):
        """Define detailed MACC technologies with temporal variations"""
        print("🔧 Defining updated MACC technologies...")

        # Base year costs and temporal multipliers
        self.macc_technologies = {
            'Early_Retirement': {
                'description': 'Early retirement of oldest/least efficient facilities',
                'base_capex_usd_per_tco2': 0,
                'base_opex_usd_per_tco2': 50,
                'fuel_savings_usd_per_tco2': 0,
                'abatement_potential_mtco2': 5.2,  # 10% of 52 Mt baseline
                'applicable_to': 'All processes',
                'min_facility_age': 20,
                'technology_readiness': 'Commercial',
                'temporal_multipliers': {
                    2030: {'capex': 1.0, 'opex': 1.0},
                    2040: {'capex': 1.0, 'opex': 1.0},
                    2050: {'capex': 1.0, 'opex': 1.0}
                }
            },

            'Efficiency_Upgrade': {
                'description': 'Energy efficiency improvements (heat integration, optimization)',
                'base_capex_usd_per_tco2': 80,
                'base_opex_usd_per_tco2': 20,
                'fuel_savings_usd_per_tco2': 45,
                'abatement_potential_mtco2': 7.8,  # 15% potential
                'applicable_to': 'Naphtha Cracker, BTX Plant, Utility',
                'min_facility_age': 5,
                'technology_readiness': 'Commercial',
                'temporal_multipliers': {
                    2030: {'capex': 1.0, 'opex': 1.0},
                    2040: {'capex': 0.85, 'opex': 1.0},
                    2050: {'capex': 0.7, 'opex': 1.0}
                }
            },

            'Fuel_Switch': {
                'description': 'Switch from naphtha/LNG to cleaner fuels (renewable electricity, biomass)',
                'base_capex_usd_per_tco2': 150,
                'base_opex_usd_per_tco2': 30,
                'fuel_savings_usd_per_tco2': -20,  # Premium for clean fuels
                'abatement_potential_mtco2': 13.0,  # 25% potential
                'applicable_to': 'Naphtha Cracker, BTX Plant, Utility',
                'min_facility_age': 0,
                'technology_readiness': 'Commercial',
                'temporal_multipliers': {
                    2030: {'capex': 1.2, 'opex': 1.0},
                    2040: {'capex': 1.0, 'opex': 1.0},
                    2050: {'capex': 0.8, 'opex': 1.0}
                }
            },

            'Electrification': {
                'description': 'Electrification of thermal processes with renewable electricity',
                'base_capex_usd_per_tco2': 200,
                'base_opex_usd_per_tco2': 40,
                'fuel_savings_usd_per_tco2': -10,  # Premium for renewable electricity
                'abatement_potential_mtco2': 10.4,  # 20% potential
                'applicable_to': 'Utility, Naphtha Cracker',
                'min_facility_age': 0,
                'technology_readiness': 'Demonstration',
                'temporal_multipliers': {
                    2030: {'capex': 1.3, 'opex': 1.0},
                    2040: {'capex': 1.0, 'opex': 1.0},
                    2050: {'capex': 0.7, 'opex': 1.0}
                }
            },

            'Bio_Naphtha': {
                'description': 'Replace fossil naphtha with bio-based naphtha feedstock',
                'base_capex_usd_per_tco2': 120,
                'base_opex_usd_per_tco2': 80,
                'fuel_savings_usd_per_tco2': -70,  # Much higher feedstock cost
                'abatement_potential_mtco2': 7.8,  # 15% of naphtha
                'applicable_to': 'Naphtha Cracker',
                'min_facility_age': 0,
                'technology_readiness': 'Demonstration',
                'temporal_multipliers': {
                    2030: {'capex': 1.5, 'opex': 1.0},
                    2040: {'capex': 1.2, 'opex': 1.0},
                    2050: {'capex': 1.0, 'opex': 1.0}
                }
            },

            'Green_Hydrogen': {
                'description': 'Use green hydrogen for high-temperature processes',
                'base_capex_usd_per_tco2': 250,
                'base_opex_usd_per_tco2': 100,
                'fuel_savings_usd_per_tco2': -50,  # Premium for green H2
                'abatement_potential_mtco2': 6.2,  # 12% potential
                'applicable_to': 'Utility, Naphtha Cracker',
                'min_facility_age': 0,
                'technology_readiness': 'Pre-commercial',
                'temporal_multipliers': {
                    2030: {'capex': 2.0, 'opex': 1.0},
                    2040: {'capex': 1.5, 'opex': 1.0},
                    2050: {'capex': 1.0, 'opex': 1.0}
                }
            },

            'Process_Replacement': {
                'description': 'Replace entire process with low-carbon alternative technology',
                'base_capex_usd_per_tco2': 350,
                'base_opex_usd_per_tco2': 70,
                'fuel_savings_usd_per_tco2': 20,  # More efficient processes
                'abatement_potential_mtco2': 4.2,  # 8% potential
                'applicable_to': 'Naphtha Cracker, BTX Plant',
                'min_facility_age': 15,
                'technology_readiness': 'Research',
                'temporal_multipliers': {
                    2030: {'capex': 1.8, 'opex': 1.0},
                    2040: {'capex': 1.3, 'opex': 1.0},
                    2050: {'capex': 1.0, 'opex': 1.0}
                }
            }
        }

    def create_macc_matrix_sheets(self):
        """Create MACC matrix sheets for each time period"""
        print("📊 Creating MACC matrix sheets...")

        macc_sheets = {}

        for year in [2030, 2040, 2050]:
            print(f"   Creating MACC matrix for {year}...")

            # Create MACC matrix for this year
            macc_data = []

            for tech_name, tech_data in self.macc_technologies.items():
                # Calculate temporal costs
                capex_multiplier = tech_data['temporal_multipliers'][year]['capex']
                opex_multiplier = tech_data['temporal_multipliers'][year]['opex']

                capex_cost = tech_data['base_capex_usd_per_tco2'] * capex_multiplier
                opex_cost = tech_data['base_opex_usd_per_tco2'] * opex_multiplier
                fuel_savings = tech_data['fuel_savings_usd_per_tco2']

                # Net cost per tCO2
                net_cost = capex_cost + opex_cost - fuel_savings

                # Total investment required
                abatement_tco2 = tech_data['abatement_potential_mtco2'] * 1e6
                total_capex_musd = (capex_cost * abatement_tco2) / 1e6
                total_opex_musd = (opex_cost * abatement_tco2) / 1e6
                total_fuel_savings_musd = (fuel_savings * abatement_tco2) / 1e6
                total_investment_musd = total_capex_musd + total_opex_musd - total_fuel_savings_musd

                macc_data.append({
                    'technology': tech_name,
                    'description': tech_data['description'],
                    'applicable_to': tech_data['applicable_to'],
                    'technology_readiness': tech_data['technology_readiness'],
                    'min_facility_age': tech_data['min_facility_age'],
                    'abatement_potential_mtco2': tech_data['abatement_potential_mtco2'],
                    'capex_usd_per_tco2': capex_cost,
                    'opex_usd_per_tco2': opex_cost,
                    'fuel_savings_usd_per_tco2': fuel_savings,
                    'net_cost_usd_per_tco2': net_cost,
                    'total_capex_musd': total_capex_musd,
                    'total_opex_musd': total_opex_musd,
                    'total_fuel_savings_musd': total_fuel_savings_musd,
                    'total_investment_musd': total_investment_musd,
                    'capex_multiplier': capex_multiplier,
                    'opex_multiplier': opex_multiplier
                })

            # Sort by net cost
            macc_df = pd.DataFrame(macc_data)
            macc_df = macc_df.sort_values('net_cost_usd_per_tco2')

            # Add cumulative columns
            macc_df['cumulative_abatement_mtco2'] = macc_df['abatement_potential_mtco2'].cumsum()
            macc_df['cumulative_investment_musd'] = macc_df['total_investment_musd'].cumsum()

            macc_sheets[year] = macc_df

        return macc_sheets

    def update_excel_file(self, macc_sheets):
        """Update the Excel file with new MACC matrices"""
        print("💾 Updating Excel file with new MACC matrices...")

        # Copy original file
        import shutil
        shutil.copy2(self.excel_path, self.output_path)

        # Load the workbook
        book = load_workbook(self.output_path)

        # Remove old MACC sheets if they exist
        sheets_to_remove = []
        for sheet_name in book.sheetnames:
            if 'MACC' in sheet_name:
                sheets_to_remove.append(sheet_name)

        for sheet_name in sheets_to_remove:
            if sheet_name in book.sheetnames:
                del book[sheet_name]
                print(f"   Removed old sheet: {sheet_name}")

        # Add new MACC sheets
        with pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a') as writer:
            # Write temporal MACC matrices
            for year, macc_df in macc_sheets.items():
                sheet_name = f'MACC_{year}'
                macc_df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"   Added sheet: {sheet_name}")

            # Create summary comparison sheet
            summary_data = []
            for year, macc_df in macc_sheets.items():
                summary_data.append({
                    'year': year,
                    'total_abatement_potential_mtco2': macc_df['abatement_potential_mtco2'].sum(),
                    'total_investment_musd': macc_df['total_investment_musd'].sum(),
                    'lowest_cost_usd_per_tco2': macc_df['net_cost_usd_per_tco2'].min(),
                    'highest_cost_usd_per_tco2': macc_df['net_cost_usd_per_tco2'].max(),
                    'technologies_count': len(macc_df)
                })

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='MACC_Summary', index=False)
            print(f"   Added sheet: MACC_Summary")

            # Create technology master sheet
            tech_master = []
            for tech_name, tech_data in self.macc_technologies.items():
                tech_master.append({
                    'technology': tech_name,
                    'description': tech_data['description'],
                    'applicable_to': tech_data['applicable_to'],
                    'technology_readiness': tech_data['technology_readiness'],
                    'base_capex_usd_per_tco2': tech_data['base_capex_usd_per_tco2'],
                    'base_opex_usd_per_tco2': tech_data['base_opex_usd_per_tco2'],
                    'fuel_savings_usd_per_tco2': tech_data['fuel_savings_usd_per_tco2'],
                    'abatement_potential_mtco2': tech_data['abatement_potential_mtco2'],
                    'min_facility_age': tech_data['min_facility_age']
                })

            tech_master_df = pd.DataFrame(tech_master)
            tech_master_df.to_excel(writer, sheet_name='Technology_Master', index=False)
            print(f"   Added sheet: Technology_Master")

    def run_update(self):
        """Run complete Excel update"""
        print("🚀 UPDATING EXCEL MACC MATRIX WITH DETAILED COSTS")
        print("=" * 80)
        print(f"📁 Source: {self.excel_path}")
        print(f"📁 Output: {self.output_path}")
        print()

        try:
            # Create MACC matrices
            macc_sheets = self.create_macc_matrix_sheets()

            # Update Excel file
            self.update_excel_file(macc_sheets)

            print("\n🎯 EXCEL UPDATE SUMMARY:")
            print(f"   📊 Technologies defined: {len(self.macc_technologies)}")
            print(f"   📈 Time periods: 2030, 2040, 2050")
            print(f"   💾 Updated file: {self.output_path}")

            for year, macc_df in macc_sheets.items():
                total_abatement = macc_df['abatement_potential_mtco2'].sum()
                total_investment = macc_df['total_investment_musd'].sum()
                cost_range = f"${macc_df['net_cost_usd_per_tco2'].min():.0f}-${macc_df['net_cost_usd_per_tco2'].max():.0f}"
                print(f"   {year}: {total_abatement:.1f} MtCO₂, ${total_investment:.0f}M, {cost_range}/tCO₂")

            return macc_sheets

        except Exception as e:
            print(f"❌ Error updating Excel file: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

if __name__ == "__main__":
    updater = ExcelMACCUpdater()
    results = updater.run_update()

    print("\n✅ EXCEL MACC MATRIX UPDATE COMPLETE!")
    print("📁 Use the updated file for all future model runs")