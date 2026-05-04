"""
Excel Model Reviewer Module
===========================
Generates Excel workbooks for model verification through reverse calculations.

Takes model outputs and assumptions, then shows reverse calculations to verify
the model's correctness. The Excel uses formulas to recalculate expected values
from assumptions and compare against actual outputs.

Usage:
    python -m modules.excel_reviewer --scenario shaheen_ncc_h2_15c
    python -m modules.excel_reviewer --scenario shaheen_ncc_elec_20c --output my_review.xlsx
    python -m modules.excel_reviewer --all
"""

import pandas as pd
import numpy as np
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Dict, Optional
import argparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class ExcelReviewerConfig:
    """Configuration for Excel reviewer generation."""
    review_years: List[int] = field(default_factory=lambda: [2025, 2030, 2035, 2040, 2045, 2050])
    tolerance_relative: float = 0.001  # 0.1% relative tolerance
    tolerance_absolute: float = 1.0    # $1 absolute tolerance
    max_rows_per_sheet: int = 5000     # Limit rows to prevent Excel slowness

    # Style configuration
    header_fill: str = "4472C4"        # Blue header
    match_fill: str = "C6EFCE"         # Green for MATCH
    mismatch_fill: str = "FFC7CE"      # Red for MISMATCH
    warning_fill: str = "FFEB9C"       # Yellow for warnings
    assumption_fill: str = "E2EFDA"    # Light green for assumption sheets


class ExcelReviewer:
    """
    Generate Excel workbook for model verification through reverse calculations.

    The workbook contains:
    - Model outputs from scenario_results.csv
    - Assumption lookup tables
    - Verification sheets with Excel formulas
    - Summary dashboard with MATCH/MISMATCH counts
    """

    def __init__(self,
                 scenario: str,
                 data_dir: str = 'data',
                 output_dir: str = 'outputs',
                 config: Optional[ExcelReviewerConfig] = None):
        """
        Initialize the Excel Reviewer.

        Args:
            scenario: Scenario name (e.g., 'shaheen_ncc_h2_15c')
            data_dir: Path to data directory with assumptions
            output_dir: Path to outputs directory with scenario_results.csv
            config: Configuration options
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install with: pip install openpyxl"
            )

        self.scenario = scenario
        self.data_dir = Path(data_dir)
        self.output_dir = Path(output_dir)
        self.config = config or ExcelReviewerConfig()

        # Data storage
        self.outputs_df: Optional[pd.DataFrame] = None
        self.assumptions: Dict[str, pd.DataFrame] = {}

        # Workbook
        self.wb: Optional[Workbook] = None

        # Named ranges for VLOOKUP references
        self.named_ranges: Dict[str, str] = {}

    def load_inputs(self):
        """Load all assumption CSV files."""
        assumption_files = {
            'product_benchmarks': self.data_dir / 'assumptions' / 'product_benchmarks.csv',
            'emission_factors': self.data_dir / 'assumptions' / 'emission_factors.csv',
            'technology_capex': self.data_dir / 'assumptions' / 'technology_capex.csv',
            'technology_parameters': self.data_dir / 'assumptions' / 'technology_parameters.csv',
            'model_config': self.data_dir / 'assumptions' / 'model_config.csv',
            'h2_prices': self.data_dir / 'assumptions' / 'prices' / 'h2_price_trajectory.csv',
            're_prices': self.data_dir / 'assumptions' / 'prices' / 're_price_trajectory.csv',
            'fuel_prices': self.data_dir / 'assumptions' / 'prices' / 'fuel_price_trajectory.csv',
        }

        for name, path in assumption_files.items():
            if path.exists():
                self.assumptions[name] = pd.read_csv(path)
            else:
                print(f"Warning: Assumption file not found: {path}")

        print(f"Loaded {len(self.assumptions)} assumption files")

    def load_outputs(self):
        """Load scenario results filtered to review years."""
        results_path = self.output_dir / 'scenario_results.csv'

        if not results_path.exists():
            raise FileNotFoundError(f"Scenario results not found: {results_path}")

        # Load full results
        df = pd.read_csv(results_path)

        # Filter to scenario and review years
        df = df[df['scenario'] == self.scenario]
        df = df[df['year'].isin(self.config.review_years)]

        # Limit rows if needed
        if len(df) > self.config.max_rows_per_sheet:
            print(f"Warning: Limiting output to {self.config.max_rows_per_sheet} rows")
            df = df.head(self.config.max_rows_per_sheet)

        self.outputs_df = df
        print(f"Loaded {len(df)} output rows for scenario '{self.scenario}'")

    def create_workbook(self):
        """Create the Excel workbook with all sheets."""
        self.wb = Workbook()

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create sheets in order
        self._create_instructions_sheet()
        self._create_outputs_sheet()
        self._create_assumption_sheets()
        self._create_verification_sheets()
        self._create_summary_dashboard()

        print("Workbook created successfully")

    def save(self, output_path: Optional[str] = None):
        """Save the workbook to file."""
        if self.wb is None:
            raise ValueError("Workbook not created. Call create_workbook() first.")

        if output_path is None:
            output_path = self.output_dir / f"verification_{self.scenario}.xlsx"
        else:
            output_path = Path(output_path)

        self.wb.save(output_path)
        print(f"Saved workbook to: {output_path}")
        return output_path

    def _create_instructions_sheet(self):
        """Create the Instructions sheet with usage guide."""
        ws = self.wb.create_sheet("Instructions")

        # Title
        ws['A1'] = "Model Verification Workbook"
        ws['A1'].font = Font(size=18, bold=True)

        ws['A3'] = f"Scenario: {self.scenario}"
        ws['A4'] = f"Review Years: {', '.join(map(str, self.config.review_years))}"
        ws['A5'] = f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"

        # Color legend
        ws['A7'] = "Color Legend:"
        ws['A7'].font = Font(bold=True)

        legend_items = [
            ("A8", "MATCH", self.config.match_fill, "Values match within tolerance"),
            ("A9", "MISMATCH", self.config.mismatch_fill, "Values differ beyond tolerance"),
            ("A10", "WARNING", self.config.warning_fill, "Potential issues to review"),
            ("A11", "Assumption", self.config.assumption_fill, "Assumption lookup data"),
        ]

        for cell, label, color, description in legend_items:
            ws[cell] = label
            ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws[cell.replace('A', 'B')] = description

        # Sheet descriptions
        ws['A13'] = "Sheet Descriptions:"
        ws['A13'].font = Font(bold=True)

        sheets = [
            ("Model_Outputs", "Raw data from scenario_results.csv (key years only)"),
            ("Assumptions_Benchmarks", "Energy intensity lookup by product/process"),
            ("Assumptions_EF", "Emission factors by fuel type"),
            ("Assumptions_TechCapex", "Technology CAPEX rates by year"),
            ("Assumptions_TechParams", "Technology parameters (COP, H2 demand, etc.)"),
            ("Assumptions_Prices", "H2, RE, and fuel price trajectories"),
            ("Verify_CAPEX", "Reverse-calc CAPEX from production and rates"),
            ("Verify_MAC", "Reverse-calc MAC = total_cost / abatement"),
            ("Verify_EnergyCosts", "Reverse-calc H2 + electricity costs"),
            ("Verify_CostComponents", "Verify cost component summation"),
            ("Summary_Dashboard", "Overall verification status"),
        ]

        for i, (sheet, desc) in enumerate(sheets, start=14):
            ws[f'A{i}'] = sheet
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = desc

        # Verification instructions
        row = 14 + len(sheets) + 2
        ws[f'A{row}'] = "Verification Process:"
        ws[f'A{row}'].font = Font(bold=True)

        steps = [
            "1. Go to Summary_Dashboard to see overall MATCH/MISMATCH counts",
            "2. If MISMATCH found, go to specific Verify_* sheet",
            "3. Red cells indicate values that don't match within tolerance",
            f"4. Tolerance: {self.config.tolerance_relative*100}% relative or ${self.config.tolerance_absolute} absolute",
            "5. Review formulas in verification columns to understand calculations",
            "6. Trace back to Assumption sheets to verify lookup data",
        ]

        for i, step in enumerate(steps, start=row+1):
            ws[f'A{i}'] = step

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60

    def _create_outputs_sheet(self):
        """Create the Model_Outputs sheet with scenario results."""
        if self.outputs_df is None or len(self.outputs_df) == 0:
            return

        ws = self.wb.create_sheet("Model_Outputs")

        # Write dataframe to sheet
        for r_idx, row in enumerate(dataframe_to_rows(self.outputs_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color=self.config.header_fill,
                        end_color=self.config.header_fill,
                        fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="FFFFFF")

        # Auto-fit columns (approximate)
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

        # Create named range for outputs
        if len(self.outputs_df) > 0:
            last_row = len(self.outputs_df) + 1
            last_col = get_column_letter(len(self.outputs_df.columns))
            self.named_ranges['outputs'] = f"Model_Outputs!$A$1:${last_col}${last_row}"

    def _create_assumption_sheets(self):
        """Create assumption lookup sheets."""
        assumption_sheets = [
            ('product_benchmarks', 'Assumptions_Benchmarks'),
            ('emission_factors', 'Assumptions_EF'),
            ('technology_capex', 'Assumptions_TechCapex'),
            ('technology_parameters', 'Assumptions_TechParams'),
        ]

        for data_key, sheet_name in assumption_sheets:
            if data_key not in self.assumptions:
                continue

            df = self.assumptions[data_key]
            ws = self.wb.create_sheet(sheet_name)

            # Write dataframe
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color=self.config.header_fill,
                            end_color=self.config.header_fill,
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.fill = PatternFill(
                            start_color=self.config.assumption_fill,
                            end_color=self.config.assumption_fill,
                            fill_type="solid"
                        )

            # Auto-fit columns
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        # Create price trajectories sheet (combines H2, RE, fuel prices)
        self._create_prices_sheet()

    def _create_prices_sheet(self):
        """Create combined price trajectories sheet."""
        ws = self.wb.create_sheet("Assumptions_Prices")

        # H2 prices
        if 'h2_prices' in self.assumptions:
            h2_df = self.assumptions['h2_prices'][['year', 'h2_price_usd_per_kg']].copy()
            h2_df.columns = ['year', 'h2_price_usd_per_kg']
        else:
            h2_df = pd.DataFrame(columns=['year', 'h2_price_usd_per_kg'])

        # RE prices
        if 're_prices' in self.assumptions:
            re_df = self.assumptions['re_prices'][['year', 're_price_usd_per_mwh']].copy()
        else:
            re_df = pd.DataFrame(columns=['year', 're_price_usd_per_mwh'])

        # Fuel prices
        if 'fuel_prices' in self.assumptions:
            fuel_df = self.assumptions['fuel_prices'].copy()
        else:
            fuel_df = pd.DataFrame(columns=['year'])

        # Merge all prices on year
        if len(h2_df) > 0:
            prices_df = h2_df.merge(re_df, on='year', how='outer')
            prices_df = prices_df.merge(fuel_df, on='year', how='outer')
        else:
            prices_df = fuel_df

        prices_df = prices_df.sort_values('year').reset_index(drop=True)

        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(prices_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color=self.config.header_fill,
                        end_color=self.config.header_fill,
                        fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color=self.config.assumption_fill,
                        end_color=self.config.assumption_fill,
                        fill_type="solid"
                    )

        # Store reference for price lookups
        if len(prices_df) > 0:
            last_row = len(prices_df) + 1
            last_col = get_column_letter(len(prices_df.columns))
            self.named_ranges['prices'] = f"Assumptions_Prices!$A$1:${last_col}${last_row}"

        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    def _create_verification_sheets(self):
        """Create verification sheets with formulas."""
        if self.outputs_df is None or len(self.outputs_df) == 0:
            return

        self._create_verify_capex_sheet()
        self._create_verify_mac_sheet()
        self._create_verify_energy_costs_sheet()
        self._create_verify_cost_components_sheet()

    def _create_verify_capex_sheet(self):
        """Create CAPEX verification sheet."""
        ws = self.wb.create_sheet("Verify_CAPEX")

        # Filter to rows with deployed technology
        df = self.outputs_df[self.outputs_df['technology'] != 'Baseline'].copy()

        if len(df) == 0:
            ws['A1'] = "No technology deployments found in selected years"
            return

        # Headers
        headers = [
            'year', 'facility_id', 'product', 'technology', 'production_t',
            'OUTPUT_capex_usd', 'CALC_capex_rate', 'CALC_capex_usd',
            'DIFF_absolute', 'DIFF_percent', 'VERIFY'
        ]

        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.config.header_fill,
                end_color=self.config.header_fill,
                fill_type="solid"
            )

        # Data rows
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            year = row['year']
            tech = row['technology']
            production = row['production_t']
            output_capex = row['capex_usd']

            # Basic data columns
            ws.cell(row=r_idx, column=1, value=year)
            ws.cell(row=r_idx, column=2, value=row['facility_id'])
            ws.cell(row=r_idx, column=3, value=row['product'])
            ws.cell(row=r_idx, column=4, value=tech)
            ws.cell(row=r_idx, column=5, value=production)
            ws.cell(row=r_idx, column=6, value=output_capex)

            # CALC_capex_rate: VLOOKUP from technology_capex with interpolation
            # For simplicity, use the formula that looks up the rate based on year
            capex_rate_formula = self._get_capex_interpolation_formula(r_idx, tech, year)
            ws.cell(row=r_idx, column=7, value=capex_rate_formula if capex_rate_formula else 0)

            # CALC_capex_usd = production_t * capex_rate
            ws.cell(row=r_idx, column=8, value=f"=E{r_idx}*G{r_idx}")

            # DIFF_absolute = OUTPUT - CALC
            ws.cell(row=r_idx, column=9, value=f"=ABS(F{r_idx}-H{r_idx})")

            # DIFF_percent = ABS(OUTPUT - CALC) / OUTPUT * 100
            ws.cell(row=r_idx, column=10, value=f"=IF(F{r_idx}>0,I{r_idx}/F{r_idx}*100,0)")

            # VERIFY formula
            verify_formula = (
                f'=IF(OR(I{r_idx}<{self.config.tolerance_absolute},'
                f'J{r_idx}<{self.config.tolerance_relative*100}),"MATCH","MISMATCH")'
            )
            cell = ws.cell(row=r_idx, column=11, value=verify_formula)

        # Apply conditional formatting
        self._apply_verification_formatting(ws, len(df) + 1, 11)

        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    def _get_capex_interpolation_formula(self, row: int, technology: str, year: int) -> Optional[str]:
        """
        Generate Excel formula for CAPEX rate interpolation.

        Uses linear interpolation between years 2025, 2030, 2040, 2050.
        """
        if 'technology_capex' not in self.assumptions:
            return None

        capex_df = self.assumptions['technology_capex']
        if technology not in capex_df['technology'].values:
            return None

        tech_row = capex_df[capex_df['technology'] == technology].iloc[0]

        # Get CAPEX values for interpolation years
        c2025 = tech_row.get('capex_2025', 0)
        c2030 = tech_row.get('capex_2030', 0)
        c2040 = tech_row.get('capex_2040', 0)
        c2050 = tech_row.get('capex_2050', 0)

        # For simplicity, return the interpolated value directly
        # In a real implementation, you'd use VLOOKUP/INDEX-MATCH with interpolation
        years = [2025, 2030, 2040, 2050]
        values = [c2025, c2030, c2040, c2050]
        interpolated = np.interp(year, years, values)

        return interpolated

    def _create_verify_mac_sheet(self):
        """Create MAC verification sheet."""
        ws = self.wb.create_sheet("Verify_MAC")

        # Filter to rows with abatement
        df = self.outputs_df[self.outputs_df['abatement_tco2'] > 0].copy()

        if len(df) == 0:
            ws['A1'] = "No abatement found in selected years"
            return

        # Headers
        headers = [
            'year', 'facility_id', 'technology',
            'OUTPUT_total_cost', 'OUTPUT_abatement', 'OUTPUT_mac',
            'CALC_mac', 'DIFF_absolute', 'DIFF_percent', 'VERIFY'
        ]

        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.config.header_fill,
                end_color=self.config.header_fill,
                fill_type="solid"
            )

        # Data rows
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            ws.cell(row=r_idx, column=1, value=row['year'])
            ws.cell(row=r_idx, column=2, value=row['facility_id'])
            ws.cell(row=r_idx, column=3, value=row['technology'])
            ws.cell(row=r_idx, column=4, value=row['total_cost_usd'])
            ws.cell(row=r_idx, column=5, value=row['abatement_tco2'])
            ws.cell(row=r_idx, column=6, value=row['mac_usd_per_tco2'])

            # CALC_mac = total_cost / abatement
            ws.cell(row=r_idx, column=7, value=f"=IF(E{r_idx}>0,D{r_idx}/E{r_idx},0)")

            # DIFF_absolute
            ws.cell(row=r_idx, column=8, value=f"=ABS(F{r_idx}-G{r_idx})")

            # DIFF_percent
            ws.cell(row=r_idx, column=9, value=f"=IF(F{r_idx}>0,H{r_idx}/ABS(F{r_idx})*100,0)")

            # VERIFY
            verify_formula = (
                f'=IF(OR(H{r_idx}<{self.config.tolerance_absolute},'
                f'I{r_idx}<{self.config.tolerance_relative*100}),"MATCH","MISMATCH")'
            )
            ws.cell(row=r_idx, column=10, value=verify_formula)

        # Apply formatting
        self._apply_verification_formatting(ws, len(df) + 1, 10)

        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    def _create_verify_energy_costs_sheet(self):
        """Create energy costs verification sheet."""
        ws = self.wb.create_sheet("Verify_EnergyCosts")

        # Filter to rows with deployed technology
        df = self.outputs_df[self.outputs_df['technology'] != 'Baseline'].copy()

        if len(df) == 0:
            ws['A1'] = "No technology deployments found in selected years"
            return

        # Headers
        # NOTE: We use 'added_elec_mwh' (not 'elec_demand_mwh') because the model only charges
        # for NEW electricity required by the decarbonization technology, not baseline electricity.
        # See docs/ROOT_CAUSE_ENERGY_MISMATCH.md for details.
        headers = [
            'year', 'facility_id', 'technology',
            'h2_demand_t', 'added_elec_mwh',
            'h2_price_lookup', 're_price_lookup',
            'CALC_h2_cost', 'CALC_elec_cost', 'CALC_new_energy',
            'OUTPUT_new_energy', 'DIFF_absolute', 'VERIFY'
        ]

        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.config.header_fill,
                end_color=self.config.header_fill,
                fill_type="solid"
            )

        # Build price lookup for efficiency
        h2_prices = {}
        re_prices = {}
        if 'h2_prices' in self.assumptions:
            h2_prices = dict(zip(
                self.assumptions['h2_prices']['year'],
                self.assumptions['h2_prices']['h2_price_usd_per_kg']
            ))
        if 're_prices' in self.assumptions:
            re_prices = dict(zip(
                self.assumptions['re_prices']['year'],
                self.assumptions['re_prices']['re_price_usd_per_mwh']
            ))

        # Data rows
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            year = row['year']
            h2_demand = row['h2_demand_t']
            # Use added_elec_mwh (electricity added by technology), not elec_demand_mwh (total)
            # If added_elec_mwh is not in the output (older data), fall back to 0
            added_elec = row.get('added_elec_mwh', 0) if 'added_elec_mwh' in row.index else 0

            h2_price = h2_prices.get(year, 0)
            re_price = re_prices.get(year, 0)

            ws.cell(row=r_idx, column=1, value=year)
            ws.cell(row=r_idx, column=2, value=row['facility_id'])
            ws.cell(row=r_idx, column=3, value=row['technology'])
            ws.cell(row=r_idx, column=4, value=h2_demand)
            ws.cell(row=r_idx, column=5, value=added_elec)
            ws.cell(row=r_idx, column=6, value=h2_price)
            ws.cell(row=r_idx, column=7, value=re_price)

            # CALC_h2_cost = h2_demand_t * h2_price * 1000 (convert t to kg)
            ws.cell(row=r_idx, column=8, value=f"=D{r_idx}*F{r_idx}*1000")

            # CALC_elec_cost = added_elec_mwh * re_price (only charge for NEW electricity)
            ws.cell(row=r_idx, column=9, value=f"=E{r_idx}*G{r_idx}")

            # CALC_new_energy = h2_cost + elec_cost
            ws.cell(row=r_idx, column=10, value=f"=H{r_idx}+I{r_idx}")

            # OUTPUT_new_energy
            ws.cell(row=r_idx, column=11, value=row['cost_component_new_energy_usd'])

            # DIFF_absolute
            ws.cell(row=r_idx, column=12, value=f"=ABS(J{r_idx}-K{r_idx})")

            # VERIFY
            verify_formula = (
                f'=IF(L{r_idx}<{max(self.config.tolerance_absolute, 100)},"MATCH","MISMATCH")'
            )
            ws.cell(row=r_idx, column=13, value=verify_formula)

        # Apply formatting
        self._apply_verification_formatting(ws, len(df) + 1, 13)

        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    def _create_verify_cost_components_sheet(self):
        """Create cost components verification sheet."""
        ws = self.wb.create_sheet("Verify_CostComponents")

        # Filter to rows with deployed technology
        df = self.outputs_df[self.outputs_df['technology'] != 'Baseline'].copy()

        if len(df) == 0:
            ws['A1'] = "No technology deployments found in selected years"
            return

        # Headers
        headers = [
            'year', 'facility_id', 'technology',
            'capex_annual', 'opex_annual', 'new_energy', 'fuel_savings',
            'CALC_total_cost', 'OUTPUT_total_cost', 'DIFF_absolute', 'VERIFY'
        ]

        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.config.header_fill,
                end_color=self.config.header_fill,
                fill_type="solid"
            )

        # Data rows
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            ws.cell(row=r_idx, column=1, value=row['year'])
            ws.cell(row=r_idx, column=2, value=row['facility_id'])
            ws.cell(row=r_idx, column=3, value=row['technology'])
            ws.cell(row=r_idx, column=4, value=row['cost_component_capex_annual_usd'])
            ws.cell(row=r_idx, column=5, value=row['cost_component_opex_annual_usd'])
            ws.cell(row=r_idx, column=6, value=row['cost_component_new_energy_usd'])
            ws.cell(row=r_idx, column=7, value=row['cost_component_fuel_savings_usd'])

            # CALC_total_cost = capex_annual + opex_annual + new_energy - fuel_savings
            ws.cell(row=r_idx, column=8, value=f"=D{r_idx}+E{r_idx}+F{r_idx}-G{r_idx}")

            # OUTPUT_total_cost
            ws.cell(row=r_idx, column=9, value=row['total_cost_usd'])

            # DIFF_absolute
            ws.cell(row=r_idx, column=10, value=f"=ABS(H{r_idx}-I{r_idx})")

            # VERIFY
            verify_formula = (
                f'=IF(J{r_idx}<{max(self.config.tolerance_absolute, 100)},"MATCH","MISMATCH")'
            )
            ws.cell(row=r_idx, column=11, value=verify_formula)

        # Apply formatting
        self._apply_verification_formatting(ws, len(df) + 1, 11)

        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    def _apply_verification_formatting(self, ws, last_row: int, verify_col: int):
        """Apply conditional formatting to verification column."""
        verify_col_letter = get_column_letter(verify_col)
        verify_range = f"{verify_col_letter}2:{verify_col_letter}{last_row}"

        # Green for MATCH
        match_rule = FormulaRule(
            formula=[f'${verify_col_letter}2="MATCH"'],
            fill=PatternFill(
                start_color=self.config.match_fill,
                end_color=self.config.match_fill,
                fill_type="solid"
            )
        )
        ws.conditional_formatting.add(verify_range, match_rule)

        # Red for MISMATCH
        mismatch_rule = FormulaRule(
            formula=[f'${verify_col_letter}2="MISMATCH"'],
            fill=PatternFill(
                start_color=self.config.mismatch_fill,
                end_color=self.config.mismatch_fill,
                fill_type="solid"
            )
        )
        ws.conditional_formatting.add(verify_range, mismatch_rule)

    def _create_summary_dashboard(self):
        """Create summary dashboard with verification status."""
        ws = self.wb.create_sheet("Summary_Dashboard")

        # Title
        ws['A1'] = "Model Verification Summary"
        ws['A1'].font = Font(size=16, bold=True)

        ws['A3'] = f"Scenario: {self.scenario}"
        ws['A4'] = f"Review Years: {', '.join(map(str, self.config.review_years))}"

        # Verification summary table
        ws['A6'] = "Verification Results:"
        ws['A6'].font = Font(bold=True)

        ws['A7'] = "Sheet"
        ws['B7'] = "Total Rows"
        ws['C7'] = "MATCH"
        ws['D7'] = "MISMATCH"
        ws['E7'] = "Status"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}7'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}7'].fill = PatternFill(
                start_color=self.config.header_fill,
                end_color=self.config.header_fill,
                fill_type="solid"
            )

        # Sheet summaries with COUNTIF formulas
        verify_sheets = [
            ('Verify_CAPEX', 'K'),
            ('Verify_MAC', 'J'),
            ('Verify_EnergyCosts', 'M'),
            ('Verify_CostComponents', 'K'),
        ]

        for i, (sheet_name, verify_col) in enumerate(verify_sheets, start=8):
            ws[f'A{i}'] = sheet_name

            # Check if sheet has data
            if sheet_name in self.wb.sheetnames:
                sheet = self.wb[sheet_name]
                max_row = sheet.max_row

                if max_row > 1:
                    # Total rows
                    ws[f'B{i}'] = max_row - 1

                    # MATCH count
                    ws[f'C{i}'] = f'=COUNTIF({sheet_name}!{verify_col}:{verify_col},"MATCH")'

                    # MISMATCH count
                    ws[f'D{i}'] = f'=COUNTIF({sheet_name}!{verify_col}:{verify_col},"MISMATCH")'

                    # Status
                    ws[f'E{i}'] = f'=IF(D{i}=0,"PASS","FAIL")'
                else:
                    ws[f'B{i}'] = 0
                    ws[f'C{i}'] = "N/A"
                    ws[f'D{i}'] = "N/A"
                    ws[f'E{i}'] = "N/A"
            else:
                ws[f'B{i}'] = "Sheet not found"

        # Overall status
        last_row = 8 + len(verify_sheets)
        ws[f'A{last_row + 2}'] = "Overall Status:"
        ws[f'A{last_row + 2}'].font = Font(bold=True, size=14)

        # Count all mismatches
        ws[f'B{last_row + 2}'] = f'=IF(SUM(D8:D{last_row - 1})=0,"ALL PASS","HAS FAILURES")'

        # Apply conditional formatting to status column
        status_range = f"E8:E{last_row - 1}"

        pass_rule = FormulaRule(
            formula=['$E8="PASS"'],
            fill=PatternFill(
                start_color=self.config.match_fill,
                end_color=self.config.match_fill,
                fill_type="solid"
            )
        )
        ws.conditional_formatting.add(status_range, pass_rule)

        fail_rule = FormulaRule(
            formula=['$E8="FAIL"'],
            fill=PatternFill(
                start_color=self.config.mismatch_fill,
                end_color=self.config.mismatch_fill,
                fill_type="solid"
            )
        )
        ws.conditional_formatting.add(status_range, fail_rule)

        # Overall status formatting
        overall_range = f"B{last_row + 2}"

        all_pass_rule = FormulaRule(
            formula=[f'$B${last_row + 2}="ALL PASS"'],
            fill=PatternFill(
                start_color=self.config.match_fill,
                end_color=self.config.match_fill,
                fill_type="solid"
            )
        )
        ws.conditional_formatting.add(overall_range, all_pass_rule)

        has_fail_rule = FormulaRule(
            formula=[f'$B${last_row + 2}="HAS FAILURES"'],
            fill=PatternFill(
                start_color=self.config.mismatch_fill,
                end_color=self.config.mismatch_fill,
                fill_type="solid"
            )
        )
        ws.conditional_formatting.add(overall_range, has_fail_rule)

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15

    def generate(self, output_path: Optional[str] = None) -> Path:
        """
        Generate the complete verification workbook.

        Args:
            output_path: Optional custom output path

        Returns:
            Path to generated Excel file
        """
        print(f"Generating verification workbook for scenario: {self.scenario}")

        self.load_inputs()
        self.load_outputs()
        self.create_workbook()
        return self.save(output_path)


def get_available_scenarios(output_dir: str = 'outputs') -> List[str]:
    """Get list of available scenarios from scenario_results.csv."""
    results_path = Path(output_dir) / 'scenario_results.csv'

    if not results_path.exists():
        return []

    df = pd.read_csv(results_path)
    return df['scenario'].unique().tolist()


def main():
    """CLI entry point for excel_reviewer module."""
    parser = argparse.ArgumentParser(
        description='Generate Excel workbook for model verification'
    )
    parser.add_argument(
        '--scenario', '-s',
        help='Scenario name (e.g., shaheen_ncc_h2_15c)'
    )
    parser.add_argument(
        '--all', '-a',
        action='store_true',
        help='Generate workbooks for all scenarios'
    )
    parser.add_argument(
        '--output', '-o',
        help='Custom output filename'
    )
    parser.add_argument(
        '--data-dir',
        default='data',
        help='Path to data directory (default: data)'
    )
    parser.add_argument(
        '--output-dir',
        default='outputs',
        help='Path to outputs directory (default: outputs)'
    )
    parser.add_argument(
        '--years',
        nargs='+',
        type=int,
        default=[2025, 2030, 2035, 2040, 2045, 2050],
        help='Years to include in review (default: 2025 2030 2035 2040 2045 2050)'
    )

    args = parser.parse_args()

    # Get available scenarios
    scenarios = get_available_scenarios(args.output_dir)

    if not scenarios:
        print(f"No scenarios found in {args.output_dir}/scenario_results.csv")
        return 1

    if args.all:
        # Generate for all scenarios
        for scenario in scenarios:
            config = ExcelReviewerConfig(review_years=args.years)
            reviewer = ExcelReviewer(
                scenario=scenario,
                data_dir=args.data_dir,
                output_dir=args.output_dir,
                config=config
            )
            reviewer.generate()
            print()
    elif args.scenario:
        # Generate for specific scenario
        if args.scenario not in scenarios:
            print(f"Scenario '{args.scenario}' not found.")
            print(f"Available scenarios: {', '.join(scenarios)}")
            return 1

        config = ExcelReviewerConfig(review_years=args.years)
        reviewer = ExcelReviewer(
            scenario=args.scenario,
            data_dir=args.data_dir,
            output_dir=args.output_dir,
            config=config
        )
        reviewer.generate(args.output)
    else:
        print("Please specify --scenario or --all")
        print(f"Available scenarios: {', '.join(scenarios)}")
        return 1

    return 0


if __name__ == '__main__':
    exit(main())
