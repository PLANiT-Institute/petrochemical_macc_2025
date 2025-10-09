#!/usr/bin/env python3
"""
Enhanced Excel Configuration with Naphtha Emission Intensities and Alternative Strategies
Korean Petrochemical Industry MACC Model
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_enhanced_excel_with_naphtha():
    """Create comprehensive Excel file with naphtha CI and alternative strategies"""
    
    print("🛢️ CREATING ENHANCED EXCEL WITH NAPHTHA STRATEGIES")
    print("=" * 70)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ============================================================================
    # 1. BASELINE DATA SHEET (Updated with naphtha details)
    # ============================================================================
    
    baseline_data = {
        'Category': [
            'Total Industry', 'Total Industry', 'Total Industry', 'Total Industry',
            'NCC Process', 'NCC Process', 'NCC Process', 'NCC Process', 'NCC Process', 'NCC Process', 'NCC Process',
            'Other Processes', 'Other Processes', 'Other Processes', 'Other Processes',
            'Naphtha Consumption', 'Naphtha Consumption', 'Naphtha Consumption', 'Naphtha Consumption'
        ],
        'Subcategory': [
            'Total Production', 'Total Energy Demand', 'Total Fuel Consumption', 'Total Emissions',
            'Ethylene Production', 'Propylene Production', 'Total Naphtha Feedstock', 'Thermal Energy Naphtha', 
            'NCC Naphtha Combustion', 'NCC Steam & Electricity', 'Total NCC Emissions',
            'BTX Production', 'Utility Operations', 'Other Processes', 'Other Process Emissions',
            'Feedstock Naphtha', 'Thermal Naphtha', 'Total Naphtha', 'Naphtha Emissions'
        ],
        'Value': [
            15.1, 0.4, 9.43, 52.0,
            6.1, 3.1, 26.7, 3.5,
            24.0, 16.0, 40.0,
            4.1, 2.4, 1.8, 12.0,
            26.7, 3.5, 30.2, 35.5
        ],
        'Unit': [
            'Mt/year', 'PJ/year', 'Mt/year', 'MtCO2/year',
            'Mt/year', 'Mt/year', 'Mt/year', 'Mt/year',
            'MtCO2/year', 'MtCO2/year', 'MtCO2/year',
            'Mt/year', 'Mt/year', 'Mt/year', 'MtCO2/year',
            'Mt/year', 'Mt/year', 'Mt/year', 'MtCO2/year'
        ],
        'Notes': [
            'Industry provided', 'All processes and utilities', 'Natural gas, LPG, fuel oil, coal', 'Industry validated baseline',
            '85% capacity utilization', '85% capacity utilization', '3.1:1 ratio to ethylene', 'For steam cracking heating',
            '46.1% of total emissions', '30.7% of total emissions', '76.8% of total emissions',
            '85% capacity utilization', 'Steam, power equivalent', 'Polymers, chemicals', '23.2% of total emissions',
            'Petrochemical feedstock', 'High-temperature heating', 'NCC process only', 'Lifecycle + combustion'
        ]
    }
    
    ws_baseline = wb.create_sheet("Baseline_Data")
    baseline_df = pd.DataFrame(baseline_data)
    
    for r in dataframe_to_rows(baseline_df, index=False, header=True):
        ws_baseline.append(r)
    
    # ============================================================================
    # 2. SOURCE SHEET (Facility-level production data)
    # ============================================================================
    
    source_data = {
        'Facility_Code': ['F001', 'F002', 'F003', 'F004', 'F005', 'F006', 'F007', 'F008'],
        'Company': ['SK Innovation', 'LG Chem', 'Lotte Chemical', 'Hanwha Solutions', 'Samsung Total', 'Kumho Petrochemical', 'S-Oil', 'GS Caltex'],
        'Process_Type': ['NCC', 'NCC', 'NCC', 'BTX', 'NCC', 'BTX', 'NCC', 'BTX'],
        'Primary_Product': ['Ethylene', 'Ethylene', 'Propylene', 'Benzene', 'Ethylene', 'Toluene', 'Propylene', 'Xylene'],
        'Production_Capacity_kt': [1200, 950, 650, 420, 780, 380, 550, 340],
        'Capacity_Utilization_%': [85, 88, 82, 90, 86, 85, 84, 87],
        'Actual_Production_kt': [1020, 836, 533, 378, 671, 323, 462, 296],
        'Naphtha_Consumption_kt': [3162, 2592, 1499, 529, 2080, 452, 1299, 414],
        'Thermal_Naphtha_kt': [434, 355, 199, 0, 285, 0, 196, 0],
        'Total_Naphtha_kt': [3596, 2947, 1698, 529, 2365, 452, 1495, 414]
    }
    
    ws_source = wb.create_sheet("source")
    source_df = pd.DataFrame(source_data)
    
    for r in dataframe_to_rows(source_df, index=False, header=True):
        ws_source.append(r)
    
    # ============================================================================
    # 3. CI SHEET (Carbon Intensity by Product and Fuel)
    # ============================================================================
    
    ci_data = {
        'Product': [
            'Ethylene', 'Propylene', 'Benzene', 'Toluene', 'Xylene', 'Butadiene',
            'Polypropylene', 'Polyethylene', 'Other_Chemicals'
        ],
        'Process': [
            'Naphtha Cracker', 'Naphtha Cracker', 'BTX Plant', 'BTX Plant', 'BTX Plant', 'Naphtha Cracker',
            'Polymerization', 'Polymerization', 'Various'
        ],
        'Naphtha_Feedstock_GJ/t': [
            134.85, 121.8, 60.9, 60.9, 60.9, 134.85,
            5.0, 5.0, 78.3
        ],
        'Naphtha_Thermal_GJ/t': [
            18.5, 16.2, 0.0, 0.0, 0.0, 18.5,
            0.0, 0.0, 5.0
        ],
        'LNG_GJ/t': [
            8.5, 7.2, 12.3, 12.3, 12.3, 8.5,
            15.2, 15.2, 10.0
        ],
        'Fuel_Gas_GJ/t': [
            15.2, 13.8, 8.5, 8.5, 8.5, 15.2,
            5.0, 5.0, 8.0
        ],
        'LPG_GJ/t': [
            2.5, 2.1, 3.2, 3.2, 3.2, 2.5,
            2.0, 2.0, 2.5
        ],
        'Fuel_Oil_GJ/t': [
            1.5, 1.2, 2.1, 2.1, 2.1, 1.5,
            1.0, 1.0, 1.5
        ],
        'Electricity_kWh/t': [
            450, 420, 380, 380, 380, 450,
            320, 320, 350
        ]
    }
    
    ws_ci = wb.create_sheet("CI")
    ci_df = pd.DataFrame(ci_data)
    
    for r in dataframe_to_rows(ci_df, index=False, header=True):
        ws_ci.append(r)
    
    # ============================================================================
    # 4. CI2 SHEET (Emission Factors)
    # ============================================================================
    
    ci2_data = {
        'Fuel_Type': [
            'Naphtha_Feedstock', 'Naphtha_Thermal', 'Bio_Naphtha_Feedstock', 'Bio_Naphtha_Thermal',
            'Recycled_Naphtha_Feedstock', 'Recycled_Naphtha_Thermal', 'LNG', 'Fuel_Gas',
            'LPG_Propane', 'LPG_Butane', 'Fuel_Oil', 'Diesel', 'Electricity_Grid'
        ],
        'Emission_Factor_tCO2/GJ': [
            0.0207, 0.0535, 0.0034, 0.0000,  # Naphtha factors
            0.0103, 0.0535, 0.0560, 0.0580,  # Alternative naphtha and gases
            0.0630, 0.0640, 0.0740, 0.0740, 0.1290  # Other fuels and electricity (tCO2/MWh)
        ],
        'Heating_Value_GJ/t': [
            43.5, 43.5, 43.5, 43.5,
            43.5, 43.5, 39.5, 45.2,
            46.1, 49.2, 40.5, 43.2, 3.6
        ],
        'Lifecycle_Emissions_tCO2/t': [
            0.90, 2.33, 0.15, 0.00,
            0.45, 2.33, 2.21, 2.62,
            2.90, 3.15, 3.00, 3.20, 0.465
        ],
        'Notes': [
            'Upstream + transport', 'Direct combustion', 'Sustainable sourcing', 'Carbon neutral combustion',
            'Recycling process', 'Same as conventional', 'Natural gas', 'Refinery gas',
            'Liquefied petroleum gas', 'Liquefied petroleum gas', 'Heavy fuel oil', 'Distillate fuel', 'Grid average'
        ]
    }
    
    ws_ci2 = wb.create_sheet("CI2")
    ci2_df = pd.DataFrame(ci2_data)
    
    for r in dataframe_to_rows(ci2_df, index=False, header=True):
        ws_ci2.append(r)
    
    # ============================================================================
    # 5. EMISSION TARGETS SHEET (Updated scenarios)
    # ============================================================================
    
    years = list(range(2025, 2051))
    targets_data = {
        'Year': years,
        'Conservative_Target_%': [5 + (i * 3.1) for i in range(26)],  # 5% to 85% by 2050
        'Moderate_Target_%': [10 + (i * 3.1) for i in range(26)],     # 10% to 90% by 2050
        'Aggressive_Target_%': [15 + (i * 3.2) for i in range(26)]    # 15% to 95% by 2050
    }
    
    # Calculate absolute emissions
    baseline_emissions = 52.0
    targets_data['Conservative_Emissions_Mt'] = [baseline_emissions * (1 - pct/100) for pct in targets_data['Conservative_Target_%']]
    targets_data['Moderate_Emissions_Mt'] = [baseline_emissions * (1 - pct/100) for pct in targets_data['Moderate_Target_%']]
    targets_data['Aggressive_Emissions_Mt'] = [baseline_emissions * (1 - pct/100) for pct in targets_data['Aggressive_Target_%']]
    
    ws_targets = wb.create_sheet("Emission_Targets")
    targets_df = pd.DataFrame(targets_data)
    
    for r in dataframe_to_rows(targets_df, index=False, header=True):
        ws_targets.append(r)
    
    # ============================================================================
    # 6. TECHNOLOGY OPTIONS SHEET (Including naphtha alternatives)
    # ============================================================================
    
    tech_data = {
        'Technology': [
            'Energy_Efficiency_NCC', 'Hydrogen_Direct_Reduction', 'Electrification_BTX', 'Renewable_Electricity',
            'Bio_Naphtha_Feedstock', 'Bio_Naphtha_Thermal', 'Recycled_Naphtha_Feedstock', 'Recycled_Naphtha_Thermal',
            'Steam_Optimization', 'Heat_Recovery', 'Process_Integration', 'Waste_Heat_Recovery'
        ],
        'Process_Application': [
            'NCC', 'NCC', 'BTX', 'All',
            'NCC', 'NCC', 'NCC', 'NCC',
            'All', 'All', 'All', 'All'
        ],
        'Max_Penetration_%': [
            15, 80, 90, 100,
            50, 60, 30, 40,
            70, 85, 60, 75
        ],
        'CAPEX_USD_per_tonne': [
            450, 1200, 800, 350,
            850, 950, 650, 750,
            300, 420, 500, 380
        ],
        'OPEX_USD_per_tonne_year': [
            25, 45, 35, 15,
            65, 70, 35, 40,
            20, 25, 30, 20
        ],
        'Emission_Reduction_%': [
            12, 75, 85, 95,
            83, 100, 50, 0,
            25, 30, 35, 20
        ],
        'Commercial_Year': [
            2025, 2028, 2026, 2025,
            2027, 2028, 2026, 2026,
            2025, 2025, 2026, 2025
        ],
        'Notes': [
            'Process optimization', 'Green hydrogen for heating', 'Electric heating/motors', 'Solar/wind power',
            'Sustainable bio-feedstock', 'Bio-naphtha for thermal', 'Recycled feedstock', 'Recycled thermal fuel',
            'Steam system efficiency', 'Heat exchanger networks', 'Process synergies', 'Waste heat utilization'
        ]
    }
    
    ws_tech = wb.create_sheet("Technology_Options")
    tech_df = pd.DataFrame(tech_data)
    
    for r in dataframe_to_rows(tech_df, index=False, header=True):
        ws_tech.append(r)
    
    # ============================================================================
    # 7. NAPHTHA ALTERNATIVES SHEET (Detailed analysis)
    # ============================================================================
    
    naphtha_alt_data = {
        'Alternative_Type': [
            'Conventional_Naphtha', 'Conventional_Naphtha', 'Bio_Naphtha', 'Bio_Naphtha',
            'Recycled_Naphtha', 'Recycled_Naphtha', 'Synthetic_Naphtha', 'Synthetic_Naphtha'
        ],
        'Application': [
            'Feedstock', 'Thermal', 'Feedstock', 'Thermal',
            'Feedstock', 'Thermal', 'Feedstock', 'Thermal'
        ],
        'Emission_Factor_tCO2/t': [
            0.90, 2.33, 0.15, 0.00,
            0.45, 2.33, 0.25, 1.20
        ],
        'Cost_Premium_%': [
            0, 0, 180, 200,
            80, 80, 150, 160
        ],
        'Availability_Mt_2030': [
            50.0, 50.0, 5.0, 5.0,
            8.0, 8.0, 2.0, 2.0
        ],
        'Availability_Mt_2040': [
            45.0, 45.0, 15.0, 15.0,
            20.0, 20.0, 8.0, 8.0
        ],
        'Availability_Mt_2050': [
            35.0, 35.0, 25.0, 25.0,
            30.0, 30.0, 15.0, 15.0
        ],
        'Technology_Readiness': [
            'Commercial', 'Commercial', 'Demonstration', 'Demonstration',
            'Pilot', 'Pilot', 'R&D', 'R&D'
        ]
    }
    
    ws_naphtha = wb.create_sheet("Naphtha_Alternatives")
    naphtha_df = pd.DataFrame(naphtha_alt_data)
    
    for r in dataframe_to_rows(naphtha_df, index=False, header=True):
        ws_naphtha.append(r)
    
    # ============================================================================
    # 8. FUEL MIX SCENARIOS SHEET (Updated with naphtha alternatives)
    # ============================================================================
    
    fuel_years = [2025, 2030, 2035, 2040, 2045, 2050]
    scenarios = ['Conservative', 'Moderate', 'Aggressive']
    
    fuel_data = []
    for year in fuel_years:
        for scenario in scenarios:
            if scenario == 'Conservative':
                bio_naphtha = min(5.0, 1.0 + (year-2025)*0.2)
                recycled_naphtha = min(8.0, 2.0 + (year-2025)*0.3)
            elif scenario == 'Moderate':
                bio_naphtha = min(15.0, 2.0 + (year-2025)*0.5)
                recycled_naphtha = min(20.0, 3.0 + (year-2025)*0.7)
            else:  # Aggressive
                bio_naphtha = min(25.0, 3.0 + (year-2025)*0.9)
                recycled_naphtha = min(30.0, 5.0 + (year-2025)*1.0)
            
            conventional_naphtha = max(5.0, 30.2 - bio_naphtha - recycled_naphtha)
            
            fuel_data.append({
                'Year': year,
                'Scenario': scenario,
                'Conventional_Naphtha_Mt': conventional_naphtha,
                'Bio_Naphtha_Mt': bio_naphtha,
                'Recycled_Naphtha_Mt': recycled_naphtha,
                'Natural_Gas_Mt': 6.5 - (year-2025)*0.1,
                'LPG_Mt': 0.8 - (year-2025)*0.02,
                'Fuel_Oil_Mt': max(0.1, 0.5 - (year-2025)*0.02),
                'Hydrogen_Mt': (year-2025)*0.1,
                'Electricity_TWh': 25 + (year-2025)*1.5
            })
    
    ws_fuel = wb.create_sheet("Fuel_Mix_Scenarios")
    fuel_df = pd.DataFrame(fuel_data)
    
    for r in dataframe_to_rows(fuel_df, index=False, header=True):
        ws_fuel.append(r)
    
    # Save file
    filename = "korean_petrochemical_scenario_configuration.xlsx"
    wb.save(filename)
    
    print(f"✅ Enhanced Excel file created: {filename}")
    print(f"📊 Sheets: {[sheet.title for sheet in wb.worksheets]}")
    
    return filename

def main():
    """Main execution"""
    
    print("🏭 KOREAN PETROCHEMICAL MACC MODEL")
    print("ENHANCED EXCEL WITH NAPHTHA ALTERNATIVES")
    print("=" * 70)
    
    filename = create_enhanced_excel_with_naphtha()
    
    print(f"\n🎯 KEY ENHANCEMENTS:")
    print(f"   • Added detailed naphtha emission intensities")
    print(f"   • Included bio-naphtha alternative (83% feedstock reduction)")
    print(f"   • Included recycled-naphtha alternative (50% feedstock reduction)")
    print(f"   • Updated CI and CI2 sheets with naphtha factors")
    print(f"   • Enhanced fuel mix scenarios with naphtha alternatives")
    print(f"   • Added naphtha alternatives availability projections")
    
    print(f"\n📋 NAPHTHA EMISSION FACTORS:")
    print(f"   • Conventional feedstock: 0.90 tCO2/tonne")
    print(f"   • Conventional thermal: 2.33 tCO2/tonne")
    print(f"   • Bio-naphtha feedstock: 0.15 tCO2/tonne")
    print(f"   • Bio-naphtha thermal: 0.00 tCO2/tonne")
    print(f"   • Recycled feedstock: 0.45 tCO2/tonne")
    print(f"   • Recycled thermal: 2.33 tCO2/tonne")

if __name__ == "__main__":
    main()