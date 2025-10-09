#!/usr/bin/env python3
"""
Recalculated Baseline with Excel Scenario Configuration
Korean Petrochemical Industry - Updated with Industry Data
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

def recalculate_baseline_from_industry_data():
    """Recalculate baseline using industry-specific emission data"""
    
    print("🔄 RECALCULATING BASELINE FROM INDUSTRY DATA")
    print("=" * 80)
    
    # Industry-provided data
    industry_data = {
        'total_emissions_mt': 52.0,           # Million tonnes CO2/year
        'ncc_naphtha_share': 0.461,           # 46.1% of total emissions
        'ncc_steam_electricity_share': 0.307, # 30.7% of total emissions
        'other_processes_share': 0.232        # 23.2% of total emissions (remaining)
    }
    
    # Calculate emissions by source
    emissions_breakdown = {
        'ncc_naphtha_combustion': industry_data['total_emissions_mt'] * industry_data['ncc_naphtha_share'],
        'ncc_steam_electricity': industry_data['total_emissions_mt'] * industry_data['ncc_steam_electricity_share'],
        'other_processes': industry_data['total_emissions_mt'] * industry_data['other_processes_share']
    }
    
    print("📊 EMISSION BREAKDOWN (Industry Data):")
    print(f"   Total Emissions: {industry_data['total_emissions_mt']:.1f} MtCO2/year")
    print(f"   NCC Naphtha Combustion: {emissions_breakdown['ncc_naphtha_combustion']:.1f} MtCO2/year ({industry_data['ncc_naphtha_share']:.1%})")
    print(f"   NCC Steam & Electricity: {emissions_breakdown['ncc_steam_electricity']:.1f} MtCO2/year ({industry_data['ncc_steam_electricity_share']:.1%})")
    print(f"   Other Processes: {emissions_breakdown['other_processes']:.1f} MtCO2/year ({industry_data['other_processes_share']:.1%})")
    
    return industry_data, emissions_breakdown

def calculate_ncc_baseline():
    """Calculate NCC-specific baseline data"""
    
    print("\n\n🏭 NCC BASELINE CALCULATION")
    print("=" * 80)
    
    industry_data, emissions_breakdown = recalculate_baseline_from_industry_data()
    
    # NCC process assumptions
    ncc_data = {
        'ethylene_capacity_mt': 7.2,          # Mt/year ethylene capacity
        'capacity_utilization': 0.85,         # 85% utilization
        'naphtha_co2_factor': 3.1,            # tCO2/tonne naphtha combusted
        'steam_electricity_intensity': 2.1    # tCO2/tonne ethylene (steam + electricity)
    }
    
    # Calculate actual ethylene production
    ethylene_production = ncc_data['ethylene_capacity_mt'] * ncc_data['capacity_utilization']
    
    # Calculate naphtha consumption from emission data
    naphtha_combustion_emissions = emissions_breakdown['ncc_naphtha_combustion']
    naphtha_consumption_mt = naphtha_combustion_emissions / ncc_data['naphtha_co2_factor']
    
    # Calculate naphtha feedstock (separate from combustion)
    # Industry standard: 3.1 tonnes naphtha feedstock per tonne ethylene
    naphtha_feedstock_ratio = 3.1
    naphtha_feedstock_mt = ethylene_production * naphtha_feedstock_ratio
    
    # Total naphtha (feedstock + fuel)
    total_naphtha_mt = naphtha_feedstock_mt + naphtha_consumption_mt
    
    print("📊 NCC Process Data:")
    print(f"   Ethylene Production: {ethylene_production:.1f} Mt/year")
    print(f"   Naphtha Feedstock: {naphtha_feedstock_mt:.1f} Mt/year")
    print(f"   Naphtha Fuel (Combustion): {naphtha_consumption_mt:.1f} Mt/year")
    print(f"   Total Naphtha: {total_naphtha_mt:.1f} Mt/year")
    print(f"   Naphtha/Ethylene Ratio: {total_naphtha_mt/ethylene_production:.1f}:1")
    
    # Steam and electricity consumption
    steam_elec_emissions = emissions_breakdown['ncc_steam_electricity']
    steam_elec_consumption = steam_elec_emissions / ncc_data['steam_electricity_intensity']
    
    print(f"   Steam & Electricity Emissions: {steam_elec_emissions:.1f} MtCO2/year")
    print(f"   Steam & Electricity Intensity: {steam_elec_emissions/ethylene_production:.1f} tCO2/tonne ethylene")
    
    ncc_baseline = {
        'ethylene_production': ethylene_production,
        'naphtha_feedstock': naphtha_feedstock_mt,
        'naphtha_fuel': naphtha_consumption_mt,
        'total_naphtha': total_naphtha_mt,
        'steam_elec_emissions': steam_elec_emissions,
        'naphtha_combustion_emissions': naphtha_combustion_emissions,
        'total_ncc_emissions': naphtha_combustion_emissions + steam_elec_emissions
    }
    
    return ncc_baseline

def calculate_other_processes_baseline():
    """Calculate baseline for non-NCC processes"""
    
    print("\n\n⚗️  OTHER PROCESSES BASELINE")
    print("=" * 80)
    
    industry_data, emissions_breakdown = recalculate_baseline_from_industry_data()
    ncc_baseline = calculate_ncc_baseline()
    
    # Other processes (BTX, Utilities, Downstream)
    other_processes = {
        'btx_capacity_mt': 4.1,               # Mt/year BTX capacity
        'utilities_capacity_mt': 2.8,         # Mt/year utilities equivalent
        'downstream_capacity_mt': 1.5,        # Mt/year downstream products
        'capacity_utilization': 0.85          # 85% utilization
    }
    
    # Calculate production
    btx_production = other_processes['btx_capacity_mt'] * other_processes['capacity_utilization']
    utilities_production = other_processes['utilities_capacity_mt'] * other_processes['capacity_utilization']
    downstream_production = other_processes['downstream_capacity_mt'] * other_processes['capacity_utilization']
    
    total_other_production = btx_production + utilities_production + downstream_production
    
    # Distribute emissions among other processes
    other_emissions = emissions_breakdown['other_processes']
    
    # Emission intensity by process (based on energy requirements)
    emission_intensities = {
        'btx': 1.8,          # tCO2/tonne (moderate energy)
        'utilities': 2.5,    # tCO2/tonne (high energy - steam, power)
        'downstream': 1.2    # tCO2/tonne (lower energy)
    }
    
    # Calculate weighted emissions
    weighted_production = (btx_production * emission_intensities['btx'] + 
                          utilities_production * emission_intensities['utilities'] +
                          downstream_production * emission_intensities['downstream'])
    
    emission_factor = other_emissions / weighted_production
    
    btx_emissions = btx_production * emission_intensities['btx'] * emission_factor
    utilities_emissions = utilities_production * emission_intensities['utilities'] * emission_factor
    downstream_emissions = downstream_production * emission_intensities['downstream'] * emission_factor
    
    print("📊 Other Processes Data:")
    print(f"   BTX Production: {btx_production:.1f} Mt/year")
    print(f"   Utilities Production: {utilities_production:.1f} Mt/year")
    print(f"   Downstream Production: {downstream_production:.1f} Mt/year")
    print(f"   Total Other Production: {total_other_production:.1f} Mt/year")
    print(f"   BTX Emissions: {btx_emissions:.1f} MtCO2/year")
    print(f"   Utilities Emissions: {utilities_emissions:.1f} MtCO2/year")
    print(f"   Downstream Emissions: {downstream_emissions:.1f} MtCO2/year")
    print(f"   Total Other Emissions: {other_emissions:.1f} MtCO2/year")
    
    other_baseline = {
        'btx_production': btx_production,
        'utilities_production': utilities_production,
        'downstream_production': downstream_production,
        'total_production': total_other_production,
        'btx_emissions': btx_emissions,
        'utilities_emissions': utilities_emissions,
        'downstream_emissions': downstream_emissions,
        'total_emissions': other_emissions
    }
    
    return other_baseline

def create_fuel_consumption_baseline():
    """Create arbitrary fuel consumption baseline as requested"""
    
    print("\n\n⛽ FUEL CONSUMPTION BASELINE (Arbitrary Values)")
    print("=" * 80)
    
    # Arbitrary fuel values as requested
    fuel_consumption = {
        'natural_gas_mt': 8.5,        # Mt/year
        'lpg_mt': 1.2,                # Mt/year  
        'fuel_oil_mt': 0.8,           # Mt/year
        'coal_mt': 0.3,               # Mt/year
        'hydrogen_mt': 0.1,           # Mt/year (minimal current use)
        'electricity_twh': 15.2       # TWh/year
    }
    
    # Emission factors
    emission_factors = {
        'natural_gas': 2.75,          # tCO2/tonne
        'lpg': 3.00,                  # tCO2/tonne
        'fuel_oil': 3.15,             # tCO2/tonne
        'coal': 2.86,                 # tCO2/tonne
        'hydrogen': 0.0,              # tCO2/tonne (if green)
        'electricity': 0.45           # tCO2/MWh (grid factor)
    }
    
    # Calculate emissions from fuels
    fuel_emissions = {}
    for fuel, consumption in fuel_consumption.items():
        if fuel == 'electricity_twh':
            emissions = consumption * 1000 * emission_factors['electricity'] / 1000  # MtCO2
        else:
            fuel_name = fuel.replace('_mt', '')
            emissions = consumption * emission_factors[fuel_name] / 1000  # MtCO2
        fuel_emissions[fuel] = emissions
    
    total_fuel_emissions = sum(fuel_emissions.values())
    
    print("📊 Fuel Consumption (Arbitrary):")
    for fuel, consumption in fuel_consumption.items():
        unit = 'TWh/year' if 'twh' in fuel else 'Mt/year'
        emissions = fuel_emissions[fuel]
        print(f"   {fuel.replace('_', ' ').title()}: {consumption:.1f} {unit} → {emissions:.1f} MtCO2/year")
    
    print(f"   Total Fuel Emissions: {total_fuel_emissions:.1f} MtCO2/year")
    
    return fuel_consumption, fuel_emissions

def create_excel_scenario_configuration():
    """Create Excel file with scenario configuration sheets"""
    
    print("\n\n📊 CREATING EXCEL SCENARIO CONFIGURATION")
    print("=" * 80)
    
    # Get baseline data
    industry_data, emissions_breakdown = recalculate_baseline_from_industry_data()
    ncc_baseline = calculate_ncc_baseline()
    other_baseline = calculate_other_processes_baseline()
    fuel_consumption, fuel_emissions = create_fuel_consumption_baseline()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. Baseline Data Sheet
    ws_baseline = wb.create_sheet("Baseline_Data")
    
    # Baseline headers and data
    baseline_headers = ['Category', 'Subcategory', 'Value', 'Unit', 'Notes']
    ws_baseline.append(baseline_headers)
    
    baseline_data = [
        ['Total Industry', 'Total Emissions', industry_data['total_emissions_mt'], 'MtCO2/year', 'Industry provided'],
        ['NCC Process', 'Ethylene Production', ncc_baseline['ethylene_production'], 'Mt/year', '85% capacity utilization'],
        ['NCC Process', 'Naphtha Feedstock', ncc_baseline['naphtha_feedstock'], 'Mt/year', '3.1:1 ratio to ethylene'],
        ['NCC Process', 'Naphtha Fuel', ncc_baseline['naphtha_fuel'], 'Mt/year', 'For combustion/heating'],
        ['NCC Process', 'Total Naphtha', ncc_baseline['total_naphtha'], 'Mt/year', 'Feedstock + fuel'],
        ['NCC Process', 'Naphtha Emissions', ncc_baseline['naphtha_combustion_emissions'], 'MtCO2/year', '46.1% of total'],
        ['NCC Process', 'Steam & Electricity Emissions', ncc_baseline['steam_elec_emissions'], 'MtCO2/year', '30.7% of total'],
        ['Other Processes', 'BTX Production', other_baseline['btx_production'], 'Mt/year', '85% capacity utilization'],
        ['Other Processes', 'Utilities Production', other_baseline['utilities_production'], 'Mt/year', 'Steam, power equivalent'],
        ['Other Processes', 'Downstream Production', other_baseline['downstream_production'], 'Mt/year', 'Polymers, chemicals'],
        ['Other Processes', 'Other Emissions', other_baseline['total_emissions'], 'MtCO2/year', '23.2% of total'],
        ['Fuel Consumption', 'Natural Gas', fuel_consumption['natural_gas_mt'], 'Mt/year', 'Arbitrary baseline'],
        ['Fuel Consumption', 'LPG', fuel_consumption['lpg_mt'], 'Mt/year', 'Arbitrary baseline'],
        ['Fuel Consumption', 'Fuel Oil', fuel_consumption['fuel_oil_mt'], 'Mt/year', 'Arbitrary baseline'],
        ['Fuel Consumption', 'Coal', fuel_consumption['coal_mt'], 'Mt/year', 'Arbitrary baseline'],
        ['Fuel Consumption', 'Hydrogen', fuel_consumption['hydrogen_mt'], 'Mt/year', 'Minimal current use'],
        ['Fuel Consumption', 'Electricity', fuel_consumption['electricity_twh'], 'TWh/year', 'Grid electricity']
    ]
    
    for row in baseline_data:
        ws_baseline.append(row)
    
    # 2. Emission Targets Sheet (3 Scenarios)
    ws_targets = wb.create_sheet("Emission_Targets")
    
    # Target headers
    target_headers = ['Year', 'Conservative_Target_%', 'Moderate_Target_%', 'Aggressive_Target_%', 
                     'Conservative_Emissions_Mt', 'Moderate_Emissions_Mt', 'Aggressive_Emissions_Mt']
    ws_targets.append(target_headers)
    
    # Years from 2025 to 2050
    years = list(range(2025, 2051))
    
    # Scenario definitions
    scenarios = {
        'conservative': {
            2025: 0.95,  # 5% reduction
            2030: 0.85,  # 15% reduction
            2035: 0.70,  # 30% reduction
            2040: 0.55,  # 45% reduction
            2045: 0.35,  # 65% reduction
            2050: 0.20   # 80% reduction
        },
        'moderate': {
            2025: 0.90,  # 10% reduction
            2030: 0.75,  # 25% reduction
            2035: 0.55,  # 45% reduction
            2040: 0.35,  # 65% reduction
            2045: 0.20,  # 80% reduction
            2050: 0.10   # 90% reduction
        },
        'aggressive': {
            2025: 0.85,  # 15% reduction
            2030: 0.65,  # 35% reduction
            2035: 0.40,  # 60% reduction
            2040: 0.20,  # 80% reduction
            2045: 0.10,  # 90% reduction
            2050: 0.05   # 95% reduction
        }
    }
    
    # Interpolate targets for all years
    baseline_emissions = industry_data['total_emissions_mt']
    
    for year in years:
        # Interpolate targets
        conservative_ratio = interpolate_target(year, scenarios['conservative'])
        moderate_ratio = interpolate_target(year, scenarios['moderate'])
        aggressive_ratio = interpolate_target(year, scenarios['aggressive'])
        
        # Calculate absolute emissions
        conservative_emissions = baseline_emissions * conservative_ratio
        moderate_emissions = baseline_emissions * moderate_ratio
        aggressive_emissions = baseline_emissions * aggressive_ratio
        
        # Percentage reductions
        conservative_pct = (1 - conservative_ratio) * 100
        moderate_pct = (1 - moderate_ratio) * 100
        aggressive_pct = (1 - aggressive_ratio) * 100
        
        ws_targets.append([
            year, 
            conservative_pct, moderate_pct, aggressive_pct,
            conservative_emissions, moderate_emissions, aggressive_emissions
        ])
    
    # 3. Technology Options Sheet
    ws_tech = wb.create_sheet("Technology_Options")
    
    tech_headers = ['Technology', 'Process_Application', 'Max_Penetration_%', 'CAPEX_USD_per_tonne', 
                   'OPEX_USD_per_tonne_year', 'Emission_Reduction_%', 'Commercial_Year', 'Notes']
    ws_tech.append(tech_headers)
    
    technologies = [
        ['NCC_Hydrogen_Retrofit', 'NCC', 80, 800, 120, 75, 2028, 'Hydrogen furnace conversion'],
        ['NCC_Energy_Efficiency', 'NCC', 15, 100, 15, 12, 2025, 'Process optimization'],
        ['BTX_Electrification', 'BTX', 60, 400, 60, 65, 2026, 'Electric heating systems'],
        ['Utility_Renewable', 'Utilities', 85, 300, 45, 80, 2025, 'Solar/wind power'],
        ['Bio_Naphtha', 'NCC', 60, 200, 150, 85, 2030, 'Sustainable feedstock'],
        ['Heat_Recovery', 'All', 40, 150, 25, 25, 2025, 'Waste heat utilization'],
        ['Green_Hydrogen_Production', 'All', 100, 1200, 200, 95, 2030, 'On-site electrolysis'],
        ['CCUS', 'All', 30, 600, 80, 90, 2035, 'Carbon capture and storage']
    ]
    
    for tech in technologies:
        ws_tech.append(tech)
    
    # 4. Fuel Mix Scenarios Sheet
    ws_fuel = wb.create_sheet("Fuel_Mix_Scenarios")
    
    fuel_headers = ['Year', 'Scenario', 'Natural_Gas_Mt', 'LPG_Mt', 'Fuel_Oil_Mt', 'Coal_Mt', 
                   'Hydrogen_Mt', 'Electricity_TWh', 'Bio_Fuels_Mt']
    ws_fuel.append(fuel_headers)
    
    # Create fuel mix projections for each scenario
    for scenario in ['Conservative', 'Moderate', 'Aggressive']:
        for year in [2025, 2030, 2035, 2040, 2045, 2050]:
            # Fuel transition rates by scenario
            if scenario == 'Conservative':
                ng_reduction = min(0.02 * (year - 2025), 0.3)
                h2_growth = min(0.05 * (year - 2025), 0.8)
            elif scenario == 'Moderate':
                ng_reduction = min(0.04 * (year - 2025), 0.6)
                h2_growth = min(0.08 * (year - 2025), 1.5)
            else:  # Aggressive
                ng_reduction = min(0.06 * (year - 2025), 0.8)
                h2_growth = min(0.12 * (year - 2025), 2.5)
            
            # Calculate fuel mix
            natural_gas = fuel_consumption['natural_gas_mt'] * (1 - ng_reduction)
            lpg = fuel_consumption['lpg_mt'] * (1 - ng_reduction * 0.5)
            fuel_oil = fuel_consumption['fuel_oil_mt'] * (1 - ng_reduction * 0.8)
            coal = fuel_consumption['coal_mt'] * (1 - ng_reduction * 1.2)
            hydrogen = fuel_consumption['hydrogen_mt'] + h2_growth
            electricity = fuel_consumption['electricity_twh'] * (1 + h2_growth * 0.3)
            bio_fuels = h2_growth * 0.5
            
            ws_fuel.append([year, scenario, natural_gas, lpg, fuel_oil, coal, hydrogen, electricity, bio_fuels])
    
    # Format Excel sheets
    format_excel_sheets(wb)
    
    # Save Excel file
    filename = 'korean_petrochemical_scenario_configuration.xlsx'
    wb.save(filename)
    
    print(f"✅ Excel scenario configuration created: {filename}")
    print("📋 Sheets created:")
    print("   • Baseline_Data - Industry baseline with corrected emission data")
    print("   • Emission_Targets - 3 scenarios with year-by-year targets")
    print("   • Technology_Options - Available technologies with specifications")
    print("   • Fuel_Mix_Scenarios - Fuel transition pathways by scenario")
    
    return filename

def interpolate_target(year, scenario_dict):
    """Interpolate emission target for any year"""
    
    # Find surrounding years
    years = sorted(scenario_dict.keys())
    
    if year <= years[0]:
        return scenario_dict[years[0]]
    if year >= years[-1]:
        return scenario_dict[years[-1]]
    
    # Find bracketing years
    for i in range(len(years) - 1):
        if years[i] <= year <= years[i + 1]:
            # Linear interpolation
            y1, y2 = scenario_dict[years[i]], scenario_dict[years[i + 1]]
            x1, x2 = years[i], years[i + 1]
            return y1 + (y2 - y1) * (year - x1) / (x2 - x1)
    
    return 1.0  # Default baseline

def format_excel_sheets(workbook):
    """Format Excel sheets for readability"""
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Format each sheet
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        
        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

def create_summary_visualization():
    """Create visualization of recalculated baseline"""
    
    print("\n\n📊 CREATING BASELINE VISUALIZATION")
    print("=" * 80)
    
    # Get data
    industry_data, emissions_breakdown = recalculate_baseline_from_industry_data()
    ncc_baseline = calculate_ncc_baseline()
    other_baseline = calculate_other_processes_baseline()
    
    # Create figure
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    
    # 1. Emission Breakdown
    labels = ['NCC Naphtha\nCombustion\n(46.1%)', 'NCC Steam &\nElectricity\n(30.7%)', 'Other\nProcesses\n(23.2%)']
    sizes = [emissions_breakdown['ncc_naphtha_combustion'], 
             emissions_breakdown['ncc_steam_electricity'],
             emissions_breakdown['other_processes']]
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1']
    
    wedges, texts, autotexts = ax1.pie(sizes, labels=labels, autopct='%1.1f\nMt', colors=colors, startangle=90)
    ax1.set_title('Total Emissions: 52.0 MtCO₂/year\n(Industry Data)', fontweight='bold')
    
    # 2. NCC Process Breakdown
    ncc_labels = ['Naphtha\nFeedstock', 'Naphtha\nFuel', 'Steam &\nElectricity']
    ncc_values = [ncc_baseline['naphtha_feedstock'], ncc_baseline['naphtha_fuel'], 
                  ncc_baseline['steam_elec_emissions']/2.1]  # Convert emissions to equivalent
    ncc_colors = ['#FFB6C1', '#FF6B6B', '#87CEEB']
    
    bars1 = ax2.bar(ncc_labels, ncc_values, color=ncc_colors, alpha=0.8)
    ax2.set_ylabel('Amount (Mt/year)')
    ax2.set_title('NCC Process Components')
    
    # Add value labels
    for bar, value in zip(bars1, ncc_values):
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                f'{value:.1f}', ha='center', va='bottom', fontweight='bold')
    
    # 3. Production Overview
    production_categories = ['Ethylene\n(NCC)', 'BTX', 'Utilities', 'Downstream']
    production_values = [ncc_baseline['ethylene_production'], other_baseline['btx_production'],
                        other_baseline['utilities_production'], other_baseline['downstream_production']]
    prod_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4']
    
    bars2 = ax3.bar(production_categories, production_values, color=prod_colors, alpha=0.8)
    ax3.set_ylabel('Production (Mt/year)')
    ax3.set_title('Production by Process')
    
    # Add value labels
    for bar, value in zip(bars2, production_values):
        height = bar.get_height()
        ax3.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                f'{value:.1f}', ha='center', va='bottom', fontweight='bold')
    
    # 4. Key Ratios
    ratios_labels = ['Naphtha/Ethylene\nRatio', 'NCC Emission\nIntensity', 'Other Process\nIntensity']
    ratios_values = [ncc_baseline['total_naphtha']/ncc_baseline['ethylene_production'],
                    (ncc_baseline['naphtha_combustion_emissions'] + ncc_baseline['steam_elec_emissions'])/ncc_baseline['ethylene_production'],
                    other_baseline['total_emissions']/other_baseline['total_production']]
    ratio_colors = ['#DDA0DD', '#F0E68C', '#98FB98']
    
    bars3 = ax4.bar(ratios_labels, ratios_values, color=ratio_colors, alpha=0.8)
    ax4.set_ylabel('Ratio (tonnes/tonne or tCO₂/tonne)')
    ax4.set_title('Key Process Ratios')
    
    # Add value labels
    for bar, value in zip(bars3, ratios_values):
        height = bar.get_height()
        ax4.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                f'{value:.1f}', ha='center', va='bottom', fontweight='bold')
    
    plt.suptitle('Korean Petrochemical Industry: Recalculated Baseline (52 MtCO₂/year)', 
                 fontsize=16, fontweight='bold')
    plt.tight_layout()
    plt.savefig('recalculated_baseline_analysis.png', dpi=300, bbox_inches='tight')
    plt.show()
    
    print("✅ Visualization saved: recalculated_baseline_analysis.png")

def main():
    """Main execution function"""
    
    print("🏭 KOREAN PETROCHEMICAL INDUSTRY")
    print("RECALCULATED BASELINE WITH SCENARIO CONFIGURATION")
    print("=" * 80)
    print(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("Base Year: 2024 (Industry Data: 52 MtCO₂/year)")
    print("=" * 80)
    
    # Run all calculations
    industry_data, emissions_breakdown = recalculate_baseline_from_industry_data()
    ncc_baseline = calculate_ncc_baseline()
    other_baseline = calculate_other_processes_baseline()
    fuel_consumption, fuel_emissions = create_fuel_consumption_baseline()
    
    # Create Excel configuration
    excel_filename = create_excel_scenario_configuration()
    
    # Create visualization
    create_summary_visualization()
    
    print(f"\n✅ RECALCULATED BASELINE COMPLETE")
    print("📊 Key Corrections Made:")
    print(f"   • Total emissions corrected to: {industry_data['total_emissions_mt']:.1f} MtCO₂/year")
    print(f"   • NCC naphtha internal: {ncc_baseline['naphtha_combustion_emissions']:.1f} MtCO₂/year (46.1%)")
    print(f"   • NCC steam/electricity: {ncc_baseline['steam_elec_emissions']:.1f} MtCO₂/year (30.7%)")
    print(f"   • Naphtha only in NCC: {ncc_baseline['total_naphtha']:.1f} Mt/year total")
    
    print(f"\n📁 Files Generated:")
    print(f"   • {excel_filename} - Scenario configuration (editable)")
    print("   • recalculated_baseline_analysis.png - Baseline visualization")
    
    print(f"\n🎯 Excel Configuration Features:")
    print("   • Baseline_Data: Industry-validated baseline data")
    print("   • Emission_Targets: 3 scenarios, editable by year (2025-2050)")
    print("   • Technology_Options: Technology specifications and costs")
    print("   • Fuel_Mix_Scenarios: Fuel transition pathways")

if __name__ == "__main__":
    main()