#!/usr/bin/env python3
"""
Update Excel MACC Model with Market Growth Rate Projections
Adds capacity growth scenario data to the Korean Petrochemical MACC Excel model
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

class ExcelGrowthScenarioUpdater:
    def __init__(self, excel_path="../data/Korean_Petrochemical_MACC_Model_English.xlsx"):
        """Initialize Excel updater with growth scenarios"""
        self.excel_path = excel_path
        self.output_path = excel_path.replace('.xlsx', '_with_Growth_Scenarios.xlsx')

        # Load the enhanced transition results
        self.results_path = "../outputs/enhanced_transition_analysis/capacity_growth_scenario_results.csv"
        self.definitions_path = "../outputs/enhanced_transition_analysis/scenario_definitions.csv"

        print("📊 EXCEL MACC MODEL GROWTH SCENARIO UPDATER")
        print("=" * 80)
        print(f"Input Excel: {self.excel_path}")
        print(f"Output Excel: {self.output_path}")
        print()

    def load_growth_scenario_data(self):
        """Load growth scenario results and definitions"""
        print("📁 Loading growth scenario data...")

        try:
            # Load results
            self.results_df = pd.read_csv(self.results_path)
            self.definitions_df = pd.read_csv(self.definitions_path)

            print(f"✅ Loaded {len(self.results_df)} result records")
            print(f"✅ Loaded {len(self.definitions_df)} scenarios")

            # Parse scenarios and years
            self.scenarios = self.results_df['Scenario'].unique()
            self.years = sorted(self.results_df['Year'].unique())

            print(f"   Scenarios: {', '.join(self.scenarios)}")
            print(f"   Years: {', '.join(map(str, self.years))}")

        except FileNotFoundError:
            print("❌ Growth scenario results not found. Please run enhanced_transition_with_growth.py first.")
            raise

        return self

    def prepare_growth_rate_matrices(self):
        """Prepare growth rate data matrices for Excel"""
        print("\n📈 Preparing growth rate matrices...")

        # Base year capacity from original data
        base_df = pd.read_excel(self.excel_path, sheet_name='source_Original')
        base_capacity = base_df['capacity_1000_t'].sum()

        print(f"   Base capacity (2025): {base_capacity:,.0f} kt/year")

        # Create growth rate matrices
        self.growth_matrices = {}

        for scenario in self.scenarios:
            scenario_data = self.results_df[self.results_df['Scenario'] == scenario].copy()
            scenario_data = scenario_data.sort_values('Year')

            # Calculate annual growth rates
            growth_rates = []
            prev_capacity = base_capacity

            for _, row in scenario_data.iterrows():
                year = row['Year']
                capacity = row['Total_Capacity_kt']

                # Calculate compound annual growth rate from base year
                years_elapsed = year - 2025
                if years_elapsed > 0:
                    annual_rate = ((capacity / base_capacity) ** (1/years_elapsed)) - 1
                else:
                    annual_rate = 0

                growth_rates.append({
                    'Year': year,
                    'Total_Capacity_kt': capacity,
                    'Capacity_Growth_Factor': capacity / base_capacity,
                    'CAGR_from_2025': annual_rate,
                    'Annual_Growth_Rate': (capacity / prev_capacity) ** (1/(year - (scenario_data.iloc[max(0, len(growth_rates)-1)]['Year'] if len(growth_rates) > 0 else 2025))) - 1 if len(growth_rates) > 0 else 0
                })
                prev_capacity = capacity

            self.growth_matrices[scenario] = pd.DataFrame(growth_rates)

            # Print scenario summary
            final_rate = growth_rates[-1]['CAGR_from_2025'] if growth_rates else 0
            final_capacity = growth_rates[-1]['Total_Capacity_kt'] if growth_rates else base_capacity

            print(f"   {scenario}:")
            print(f"     2050 capacity: {final_capacity:,.0f} kt/year")
            print(f"     CAGR (2025-2050): {final_rate*100:+.1f}% annual")

        return self

    def create_market_projection_sheets(self):
        """Create detailed market projection sheets"""
        print("\n📋 Creating market projection sheets...")

        # Load the existing workbook
        wb = openpyxl.load_workbook(self.excel_path)

        # 1. Growth Scenario Overview Sheet
        if 'Growth_Scenarios' in wb.sheetnames:
            wb.remove(wb['Growth_Scenarios'])
        ws_overview = wb.create_sheet('Growth_Scenarios', 0)

        # Add title
        ws_overview['A1'] = "Korean Petrochemical Industry - Market Growth Scenarios"
        ws_overview['A1'].font = Font(size=16, bold=True)
        ws_overview.merge_cells('A1:H1')

        # Add scenario definitions
        ws_overview['A3'] = "Scenario Definitions"
        ws_overview['A3'].font = Font(size=14, bold=True)

        headers = ['Scenario', 'Description', 'Annual Growth Rate', 'Key Drivers']
        for col, header in enumerate(headers, 1):
            cell = ws_overview.cell(4, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Add scenario data
        for row_idx, (_, scenario) in enumerate(self.definitions_df.iterrows(), 5):
            ws_overview.cell(row_idx, 1, scenario['Scenario'].replace('_', ' '))
            ws_overview.cell(row_idx, 2, scenario['Description'])
            ws_overview.cell(row_idx, 3, f"{scenario['Overall_Growth_Rate']*100:+.1f}%")
            ws_overview.cell(row_idx, 4, scenario['Key_Drivers'])

        # Add capacity projections summary
        start_row = len(self.definitions_df) + 7
        ws_overview.cell(start_row, 1, "Capacity Projections Summary (kt/year)").font = Font(size=14, bold=True)

        # Headers for capacity table
        capacity_headers = ['Year'] + [s.replace('_', ' ') for s in self.scenarios]
        for col, header in enumerate(capacity_headers, 1):
            cell = ws_overview.cell(start_row + 1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Add capacity data by year
        for row_idx, year in enumerate(self.years, start_row + 2):
            ws_overview.cell(row_idx, 1, year)
            for col_idx, scenario in enumerate(self.scenarios, 2):
                scenario_data = self.results_df[
                    (self.results_df['Scenario'] == scenario) &
                    (self.results_df['Year'] == year)
                ]
                if not scenario_data.empty:
                    capacity = scenario_data.iloc[0]['Total_Capacity_kt']
                    ws_overview.cell(row_idx, col_idx, f"{capacity:,.0f}")

        # 2. Detailed Growth Rates Sheet
        if 'Growth_Rates_Detail' in wb.sheetnames:
            wb.remove(wb['Growth_Rates_Detail'])
        ws_detail = wb.create_sheet('Growth_Rates_Detail')

        # Add title
        ws_detail['A1'] = "Detailed Growth Rate Analysis"
        ws_detail['A1'].font = Font(size=16, bold=True)

        # Add detailed growth rate data for each scenario
        current_row = 3
        for scenario in self.scenarios:
            # Scenario header
            ws_detail.cell(current_row, 1, f"{scenario.replace('_', ' ')} Growth Analysis").font = Font(size=12, bold=True)
            current_row += 1

            # Headers
            detail_headers = ['Year', 'Capacity (kt)', 'Growth Factor', 'CAGR from 2025', 'Baseline Emissions (Mt)', 'Abatement (Mt)', 'Investment ($M)']
            for col, header in enumerate(detail_headers, 1):
                cell = ws_detail.cell(current_row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            current_row += 1

            # Add data
            scenario_results = self.results_df[self.results_df['Scenario'] == scenario].sort_values('Year')
            base_capacity = 100066  # From the model output

            for _, row in scenario_results.iterrows():
                year = row['Year']
                capacity = row['Total_Capacity_kt']
                growth_factor = capacity / base_capacity
                years_elapsed = year - 2025
                cagr = ((capacity / base_capacity) ** (1/years_elapsed)) - 1 if years_elapsed > 0 else 0

                ws_detail.cell(current_row, 1, year)
                ws_detail.cell(current_row, 2, f"{capacity:,.0f}")
                ws_detail.cell(current_row, 3, f"{growth_factor:.2f}x")
                ws_detail.cell(current_row, 4, f"{cagr*100:+.1f}%")
                ws_detail.cell(current_row, 5, f"{row['Baseline_Emissions_Mt']:.1f}")
                ws_detail.cell(current_row, 6, f"{row['Abatement_Mt']:.1f}")
                ws_detail.cell(current_row, 7, f"{row['CAPEX_M_USD']:.0f}")
                current_row += 1

            current_row += 2  # Space between scenarios

        # 3. MACC Results with Growth Sheet
        if 'MACC_Growth_Results' in wb.sheetnames:
            wb.remove(wb['MACC_Growth_Results'])
        ws_macc = wb.create_sheet('MACC_Growth_Results')

        # Add title
        ws_macc['A1'] = "MACC Results by Growth Scenario"
        ws_macc['A1'].font = Font(size=16, bold=True)

        # Create comprehensive MACC table
        ws_macc['A3'] = "Complete MACC Analysis Results"
        ws_macc['A3'].font = Font(size=14, bold=True)

        # Add all results data
        macc_headers = list(self.results_df.columns)
        for col, header in enumerate(macc_headers, 1):
            cell = ws_macc.cell(4, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Add all data
        for row_idx, (_, row) in enumerate(self.results_df.iterrows(), 5):
            for col_idx, value in enumerate(row, 1):
                if isinstance(value, (int, float)):
                    if col_idx in [3, 4, 5, 6]:  # Capacity and emission columns
                        ws_macc.cell(row_idx, col_idx, f"{value:.1f}")
                    elif col_idx in [8, 9, 10, 11]:  # Cost columns
                        ws_macc.cell(row_idx, col_idx, f"{value:.0f}")
                    else:
                        ws_macc.cell(row_idx, col_idx, f"{value:.2f}")
                else:
                    ws_macc.cell(row_idx, col_idx, str(value))

        # 4. Update existing MACC template with growth factors
        if 'MACC_Template_with_Growth' in wb.sheetnames:
            wb.remove(wb['MACC_Template_with_Growth'])

        # Copy original MACC template
        original_macc = wb['MACC_Template_Corrected']
        ws_new_macc = wb.copy_worksheet(original_macc)
        ws_new_macc.title = 'MACC_Template_with_Growth'

        # Add growth scenario columns
        max_col = ws_new_macc.max_column
        growth_start_col = max_col + 2

        # Add headers
        ws_new_macc.cell(1, growth_start_col, "Growth Scenario Impact Analysis").font = Font(size=12, bold=True)

        scenario_headers = []
        for scenario in self.scenarios:
            scenario_headers.extend([
                f"{scenario}_2030_Factor",
                f"{scenario}_2040_Factor",
                f"{scenario}_2050_Factor"
            ])

        for col_idx, header in enumerate(scenario_headers, growth_start_col):
            cell = ws_new_macc.cell(2, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Add growth factor data (simplified for demonstration)
        base_capacity = 100066
        for row in range(3, min(50, ws_new_macc.max_row + 1)):  # First 47 data rows
            col_idx = growth_start_col
            for scenario in self.scenarios:
                # Get capacity factors for this scenario
                scenario_data = self.results_df[self.results_df['Scenario'] == scenario]

                for year in [2030, 2040, 2050]:
                    year_data = scenario_data[scenario_data['Year'] == year]
                    if not year_data.empty:
                        capacity_factor = year_data.iloc[0]['Total_Capacity_kt'] / base_capacity
                        ws_new_macc.cell(row, col_idx, f"{capacity_factor:.3f}")
                    else:
                        ws_new_macc.cell(row, col_idx, "1.000")
                    col_idx += 1

        # Auto-adjust column widths
        for ws in [ws_overview, ws_detail, ws_macc, ws_new_macc]:
            for col_num in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_num)
                for row_num in range(1, min(ws.max_row + 1, 100)):  # Limit to first 100 rows for performance
                    try:
                        cell = ws.cell(row_num, col_num)
                        if cell.value and hasattr(cell, 'column_letter'):
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 30)  # Min 10, max 30
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save the updated workbook
        wb.save(self.output_path)
        print(f"✅ Updated Excel saved: {self.output_path}")

        return self.output_path

    def create_growth_summary_report(self):
        """Create a summary report of growth scenarios"""
        print("\n📋 Creating growth scenario summary report...")

        summary_data = []
        base_capacity = 100066  # kt/year

        for scenario in self.scenarios:
            scenario_def = self.definitions_df[self.definitions_df['Scenario'] == scenario].iloc[0]
            scenario_results_2050 = self.results_df[
                (self.results_df['Scenario'] == scenario) &
                (self.results_df['Year'] == 2050)
            ].iloc[0]

            # Calculate 25-year CAGR
            final_capacity = scenario_results_2050['Total_Capacity_kt']
            cagr_25yr = ((final_capacity / base_capacity) ** (1/25)) - 1

            summary_data.append({
                'Scenario': scenario.replace('_', ' '),
                'Description': scenario_def['Description'],
                'Annual_Growth_Rate': f"{scenario_def['Overall_Growth_Rate']*100:+.1f}%",
                'Base_Capacity_2025_kt': f"{base_capacity:,.0f}",
                'Final_Capacity_2050_kt': f"{final_capacity:,.0f}",
                'Total_Growth_Factor': f"{final_capacity/base_capacity:.2f}x",
                'CAGR_2025_2050': f"{cagr_25yr*100:+.1f}%",
                'Baseline_Emissions_2050_Mt': f"{scenario_results_2050['Baseline_Emissions_Mt']:.1f}",
                'Abatement_Potential_2050_Mt': f"{scenario_results_2050['Abatement_Mt']:.1f}",
                'Investment_Required_2050_M_USD': f"{scenario_results_2050['CAPEX_M_USD']:.0f}",
                'Abatement_Cost_USD_per_tCO2': f"{scenario_results_2050['Abatement_Cost_USD_per_tCO2']:.0f}",
                'Key_Drivers': scenario_def['Key_Drivers']
            })

        summary_df = pd.DataFrame(summary_data)
        summary_path = "../outputs/enhanced_transition_analysis/growth_scenario_summary.csv"
        summary_df.to_csv(summary_path, index=False)

        print(f"✅ Growth scenario summary saved: {summary_path}")
        return summary_path

    def print_update_summary(self):
        """Print summary of Excel updates"""
        print(f"\n{'='*80}")
        print("📊 EXCEL GROWTH SCENARIO UPDATE SUMMARY")
        print(f"{'='*80}")

        print(f"\n✅ SHEETS ADDED TO EXCEL:")
        print(f"   • Growth_Scenarios: Overview and capacity projections")
        print(f"   • Growth_Rates_Detail: Detailed growth rate analysis")
        print(f"   • MACC_Growth_Results: Complete MACC results with growth")
        print(f"   • MACC_Template_with_Growth: Enhanced MACC template with growth factors")

        print(f"\n📊 SCENARIO COVERAGE:")
        for scenario in self.scenarios:
            scenario_def = self.definitions_df[self.definitions_df['Scenario'] == scenario].iloc[0]
            print(f"   • {scenario.replace('_', ' ')}: {scenario_def['Overall_Growth_Rate']*100:+.1f}% annual")

        print(f"\n📅 TIME HORIZON:")
        print(f"   • Years covered: {', '.join(map(str, self.years))}")
        print(f"   • Growth factors calculated for each year")
        print(f"   • CAGR analysis from base year (2025)")

        print(f"\n💼 BUSINESS IMPACT METRICS INCLUDED:")
        print(f"   • Capacity projections by scenario and year")
        print(f"   • Baseline emission trajectories")
        print(f"   • Abatement potential scaling")
        print(f"   • Investment requirements")
        print(f"   • Technology-specific cost analysis")

        print(f"\n🎯 KEY APPLICATIONS:")
        print(f"   • Strategic planning with different market assumptions")
        print(f"   • Investment scenario analysis")
        print(f"   • Policy impact assessment")
        print(f"   • Risk analysis under different growth trajectories")

        print(f"\n📁 FILES CREATED:")
        print(f"   • Updated Excel: {self.output_path}")
        print(f"   • Growth summary: ../outputs/enhanced_transition_analysis/growth_scenario_summary.csv")

    def run_excel_update(self):
        """Run complete Excel update with growth scenarios"""
        print("🚀 Starting Excel MACC model update with growth scenarios...")

        # Execute all steps
        self.load_growth_scenario_data()
        self.prepare_growth_rate_matrices()
        excel_path = self.create_market_projection_sheets()
        summary_path = self.create_growth_summary_report()

        # Print summary
        self.print_update_summary()

        print(f"\n✅ EXCEL UPDATE COMPLETE!")
        print(f"📊 Enhanced Excel: {excel_path}")
        print(f"📋 Summary report: {summary_path}")
        print(f"\n🔄 Ready for strategic analysis and decision-making")

        return {
            'excel_path': excel_path,
            'summary_path': summary_path,
            'scenarios': self.scenarios,
            'years': self.years
        }

if __name__ == "__main__":
    # Run the complete Excel update
    updater = ExcelGrowthScenarioUpdater()
    results = updater.run_excel_update()