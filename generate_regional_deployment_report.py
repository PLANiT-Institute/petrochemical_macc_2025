"""
Regional Deployment Report Generator
Shows technology deployment, energy demand, and emission pathways by region
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Paths
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUTS_DIR = BASE_DIR / "outputs"
REPORTS_DIR = BASE_DIR / "reports"

REPORTS_DIR.mkdir(exist_ok=True)

# Style definitions
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def load_data():
    """Load all necessary data"""
    data = {}

    # Facility data
    data['facilities'] = pd.read_csv(DATA_DIR / "facility_database_with_regions.csv")

    # Technology parameters
    data['tech_params'] = pd.read_csv(DATA_DIR / "technology_parameters.csv")

    # Energy intensities
    data['energy_intensity'] = pd.read_csv(DATA_DIR / "energy_intensities.csv")

    # Price trajectories
    data['h2_prices'] = pd.read_csv(DATA_DIR / "h2_price_trajectory.csv")
    data['re_prices'] = pd.read_csv(DATA_DIR / "re_price_trajectory.csv")
    data['grid_ef'] = pd.read_csv(DATA_DIR / "grid_emission_trajectory.csv")

    # Scenario summary
    data['scenarios'] = pd.read_csv(OUTPUTS_DIR / "scenario_summary_final.csv")

    # Load scenario-specific data
    data['scenario_data'] = {}
    scenario_ids = [
        'shaheen_ncc_h2', 'shaheen_ncc_electricity',
        'restructure_25pct_ncc_h2', 'restructure_25pct_ncc_electricity',
        'restructure_40pct_ncc_h2', 'restructure_40pct_ncc_electricity'
    ]

    for scenario_id in scenario_ids:
        scenario_dir = OUTPUTS_DIR / f'scenario_{scenario_id}'
        if scenario_dir.exists():
            data['scenario_data'][scenario_id] = {
                'facilities': pd.read_csv(scenario_dir / 'scenario_facilities.csv'),
                'emissions': pd.read_csv(scenario_dir / 'facility_emissions_2050.csv'),
                'deployment': pd.read_csv(scenario_dir / 'deployment_trajectory.csv'),
            }

    return data


def calculate_regional_emissions(facilities, energy_intensity, grid_ef_2025=0.459):
    """Calculate emissions by region using energy intensity data"""
    # The energy_intensity file has detailed per-facility data
    # Columns: Naphtha_GJ_per_tonne, Electricity_kWh_per_tonne, LNG_GJ_per_tonne, etc.

    # Calculate total fuel GJ per tonne (sum of all fuel types)
    energy_intensity['total_fuel_gj_per_ton'] = (
        energy_intensity['Naphtha_GJ_per_tonne'].fillna(0) +
        energy_intensity['LNG_GJ_per_tonne'].fillna(0) +
        energy_intensity['Fuel_Gas_GJ_per_tonne'].fillna(0) +
        energy_intensity['Byproduct_Gas_GJ_per_tonne'].fillna(0) +
        energy_intensity['LPG_GJ_per_tonne'].fillna(0) +
        energy_intensity['Fuel_Oil_GJ_per_tonne'].fillna(0) +
        energy_intensity['Diesel_GJ_per_tonne'].fillna(0)
    )

    # Calculate electricity MWh per tonne (convert from kWh)
    energy_intensity['electricity_mwh_per_ton'] = (
        energy_intensity['Electricity_kWh_per_tonne'].fillna(0) / 1000
    )

    # Aggregate by process type for simpler regional analysis
    process_energy = energy_intensity.groupby('process').agg({
        'total_fuel_gj_per_ton': 'mean',
        'electricity_mwh_per_ton': 'mean'
    }).reset_index()

    # Merge facilities with process-level energy intensity
    fac_with_energy = facilities.merge(
        process_energy,
        on='process',
        how='left'
    )

    # Fill NaN with defaults
    fac_with_energy['total_fuel_gj_per_ton'] = fac_with_energy['total_fuel_gj_per_ton'].fillna(5.0)
    fac_with_energy['electricity_mwh_per_ton'] = fac_with_energy['electricity_mwh_per_ton'].fillna(0.5)

    # Calculate emissions (kt CO2)
    # Fuel emissions: GJ * 0.0561 tCO2/GJ (natural gas equivalent)
    fac_with_energy['fuel_emissions_kt'] = (
        fac_with_energy['capacity_kt'] * 1000 *
        fac_with_energy['total_fuel_gj_per_ton'] * 0.0561 / 1000
    )

    # Electricity emissions: MWh * grid_ef
    fac_with_energy['elec_emissions_kt'] = (
        fac_with_energy['capacity_kt'] * 1000 *
        fac_with_energy['electricity_mwh_per_ton'] * grid_ef_2025 / 1000
    )

    # Process emissions (assume 0.1 tCO2/ton for all processes)
    fac_with_energy['process_emissions_kt'] = (
        fac_with_energy['capacity_kt'] * 1000 * 0.1 / 1000
    )

    # Total
    fac_with_energy['total_emissions_kt'] = (
        fac_with_energy['fuel_emissions_kt'] +
        fac_with_energy['elec_emissions_kt'] +
        fac_with_energy['process_emissions_kt']
    )

    return fac_with_energy


def create_overview_sheet(wb, data):
    """Create overview sheet"""
    ws = wb.create_sheet("1. Overview")

    ws['A1'] = "Regional Technology Deployment Report"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = "Korea Petrochemical Net Zero Transition - By Region"
    ws['A2'].font = Font(italic=True, size=10)

    row = 4

    # About this report
    ws.cell(row=row, column=1, value="About This Report").font = SECTION_FONT
    row += 2

    about_text = [
        "This report provides regional breakdown of:",
        "1. Technology deployment schedules by region",
        "2. Energy demand changes (electricity, hydrogen) by region",
        "3. Emission reduction pathways by region",
        "4. Total national summaries for all 6 scenarios",
        "",
        "Residual 2050 Emissions Explanation:",
        "The ~3.9 Mt residual emissions in 2050 come from:",
        "- BTX Plants: Not covered by NCC-H2 or NCC-Electricity technologies",
        "- Utility processes: Only partially covered by Heat Pump",
        "- High-temperature heat: RDH technology not deployed in scenarios",
        "",
        "To achieve true Net Zero, additional technologies would be needed for:",
        "- BTX plant decarbonization",
        "- High-temperature heat electrification",
        "- Carbon capture for process emissions",
    ]

    for text in about_text:
        ws.cell(row=row, column=1, value=text)
        row += 1

    row += 2

    # Scenario summary
    ws.cell(row=row, column=1, value="6 Scenario Summary").font = SECTION_FONT
    row += 2

    headers = ["Scenario", "Technology", "Facilities", "NCC", "BAU 2050 (Mt)", "Net 2050 (Mt)", "CAPEX ($B)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for _, s in data['scenarios'].iterrows():
        ws.cell(row=row, column=1, value=s['scenario'])
        ws.cell(row=row, column=2, value=s['technology'])
        ws.cell(row=row, column=3, value=s['n_facilities'])
        ws.cell(row=row, column=4, value=s['n_ncc_facilities'])
        ws.cell(row=row, column=5, value=f"{s['bau_2050_mt']:.1f}")
        ws.cell(row=row, column=6, value=f"{s['net_2050_mt']:.1f}")
        ws.cell(row=row, column=7, value=f"${s['capex_billion_usd']:.1f}")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12

    return wb


def create_regional_facilities_sheet(wb, data):
    """Create regional facilities breakdown sheet"""
    ws = wb.create_sheet("2. Regional Facilities")

    ws['A1'] = "Facility Distribution by Region"
    ws['A1'].font = TITLE_FONT

    facilities = data['facilities']

    row = 3

    # Overall regional summary
    ws.cell(row=row, column=1, value="1. Regional Capacity Summary (Baseline)").font = SECTION_FONT
    row += 2

    regional = facilities.groupby('location').agg({
        'capacity_kt': 'sum',
        'product': 'count'
    }).reset_index()
    regional.columns = ['Region', 'Total Capacity (kt)', 'Facility Count']
    regional = regional.sort_values('Total Capacity (kt)', ascending=False)

    # Calculate NCC capacity per region
    ncc_regional = facilities[facilities['process'] == 'Naphtha Cracker'].groupby('location')['capacity_kt'].sum()
    regional['NCC Capacity (kt)'] = regional['Region'].map(ncc_regional).fillna(0)
    regional['Non-NCC Capacity (kt)'] = regional['Total Capacity (kt)'] - regional['NCC Capacity (kt)']

    headers = ["Region", "Total Capacity (kt)", "Facility Count", "NCC Capacity (kt)", "Non-NCC Capacity (kt)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for _, r in regional.iterrows():
        ws.cell(row=row, column=1, value=r['Region'])
        ws.cell(row=row, column=2, value=f"{r['Total Capacity (kt)']:,.0f}")
        ws.cell(row=row, column=3, value=r['Facility Count'])
        ws.cell(row=row, column=4, value=f"{r['NCC Capacity (kt)']:,.0f}")
        ws.cell(row=row, column=5, value=f"{r['Non-NCC Capacity (kt)']:,.0f}")
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{regional['Total Capacity (kt)'].sum():,.0f}").font = Font(bold=True)
    ws.cell(row=row, column=3, value=regional['Facility Count'].sum()).font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"{regional['NCC Capacity (kt)'].sum():,.0f}").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"{regional['Non-NCC Capacity (kt)'].sum():,.0f}").font = Font(bold=True)

    row += 3

    # Process breakdown by region
    ws.cell(row=row, column=1, value="2. Process Type by Region").font = SECTION_FONT
    row += 2

    process_regional = facilities.pivot_table(
        index='location',
        columns='process',
        values='capacity_kt',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    headers = ['Region'] + list(process_regional.columns[1:])
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for _, r in process_regional.iterrows():
        ws.cell(row=row, column=1, value=r['location'])
        for j, col in enumerate(process_regional.columns[1:], 2):
            ws.cell(row=row, column=j, value=f"{r[col]:,.0f}" if r[col] > 0 else "-")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    return wb


def create_regional_emissions_sheet(wb, data):
    """Create regional emissions breakdown sheet"""
    ws = wb.create_sheet("3. Regional Emissions")

    ws['A1'] = "Emission Distribution by Region"
    ws['A1'].font = TITLE_FONT

    facilities = data['facilities']
    energy_intensity = data['energy_intensity']

    # Calculate regional emissions
    fac_emissions = calculate_regional_emissions(facilities, energy_intensity)

    row = 3

    # Regional emissions summary
    ws.cell(row=row, column=1, value="1. BAU 2025 Emissions by Region (kt CO2)").font = SECTION_FONT
    row += 2

    regional_emissions = fac_emissions.groupby('location').agg({
        'fuel_emissions_kt': 'sum',
        'elec_emissions_kt': 'sum',
        'process_emissions_kt': 'sum',
        'total_emissions_kt': 'sum'
    }).reset_index()
    regional_emissions = regional_emissions.sort_values('total_emissions_kt', ascending=False)

    headers = ["Region", "Fuel Emissions", "Electricity Emissions", "Process Emissions", "Total Emissions", "Share (%)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    total_emissions = regional_emissions['total_emissions_kt'].sum()

    for _, r in regional_emissions.iterrows():
        share = r['total_emissions_kt'] / total_emissions * 100
        ws.cell(row=row, column=1, value=r['location'])
        ws.cell(row=row, column=2, value=f"{r['fuel_emissions_kt']:,.0f}")
        ws.cell(row=row, column=3, value=f"{r['elec_emissions_kt']:,.0f}")
        ws.cell(row=row, column=4, value=f"{r['process_emissions_kt']:,.0f}")
        ws.cell(row=row, column=5, value=f"{r['total_emissions_kt']:,.0f}")
        ws.cell(row=row, column=6, value=f"{share:.1f}%")
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{regional_emissions['fuel_emissions_kt'].sum():,.0f}").font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"{regional_emissions['elec_emissions_kt'].sum():,.0f}").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"{regional_emissions['process_emissions_kt'].sum():,.0f}").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"{regional_emissions['total_emissions_kt'].sum():,.0f}").font = Font(bold=True)
    ws.cell(row=row, column=6, value="100.0%").font = Font(bold=True)

    row += 3

    # Emissions by process type
    ws.cell(row=row, column=1, value="2. Emissions by Process Type (kt CO2)").font = SECTION_FONT
    row += 2

    process_emissions = fac_emissions.groupby('process')['total_emissions_kt'].sum().sort_values(ascending=False)

    headers = ["Process", "Total Emissions (kt)", "Share (%)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for process, emissions in process_emissions.items():
        share = emissions / total_emissions * 100
        ws.cell(row=row, column=1, value=process)
        ws.cell(row=row, column=2, value=f"{emissions:,.0f}")
        ws.cell(row=row, column=3, value=f"{share:.1f}%")
        row += 1

    row += 3

    # NCC vs Non-NCC breakdown
    ws.cell(row=row, column=1, value="3. NCC vs Non-NCC Emissions").font = SECTION_FONT
    row += 2

    ncc_emissions = fac_emissions[fac_emissions['process'] == 'Naphtha Cracker']['total_emissions_kt'].sum()
    non_ncc_emissions = total_emissions - ncc_emissions

    ws.cell(row=row, column=1, value="NCC Facilities")
    ws.cell(row=row, column=2, value=f"{ncc_emissions:,.0f} kt")
    ws.cell(row=row, column=3, value=f"{ncc_emissions/total_emissions*100:.1f}%")
    row += 1

    ws.cell(row=row, column=1, value="Non-NCC Facilities")
    ws.cell(row=row, column=2, value=f"{non_ncc_emissions:,.0f} kt")
    ws.cell(row=row, column=3, value=f"{non_ncc_emissions/total_emissions*100:.1f}%")
    row += 1

    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{total_emissions:,.0f} kt").font = Font(bold=True)
    ws.cell(row=row, column=3, value="100.0%").font = Font(bold=True)

    row += 3

    # Explanation of residual emissions
    ws.cell(row=row, column=1, value="4. Why ~3.9 Mt Residual Emissions in 2050?").font = SECTION_FONT
    row += 2

    explanation = [
        "The deployed technologies (NCC-H2, NCC-Electricity, Heat Pump, RE-PPA) primarily target:",
        "- Naphtha Cracker facilities (covered by NCC-H2 or NCC-Electricity)",
        "- Low-temperature heat (<165°C) (covered by Heat Pump)",
        "- Grid electricity emissions (covered by RE-PPA)",
        "",
        "NOT covered by deployed technologies:",
        f"- BTX Plant emissions: ~{fac_emissions[fac_emissions['process'] == 'BTX Plant']['total_emissions_kt'].sum()/1000:.1f} Mt",
        f"- Utility process emissions: ~{fac_emissions[fac_emissions['process'] == 'Utility']['total_emissions_kt'].sum()/1000:.1f} Mt",
        "- High-temperature process emissions",
        "",
        "To achieve true Net Zero, additional measures needed:",
        "- RDH (Reformer Duty Heater) technology for high-temp heat",
        "- BTX plant electrification or hydrogen",
        "- Carbon capture for unavoidable process emissions",
    ]

    for text in explanation:
        ws.cell(row=row, column=1, value=text)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    return wb


def create_scenario_regional_sheet(wb, data, scenario_id, sheet_num):
    """Create regional breakdown for a specific scenario"""
    scenario_names = {
        'shaheen_ncc_h2': 'Shaheen + NCC-H2',
        'shaheen_ncc_electricity': 'Shaheen + NCC-Elec',
        'restructure_25pct_ncc_h2': 'Restructure 25% + H2',
        'restructure_25pct_ncc_electricity': 'Restructure 25% + Elec',
        'restructure_40pct_ncc_h2': 'Restructure 40% + H2',
        'restructure_40pct_ncc_electricity': 'Restructure 40% + Elec',
    }

    ws = wb.create_sheet(f"{sheet_num}. {scenario_names[scenario_id][:18]}")

    ws['A1'] = f"Scenario: {scenario_names[scenario_id]}"
    ws['A1'].font = TITLE_FONT

    scenario_data = data['scenario_data'].get(scenario_id, {})
    scenario_summary = data['scenarios'][data['scenarios']['scenario_id'] == scenario_id]

    if scenario_data.get('facilities') is None:
        ws['A3'] = "Scenario data not available"
        return wb

    facilities = scenario_data['facilities']
    emissions = scenario_data['emissions']
    deployment = scenario_data['deployment']
    energy_intensity = data['energy_intensity']

    row = 3

    # Scenario summary
    ws.cell(row=row, column=1, value="1. Scenario Summary").font = SECTION_FONT
    row += 2

    if len(scenario_summary) > 0:
        s = scenario_summary.iloc[0]
        summary_data = [
            ("Total Facilities", s['n_facilities']),
            ("NCC Facilities", s['n_ncc_facilities']),
            ("NCC Capacity (kt)", f"{s['ncc_capacity_kt']:,.0f}"),
            ("BAU 2050 (Mt)", f"{s['bau_2050_mt']:.2f}"),
            ("Net 2050 (Mt)", f"{s['net_2050_mt']:.2f}"),
            ("NCC Abatement (Mt)", f"{s['ncc_abatement_mt']:.2f}"),
            ("Heat Pump Abatement (Mt)", f"{s['heat_pump_mt']:.2f}"),
            ("RE-PPA Abatement (Mt)", f"{s['re_ppa_mt']:.2f}"),
            ("Total CAPEX ($B)", f"${s['capex_billion_usd']:.1f}"),
        ]

        if 'NCC-H2' in scenario_id or 'ncc_h2' in scenario_id:
            summary_data.append(("H2 Demand (kt)", f"{s['h2_kt']:,.0f}"))
        else:
            summary_data.append(("Electricity Demand (TWh)", f"{s['electricity_twh']:.1f}"))

        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

    row += 2

    # Regional NCC distribution
    ws.cell(row=row, column=1, value="2. NCC Facilities by Region").font = SECTION_FONT
    row += 2

    ncc_facilities = facilities[facilities['process'] == 'Naphtha Cracker']
    ncc_regional = ncc_facilities.groupby('location').agg({
        'capacity_kt': 'sum',
        'product': 'count'
    }).reset_index()
    ncc_regional.columns = ['Region', 'NCC Capacity (kt)', 'NCC Count']
    ncc_regional = ncc_regional.sort_values('NCC Capacity (kt)', ascending=False)

    headers = ["Region", "NCC Capacity (kt)", "NCC Count", "Share (%)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    total_ncc_cap = ncc_regional['NCC Capacity (kt)'].sum()

    for _, r in ncc_regional.iterrows():
        share = r['NCC Capacity (kt)'] / total_ncc_cap * 100
        ws.cell(row=row, column=1, value=r['Region'])
        ws.cell(row=row, column=2, value=f"{r['NCC Capacity (kt)']:,.0f}")
        ws.cell(row=row, column=3, value=r['NCC Count'])
        ws.cell(row=row, column=4, value=f"{share:.1f}%")
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{total_ncc_cap:,.0f}").font = Font(bold=True)
    ws.cell(row=row, column=3, value=ncc_regional['NCC Count'].sum()).font = Font(bold=True)
    ws.cell(row=row, column=4, value="100.0%").font = Font(bold=True)

    row += 3

    # Deployment trajectory
    ws.cell(row=row, column=1, value="3. Emission Pathway (2025-2050)").font = SECTION_FONT
    row += 2

    headers = ["Year", "BAU (Mt)", "Abatement (Mt)", "Net Emissions (Mt)", "Reduction (%)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for year in [2025, 2030, 2035, 2040, 2045, 2050]:
        year_data = deployment[deployment['year'] == year]
        if len(year_data) > 0:
            d = year_data.iloc[0]
            reduction = (1 - d['actual_emissions_mt'] / d['bau_mt']) * 100 if d['bau_mt'] > 0 else 0
            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=f"{d['bau_mt']:.2f}")
            ws.cell(row=row, column=3, value=f"{d['deployed_abatement_mt']:.2f}")
            ws.cell(row=row, column=4, value=f"{d['actual_emissions_mt']:.2f}")
            ws.cell(row=row, column=5, value=f"{reduction:.1f}%")
            row += 1

    row += 2

    # Regional emissions 2050
    if emissions is not None:
        ws.cell(row=row, column=1, value="4. Regional Emissions 2050 (Top Regions)").font = SECTION_FONT
        row += 2

        regional_2050 = emissions.groupby('location')['total_emissions_kt'].sum().sort_values(ascending=False)

        headers = ["Region", "Emissions 2050 (kt)", "Share (%)"]
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=j, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row += 1

        total_2050 = regional_2050.sum()

        for region, em in regional_2050.head(10).items():
            share = em / total_2050 * 100 if total_2050 > 0 else 0
            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value=f"{em:,.0f}")
            ws.cell(row=row, column=3, value=f"{share:.1f}%")
            row += 1

        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{total_2050:,.0f}").font = Font(bold=True)
        ws.cell(row=row, column=3, value="100.0%").font = Font(bold=True)

    row += 3

    # Energy demand by region (estimated)
    ws.cell(row=row, column=1, value="5. Regional Energy Demand for Decarbonization").font = SECTION_FONT
    row += 2

    # Estimate energy demand proportional to NCC capacity
    tech_type = 'H2' if 'h2' in scenario_id else 'Electricity'

    if len(scenario_summary) > 0:
        s = scenario_summary.iloc[0]

        if 'h2' in scenario_id:
            total_energy = s['h2_kt']
            energy_unit = "H2 (kt)"
        else:
            total_energy = s['electricity_twh']
            energy_unit = "Electricity (TWh)"

        ncc_regional['Energy Demand'] = (ncc_regional['NCC Capacity (kt)'] / total_ncc_cap) * total_energy

        headers = ["Region", "NCC Capacity (kt)", f"{energy_unit}", "Share (%)"]
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=j, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row += 1

        for _, r in ncc_regional.iterrows():
            share = r['NCC Capacity (kt)'] / total_ncc_cap * 100
            ws.cell(row=row, column=1, value=r['Region'])
            ws.cell(row=row, column=2, value=f"{r['NCC Capacity (kt)']:,.0f}")
            ws.cell(row=row, column=3, value=f"{r['Energy Demand']:,.0f}" if 'h2' in scenario_id else f"{r['Energy Demand']:.2f}")
            ws.cell(row=row, column=4, value=f"{share:.1f}%")
            row += 1

        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{total_ncc_cap:,.0f}").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"{total_energy:,.0f}" if 'h2' in scenario_id else f"{total_energy:.2f}").font = Font(bold=True)
        ws.cell(row=row, column=4, value="100.0%").font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12

    return wb


def create_technology_deployment_sheet(wb, data):
    """Create technology deployment schedule sheet"""
    ws = wb.create_sheet("4. Tech Deployment")

    ws['A1'] = "Technology Deployment Schedule"
    ws['A1'].font = TITLE_FONT

    row = 3

    # Technology overview
    ws.cell(row=row, column=1, value="1. Technology Parameters").font = SECTION_FONT
    row += 2

    tech_params = data['tech_params']

    headers = ["Technology", "Applies To", "CAPEX 2025", "CAPEX 2050", "Decline", "Available"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for _, t in tech_params.iterrows():
        capex_2025 = t['capex_2025_musd_per_mtco2']
        capex_2050 = t['capex_2050_musd_per_mtco2']
        decline = (1 - capex_2050 / capex_2025) * 100 if capex_2025 > 0 else 0

        ws.cell(row=row, column=1, value=t['technology'])
        ws.cell(row=row, column=2, value=t['applies_to'])
        ws.cell(row=row, column=3, value=f"${capex_2025:,.0f}M" if capex_2025 > 0 else "-")
        ws.cell(row=row, column=4, value=f"${capex_2050:,.0f}M" if capex_2050 > 0 else "-")
        ws.cell(row=row, column=5, value=f"{decline:.0f}%" if decline > 0 else "-")
        ws.cell(row=row, column=6, value=int(t['available_year']))
        row += 1

    row += 2

    # CAPEX learning curve
    ws.cell(row=row, column=1, value="2. CAPEX Learning Curve (50% decline by 2050)").font = SECTION_FONT
    row += 2

    years = [2025, 2030, 2035, 2040, 2045, 2050]
    headers = ["Technology"] + [str(y) for y in years]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for _, t in tech_params.iterrows():
        if t['capex_2025_musd_per_mtco2'] > 0:
            capex_2025 = t['capex_2025_musd_per_mtco2']
            capex_2050 = t['capex_2050_musd_per_mtco2']

            ws.cell(row=row, column=1, value=t['technology'])
            for j, year in enumerate(years, 2):
                capex = capex_2025 - (capex_2025 - capex_2050) * (year - 2025) / 25
                ws.cell(row=row, column=j, value=f"${capex:,.0f}")
            row += 1

    row += 2

    # H2 price trajectory
    ws.cell(row=row, column=1, value="3. Green Hydrogen Price Trajectory (LCOH-based)").font = SECTION_FONT
    row += 2

    h2_prices = data['h2_prices']

    headers = ["Year", "H2 Price ($/kg)", "RE Price ($/MWh)", "Notes"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    re_prices = data['re_prices']

    for year in years:
        h2_row = h2_prices[h2_prices['year'] == year]
        re_row = re_prices[re_prices['year'] == year]

        if len(h2_row) > 0:
            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=f"${h2_row.iloc[0]['h2_price_usd_per_kg']:.2f}")
            ws.cell(row=row, column=3, value=f"${re_row.iloc[0]['re_price_usd_per_mwh']:.1f}" if len(re_row) > 0 else "-")
            ws.cell(row=row, column=4, value=h2_row.iloc[0]['notes'] if 'notes' in h2_row.columns else "")
            row += 1

    row += 2

    # Grid emission factor trajectory
    ws.cell(row=row, column=1, value="4. Grid Decarbonization Trajectory").font = SECTION_FONT
    row += 2

    grid_ef = data['grid_ef']

    headers = ["Year", "Grid EF (tCO2/MWh)", "Reduction vs 2025"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    ef_2025 = grid_ef[grid_ef['year'] == 2025]['grid_ef_tco2_per_mwh'].values[0] if len(grid_ef[grid_ef['year'] == 2025]) > 0 else 0.459

    for year in years:
        ef_row = grid_ef[grid_ef['year'] == year]
        if len(ef_row) > 0:
            ef = ef_row.iloc[0]['grid_ef_tco2_per_mwh']
            reduction = (1 - ef / ef_2025) * 100 if ef_2025 > 0 else 0
            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=f"{ef:.3f}")
            ws.cell(row=row, column=3, value=f"{reduction:.0f}%")
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15

    return wb


def create_total_summary_sheet(wb, data):
    """Create total national summary sheet"""
    ws = wb.create_sheet("10. Total Summary")

    ws['A1'] = "National Total Summary - All 6 Scenarios"
    ws['A1'].font = TITLE_FONT

    scenarios = data['scenarios']

    row = 3

    # Complete scenario comparison
    ws.cell(row=row, column=1, value="1. Complete Scenario Comparison").font = SECTION_FONT
    row += 2

    headers = ["Metric", "Shaheen+H2", "Shaheen+Elec", "25%+H2", "25%+Elec", "40%+H2", "40%+Elec"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    metrics = [
        ('Total Facilities', 'n_facilities', '{:.0f}'),
        ('NCC Facilities', 'n_ncc_facilities', '{:.0f}'),
        ('NCC Capacity (kt)', 'ncc_capacity_kt', '{:,.0f}'),
        ('BAU 2050 (Mt)', 'bau_2050_mt', '{:.1f}'),
        ('Net 2050 (Mt)', 'net_2050_mt', '{:.1f}'),
        ('NCC Abatement (Mt)', 'ncc_abatement_mt', '{:.1f}'),
        ('Heat Pump (Mt)', 'heat_pump_mt', '{:.2f}'),
        ('RE-PPA (Mt)', 're_ppa_mt', '{:.2f}'),
        ('CAPEX ($B)', 'capex_billion_usd', '${:.1f}'),
        ('H2 Demand (kt)', 'h2_kt', '{:,.0f}'),
        ('Electricity (TWh)', 'electricity_twh', '{:.2f}'),
    ]

    scenario_order = [
        'shaheen_ncc_h2', 'shaheen_ncc_electricity',
        'restructure_25pct_ncc_h2', 'restructure_25pct_ncc_electricity',
        'restructure_40pct_ncc_h2', 'restructure_40pct_ncc_electricity'
    ]

    for metric_name, col_name, fmt in metrics:
        ws.cell(row=row, column=1, value=metric_name)
        for j, scenario_id in enumerate(scenario_order, 2):
            scenario_row = scenarios[scenarios['scenario_id'] == scenario_id]
            if len(scenario_row) > 0:
                value = scenario_row.iloc[0][col_name]
                ws.cell(row=row, column=j, value=fmt.format(value))
            else:
                ws.cell(row=row, column=j, value="-")
        row += 1

    row += 2

    # Key insights
    ws.cell(row=row, column=1, value="2. Key Insights").font = SECTION_FONT
    row += 2

    insights = [
        "Production Pathway Impact:",
        f"  - Shaheen scenario adds 6 new facilities, increasing emissions potential",
        f"  - 25% restructure retires oldest NCC (9 facilities, ~25% capacity)",
        f"  - 40% restructure retires 16 oldest NCC facilities (~40% capacity)",
        "",
        "Technology Pathway Impact:",
        f"  - NCC-H2 is ~10% cheaper than NCC-Electricity",
        f"  - NCC-Electricity requires massive grid expansion",
        f"  - NCC-H2 requires H2 infrastructure (~50-100 kt H2/year)",
        "",
        "Residual Emissions (~3.9 Mt in 2050):",
        "  - BTX Plants: Not covered by deployed technologies",
        "  - Utility processes: Partially covered by Heat Pump",
        "  - Additional technologies needed for true Net Zero",
    ]

    for text in insights:
        ws.cell(row=row, column=1, value=text)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 14

    return wb


def generate_report():
    """Generate the complete regional deployment report"""
    print("="*70)
    print("Regional Deployment Report Generator")
    print("="*70)

    print("\nLoading data...")
    data = load_data()

    print("Creating workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    print("Creating Sheet 1: Overview...")
    wb = create_overview_sheet(wb, data)

    print("Creating Sheet 2: Regional Facilities...")
    wb = create_regional_facilities_sheet(wb, data)

    print("Creating Sheet 3: Regional Emissions...")
    wb = create_regional_emissions_sheet(wb, data)

    print("Creating Sheet 4: Technology Deployment...")
    wb = create_technology_deployment_sheet(wb, data)

    # Create scenario-specific sheets
    scenario_ids = [
        'shaheen_ncc_h2', 'shaheen_ncc_electricity',
        'restructure_25pct_ncc_h2', 'restructure_25pct_ncc_electricity',
        'restructure_40pct_ncc_h2', 'restructure_40pct_ncc_electricity'
    ]

    for i, scenario_id in enumerate(scenario_ids, 5):
        print(f"Creating Sheet {i}: {scenario_id}...")
        wb = create_scenario_regional_sheet(wb, data, scenario_id, i)

    print("Creating Sheet 10: Total Summary...")
    wb = create_total_summary_sheet(wb, data)

    # Save
    output_path = REPORTS_DIR / "Regional_Deployment_Report.xlsx"
    wb.save(output_path)

    print("\n" + "="*70)
    print(f"Report saved to: {output_path}")
    print("="*70)

    print("\nReport Contents:")
    print("  1. Overview - Report summary and residual emissions explanation")
    print("  2. Regional Facilities - Capacity distribution by region")
    print("  3. Regional Emissions - BAU emissions breakdown by region")
    print("  4. Tech Deployment - Technology parameters and learning curves")
    print("  5. Shaheen + NCC-H2 - Regional breakdown")
    print("  6. Shaheen + NCC-Elec - Regional breakdown")
    print("  7. Restructure 25% + H2 - Regional breakdown")
    print("  8. Restructure 25% + Elec - Regional breakdown")
    print("  9. Restructure 40% + H2 - Regional breakdown")
    print("  10. Total Summary - National comparison all scenarios")

    return output_path


if __name__ == "__main__":
    generate_report()
