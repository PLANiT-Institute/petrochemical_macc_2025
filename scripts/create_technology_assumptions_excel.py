"""
Create Comprehensive Excel File of Technology Assumptions
For review and discussion before finalizing academic paper
"""

import pandas as pd
import numpy as np
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_technology_assumptions_excel():
    """Create comprehensive Excel file with all technology assumptions"""

    print("="*80)
    print("CREATING TECHNOLOGY ASSUMPTIONS EXCEL FILE")
    print("="*80)

    # Load all data
    print("\n📊 Loading data...")

    # Technology parameters
    df_tech = pd.read_csv('data/technology_parameters.csv')

    # Price trajectories
    df_h2 = pd.read_csv('data/h2_price_trajectory.csv')
    df_re = pd.read_csv('data/re_price_trajectory.csv')
    df_fuel = pd.read_csv('data/fuel_price_trajectory.csv')
    df_grid = pd.read_csv('data/grid_emission_trajectory.csv')

    # Emission factors
    df_ef = pd.read_csv('data/emission_factors.csv')

    # Energy intensities (sample)
    df_energy = pd.read_csv('data/energy_intensities.csv').head(20)  # First 20 facilities

    print("   ✓ All data loaded")

    # Create Excel file
    output_file = Path('docs/TECHNOLOGY_ASSUMPTIONS_FOR_REVIEW.xlsx')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # =================================================================
        # SHEET 1: TECHNOLOGY PARAMETERS OVERVIEW
        # =================================================================
        print("\n📋 Creating Sheet 1: Technology Parameters...")

        # Select key columns
        tech_overview = df_tech[[
            'technology', 'applies_to', 'available_year', 'trl',
            'capex_2025_musd_per_mtco2', 'capex_2030_musd_per_mtco2',
            'capex_2040_musd_per_mtco2', 'capex_2050_musd_per_mtco2',
            'opex_pct_capex', 'lifetime_years', 'notes'
        ]].copy()

        tech_overview.to_excel(writer, sheet_name='1_Tech_Parameters', index=False)

        # Format
        ws = writer.sheets['1_Tech_Parameters']

        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Column widths
        ws.column_dimensions['A'].width = 18  # technology
        ws.column_dimensions['B'].width = 25  # applies_to
        ws.column_dimensions['C'].width = 15  # available_year
        ws.column_dimensions['D'].width = 10  # trl
        ws.column_dimensions['E'].width = 15  # capex_2025
        ws.column_dimensions['F'].width = 15  # capex_2030
        ws.column_dimensions['G'].width = 15  # capex_2040
        ws.column_dimensions['H'].width = 15  # capex_2050
        ws.column_dimensions['I'].width = 12  # opex_pct
        ws.column_dimensions['J'].width = 12  # lifetime
        ws.column_dimensions['K'].width = 50  # notes

        # =================================================================
        # SHEET 2: ENERGY CONSUMPTION BY TECHNOLOGY
        # =================================================================
        print("📋 Creating Sheet 2: Energy Consumption...")

        energy_consumption = pd.DataFrame([
            {
                'Technology': 'Heat Pump',
                'Applicable_To': 'Non-NCC facilities (<165°C)',
                'COP': 4.0,
                'Energy_Input': 'Renewable Electricity',
                'Input_Unit': 'MWh per GJ thermal',
                'Input_Value': 0.25 / 3.6,  # 1/COP / 3.6 (GJ to MWh)
                'Energy_Output': 'Heat (thermal)',
                'Output_Unit': 'GJ thermal',
                'Fuel_Replaced': 'All fossil fuels (LNG, fuel gas, etc.)',
                'Emission_Factor_Input': 0.05,  # tCO2/MWh (RE lifecycle)
                'Emission_Factor_Replaced': 0.050,  # tCO2/GJ (avg fossil fuel)
                'Notes': 'COP=4 means 1 kWh electricity → 4 kWh heat'
            },
            {
                'Technology': 'NCC-H2',
                'Applicable_To': 'Naphtha crackers only',
                'COP': None,
                'Energy_Input': 'Green Hydrogen',
                'Input_Unit': 'kg H2 per ton ethylene',
                'Input_Value': 0.56,  # From energy balance: 11 GJ / 120 MJ/kg = 92 kg, adjusted
                'Energy_Output': 'Heat for cracking',
                'Output_Unit': 'GJ per ton ethylene',
                'Fuel_Replaced': 'LNG + Fuel Gas combustion (~11 GJ/ton)',
                'Emission_Factor_Input': 0.0,  # Green H2
                'Emission_Factor_Replaced': 0.050,  # Mixed fuel gas
                'Notes': 'Naphtha feedstock (29 GJ/ton) UNCHANGED - only combustion replaced'
            },
            {
                'Technology': 'NCC-Electricity',
                'Applicable_To': 'Naphtha crackers only',
                'COP': None,
                'Energy_Input': 'Renewable Electricity',
                'Input_Unit': 'MWh per ton ethylene',
                'Input_Value': 3.0,  # From tech parameters
                'Energy_Output': 'Heat for cracking',
                'Output_Unit': 'GJ per ton ethylene',
                'Fuel_Replaced': 'LNG + Fuel Gas combustion (~11 GJ/ton)',
                'Emission_Factor_Input': 0.05,  # tCO2/MWh (RE lifecycle)
                'Emission_Factor_Replaced': 0.050,  # Mixed fuel gas
                'Notes': 'Naphtha feedstock (29 GJ/ton) UNCHANGED - only combustion replaced'
            },
            {
                'Technology': 'RE PPA',
                'Applicable_To': 'NCC facilities (electricity users)',
                'COP': None,
                'Energy_Input': 'Renewable Electricity',
                'Input_Unit': 'MWh',
                'Input_Value': 1.0,  # 1:1 replacement
                'Energy_Output': 'Electricity',
                'Output_Unit': 'MWh',
                'Fuel_Replaced': 'Grid electricity',
                'Emission_Factor_Input': 0.05,  # tCO2/MWh (RE lifecycle)
                'Emission_Factor_Replaced': 0.45,  # tCO2/MWh (2025 grid, decreasing over time)
                'Notes': 'Simple fuel switching - no infrastructure change'
            }
        ])

        energy_consumption.to_excel(writer, sheet_name='2_Energy_Consumption', index=False)

        # Format
        ws = writer.sheets['2_Energy_Consumption']
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Adjust column widths
        for col in range(1, 13):
            ws.column_dimensions[chr(64+col)].width = 20
        ws.column_dimensions['M'].width = 60  # Notes

        # =================================================================
        # SHEET 3: MACC CALCULATION BREAKDOWN (2030 EXAMPLE)
        # =================================================================
        print("📋 Creating Sheet 3: MACC Calculation (2030)...")

        # Get 2030 prices
        h2_2030 = df_h2[df_h2['year'] == 2030]['h2_price_usd_per_kg'].iloc[0]
        re_2030 = df_re[df_re['year'] == 2030]['re_price_usd_per_mwh'].iloc[0]
        grid_ef_2030 = df_grid[df_grid['year'] == 2030]['grid_ef_tco2_per_mwh'].iloc[0]

        # Calculate MACC components for each technology
        macc_breakdown = []

        # Heat Pump
        hp_capex_2030 = np.interp(2030, [2025, 2030, 2040, 2050],
                                   [900, 720, 540, 450])
        hp_lifetime = 20
        hp_opex_pct = 3.0

        macc_breakdown.append({
            'Technology': 'Heat Pump',
            'Year': 2030,
            'CAPEX_Total_MUSD_per_MtCO2': hp_capex_2030,
            'Lifetime_years': hp_lifetime,
            'CAPEX_Annual_USD_per_tCO2': hp_capex_2030 / hp_lifetime,
            'OPEX_pct_CAPEX': hp_opex_pct,
            'OPEX_Annual_USD_per_tCO2': hp_capex_2030 * (hp_opex_pct/100),
            'Fuel_Type_New': 'RE Electricity',
            'Fuel_Price_New': f'${re_2030}/MWh',
            'Fuel_Consumption_per_tCO2': 'Varies by baseline',
            'Fuel_Cost_Differential_USD_per_tCO2': '(Calculated in model)',
            'Total_MACC_USD_per_tCO2': '(See Module 2 outputs)',
            'Notes': f'COP=4, RE price=${re_2030}/MWh'
        })

        # NCC-H2
        ncc_h2_capex_2030 = np.interp(2030, [2025, 2030, 2040, 2050],
                                       [1725, 1440, 1035, 863])
        ncc_h2_lifetime = 25
        ncc_h2_opex_pct = 4.0
        h2_per_ton = 0.56  # kg H2 per ton ethylene
        abatement_per_ton = 0.53  # tCO2 per ton ethylene (approximate)
        h2_cost_per_tco2 = (h2_per_ton * h2_2030) / abatement_per_ton

        macc_breakdown.append({
            'Technology': 'NCC-H2',
            'Year': 2030,
            'CAPEX_Total_MUSD_per_MtCO2': ncc_h2_capex_2030,
            'Lifetime_years': ncc_h2_lifetime,
            'CAPEX_Annual_USD_per_tCO2': ncc_h2_capex_2030 / ncc_h2_lifetime,
            'OPEX_pct_CAPEX': ncc_h2_opex_pct,
            'OPEX_Annual_USD_per_tCO2': ncc_h2_capex_2030 * (ncc_h2_opex_pct/100),
            'Fuel_Type_New': 'Green H2',
            'Fuel_Price_New': f'${h2_2030}/kg',
            'Fuel_Consumption_per_tCO2': f'{h2_per_ton:.2f} kg H2 / ton ethylene',
            'Fuel_Cost_Differential_USD_per_tCO2': f'~${h2_cost_per_tco2:.0f}',
            'Total_MACC_USD_per_tCO2': f'~${ncc_h2_capex_2030/ncc_h2_lifetime + ncc_h2_capex_2030*(ncc_h2_opex_pct/100) + h2_cost_per_tco2:.0f}',
            'Notes': f'H2 price=${h2_2030}/kg, {h2_per_ton} kg/ton ethylene'
        })

        # NCC-Electricity
        ncc_elec_capex_2030 = np.interp(2030, [2025, 2030, 2040, 2050],
                                         [1840, 1560, 1150, 940])
        ncc_elec_lifetime = 25
        ncc_elec_opex_pct = 3.5
        elec_per_ton = 3.0  # MWh per ton ethylene
        elec_cost_per_tco2 = (elec_per_ton * re_2030) / abatement_per_ton

        macc_breakdown.append({
            'Technology': 'NCC-Electricity',
            'Year': 2030,
            'CAPEX_Total_MUSD_per_MtCO2': ncc_elec_capex_2030,
            'Lifetime_years': ncc_elec_lifetime,
            'CAPEX_Annual_USD_per_tCO2': ncc_elec_capex_2030 / ncc_elec_lifetime,
            'OPEX_pct_CAPEX': ncc_elec_opex_pct,
            'OPEX_Annual_USD_per_tCO2': ncc_elec_capex_2030 * (ncc_elec_opex_pct/100),
            'Fuel_Type_New': 'RE Electricity',
            'Fuel_Price_New': f'${re_2030}/MWh',
            'Fuel_Consumption_per_tCO2': f'{elec_per_ton:.1f} MWh / ton ethylene',
            'Fuel_Cost_Differential_USD_per_tCO2': f'~${elec_cost_per_tco2:.0f}',
            'Total_MACC_USD_per_tCO2': f'~${ncc_elec_capex_2030/ncc_elec_lifetime + ncc_elec_capex_2030*(ncc_elec_opex_pct/100) + elec_cost_per_tco2:.0f}',
            'Notes': f'RE price=${re_2030}/MWh, {elec_per_ton} MWh/ton ethylene'
        })

        # RE PPA
        re_ppa_abatement_per_mwh = grid_ef_2030 - 0.05  # Grid EF - RE lifecycle
        re_ppa_cost = re_2030 / re_ppa_abatement_per_mwh

        macc_breakdown.append({
            'Technology': 'RE PPA',
            'Year': 2030,
            'CAPEX_Total_MUSD_per_MtCO2': 0,
            'Lifetime_years': 99,
            'CAPEX_Annual_USD_per_tCO2': 0,
            'OPEX_pct_CAPEX': 0,
            'OPEX_Annual_USD_per_tCO2': 0,
            'Fuel_Type_New': 'RE Electricity',
            'Fuel_Price_New': f'${re_2030}/MWh',
            'Fuel_Consumption_per_tCO2': f'1 MWh / {re_ppa_abatement_per_mwh:.3f} tCO2',
            'Fuel_Cost_Differential_USD_per_tCO2': f'${re_ppa_cost:.0f}',
            'Total_MACC_USD_per_tCO2': f'${re_ppa_cost:.0f}',
            'Notes': f'Grid EF={grid_ef_2030:.3f}, RE EF=0.05 tCO2/MWh'
        })

        df_macc = pd.DataFrame(macc_breakdown)
        df_macc.to_excel(writer, sheet_name='3_MACC_Calculation_2030', index=False)

        # Format
        ws = writer.sheets['3_MACC_Calculation_2030']
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1, 13):
            ws.column_dimensions[chr(64+col)].width = 18
        ws.column_dimensions['M'].width = 50

        # =================================================================
        # SHEET 4: PRICE TRAJECTORIES
        # =================================================================
        print("📋 Creating Sheet 4: Price Trajectories...")

        # Combine price data
        price_traj = pd.DataFrame({
            'Year': df_h2['year'],
            'H2_USD_per_kg': df_h2['h2_price_usd_per_kg'],
            'RE_USD_per_MWh': df_re['re_price_usd_per_mwh'],
            'Grid_EF_tCO2_per_MWh': df_grid['grid_ef_tco2_per_mwh'],
            'Naphtha_USD_per_GJ': df_fuel['naphtha_usd_per_gj'],
            'LNG_USD_per_GJ': df_fuel['lng_usd_per_gj'],
            'Electricity_USD_per_kWh': df_fuel['electricity_usd_per_kwh']
        })

        price_traj.to_excel(writer, sheet_name='4_Price_Trajectories', index=False)

        # Format
        ws = writer.sheets['4_Price_Trajectories']
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1, 8):
            ws.column_dimensions[chr(64+col)].width = 20

        # =================================================================
        # SHEET 5: EMISSION FACTORS
        # =================================================================
        print("📋 Creating Sheet 5: Emission Factors...")

        df_ef.to_excel(writer, sheet_name='5_Emission_Factors', index=False)

        # Format
        ws = writer.sheets['5_Emission_Factors']
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 50

        # =================================================================
        # SHEET 6: ENERGY INTENSITIES (Sample)
        # =================================================================
        print("📋 Creating Sheet 6: Energy Intensities (Sample)...")

        df_energy.to_excel(writer, sheet_name='6_Energy_Intensities_Sample', index=False)

        # Format
        ws = writer.sheets['6_Energy_Intensities_Sample']
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1, 15):
            ws.column_dimensions[chr(64+col)].width = 18

        # =================================================================
        # SHEET 7: KEY ASSUMPTIONS SUMMARY
        # =================================================================
        print("📋 Creating Sheet 7: Key Assumptions Summary...")

        assumptions = pd.DataFrame([
            ['CATEGORY', 'PARAMETER', 'VALUE', 'SOURCE / JUSTIFICATION'],
            ['', '', '', ''],
            ['Economic', 'Discount Rate', '0% (simple annualization)', 'Simplified approach: CAPEX_annual = CAPEX / Lifetime'],
            ['Economic', 'Analysis Period', '2025-2050 (25 years)', 'Aligned with Korea NDC timeline'],
            ['Economic', 'Currency', 'USD (2025 constant)', 'All costs in nominal 2025 USD'],
            ['', '', '', ''],
            ['Facility', 'Number of Facilities', '248', 'Complete Korea petrochemical sector coverage'],
            ['Facility', 'Facility Lifetime', 'Infinite (no retirement)', 'Conservative assumption - all facilities operate through 2050'],
            ['Facility', 'Demand Growth', '28.8% total (2025-2050)', 'Based on industry projections, 1.5% → 0.5% annual'],
            ['', '', '', ''],
            ['Technology', 'Number of Technologies', '4 (Heat Pump, NCC-H2, NCC-Elec, RE PPA)', 'Focus on major decarbonization options'],
            ['Technology', 'NCC Mutual Exclusivity', 'YES (H2 OR Electricity, not both)', 'Physical constraint - same equipment'],
            ['Technology', 'Technology Irreversibility', 'YES (capacity cannot decrease)', 'Capital lock-in - cannot un-build'],
            ['Technology', 'Industry-wide Selection', 'YES (all facilities choose same NCC tech)', 'Simplification - technology standardization'],
            ['', '', '', ''],
            ['Baseline', '2025 Baseline Emissions', '66.2 MtCO2', 'Corrected with proper LNG/fuel gas emission factors'],
            ['Baseline', 'LNG Emission Factor', '0.0561 tCO2/GJ', 'IPCC 2019 Guidelines (corrected from 0.0149)'],
            ['Baseline', 'Naphtha Emission Factor', '0.0542 tCO2/GJ', 'IPCC 2019 Guidelines'],
            ['Baseline', 'Grid Decarbonization', '0.45 → 0.25 tCO2/MWh (2025→2050)', 'Korea grid trajectory (conservative)'],
            ['', '', '', ''],
            ['Prices', 'H2 Price (2025)', '$6.00/kg', 'IEA/IRENA 2024 (corrected from $12)'],
            ['Prices', 'H2 Price (2030)', '$3.50/kg', 'IEA/IRENA 2024 projections (corrected from $10)'],
            ['Prices', 'H2 Price (2050)', '$2.00/kg', 'Long-term floor estimate'],
            ['Prices', 'RE Electricity (2025)', '$90/MWh', 'Korea auctions 2024 (corrected from $130)'],
            ['Prices', 'RE Electricity (2030)', '$75/MWh', 'IRENA 2024 projections (corrected from $115)'],
            ['Prices', 'RE Electricity (2050)', '$50/MWh', 'Long-term cost floor'],
            ['Prices', 'Fossil Fuels', 'Constant at 2025 levels', 'Conservative assumption - no price increase'],
            ['', '', '', ''],
            ['NCC', 'NCC Definition', 'Ethylene, Propylene, Butadiene only', 'True cracker products (NOT BTX)'],
            ['NCC', 'NCC Facilities', '~40 facilities', 'Out of 248 total'],
            ['NCC', 'Naphtha Role', 'Feedstock only (29 GJ/ton)', 'Chemically converted, NOT burned'],
            ['NCC', 'Combustion Energy', 'LNG + Fuel Gas (~11 GJ/ton)', 'THIS is what gets replaced by H2 or electricity'],
            ['', '', '', ''],
            ['Methodology', 'Optimization Algorithm', 'Greedy cost-ordered deployment', 'Deploy cheapest technology first each year'],
            ['Methodology', 'Cost Calculation', 'Energy-based (not LCOE)', 'Explicit fuel consumption tracking'],
            ['Methodology', 'MACC Formula', 'CAPEX_ann + OPEX_ann + Fuel_diff', 'Standard MACC methodology'],
            ['Methodology', 'Fuel Cost Differential', 'New fuel cost only (NO baseline subtraction)', '⚠️ DISCUSS: Should we change to standard (New - Old)?'],
        ])

        assumptions.to_excel(writer, sheet_name='7_Key_Assumptions', index=False, header=False)

        # Format
        ws = writer.sheets['7_Key_Assumptions']

        # Header row
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Category rows (bold)
        category_rows = [3, 6, 10, 15, 20, 27, 31]
        for row in category_rows:
            for cell in ws[row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 60

        # Highlight discussion point
        ws['D35'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws['D35'].font = Font(bold=True)

    print(f"\n✅ Excel file created: {output_file}")
    print(f"\nSheets included:")
    print("  1. Technology Parameters - CAPEX, OPEX, lifetime by year")
    print("  2. Energy Consumption - Fuel/energy use for each technology")
    print("  3. MACC Calculation (2030) - Detailed cost breakdown")
    print("  4. Price Trajectories - H2, RE, fuels (2025-2050)")
    print("  5. Emission Factors - All fuels with sources")
    print("  6. Energy Intensities - Sample facilities (first 20)")
    print("  7. Key Assumptions - Summary of all major assumptions")

    print("\n" + "="*80)
    print("EXCEL FILE READY FOR REVIEW")
    print("="*80)
    print(f"\nFile: {output_file}")
    print("\n💡 Key Points to Review:")
    print("  1. Sheet 3: Check MACC calculation methodology")
    print("  2. Sheet 7: Review KEY ASSUMPTION highlighted in yellow")
    print("     → Should fuel cost differential include baseline savings?")
    print("  3. All price trajectories (corrected values)")
    print("  4. Technology parameters and energy consumption")
    print("\n" + "="*80)

    return output_file

if __name__ == '__main__':
    create_technology_assumptions_excel()
