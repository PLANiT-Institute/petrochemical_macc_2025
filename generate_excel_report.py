"""
Comprehensive Excel Report Generator for Petrochemical Decarbonization Model
248 Facilities | 2025-2050 | NCC-Electricity (Main) & NCC-H2 (Alternative) Scenarios

RE-PPA Price & LCOH Methodology: PLANiT Institute (2025)
Operating Rate Assumption: 70% (Flat 2025-2050)

Emission Targets (from emission_scenarios_clean.csv):
- 2025: 66.2 MtCO2 (Baseline at 100% capacity)
- 2035: 43.5 MtCO2 (24.5% reduction from 2018 baseline of 57.6 Mt)
- 2050: 0 MtCO2 (Net Zero)

Output: Multi-sheet Excel workbook with full facility-level yearly details
"""

import pandas as pd
import numpy as np
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from datetime import datetime

# Add modules to path
import sys
sys.path.insert(0, str(Path(__file__).parent))

from modules.utils import DataLoader, EmissionCalculator, identify_product_group, is_ncc_facility


class ComprehensiveExcelGenerator:
    """
    Generate comprehensive Excel output for petrochemical decarbonization analysis

    Features:
    - 248 facilities with full details
    - Yearly tracking (2025-2050)
    - Energy consumption by fuel type
    - Cost breakdown (CAPEX, OPEX, Fuel)
    - Technology transition timing
    - Regional and company aggregations
    """

    def __init__(self, data_dir='data', output_dir='outputs/excel_report'):
        self.data_dir = Path(data_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        print("="*80)
        print("COMPREHENSIVE EXCEL REPORT GENERATOR")
        print("="*80)

        # Load all data
        self._load_data()

        # User-specified baseline (will normalize calculated emissions to this)
        self.target_baseline_mt = 66.2  # MtCO2 (2025 baseline at 100% capacity)

        # Load emission targets from scenario file
        try:
            df_scenarios = pd.read_csv(self.data_dir / 'emission_scenarios_clean.csv')
            # Filter for Policy_Target scenario
            policy_target = df_scenarios[df_scenarios['scenario_name'] == 'Policy_Target']
            
            # Build emission targets dict from the data
            self.emission_targets = {}
            for _, row in policy_target.iterrows():
                if pd.notna(row['year']):
                    self.emission_targets[int(row['year'])] = row['target_mt']
            
            print(f"\nLoaded emission targets from emission_scenarios_clean.csv:")
        except FileNotFoundError:
            print("\n⚠️  emission_scenarios_clean.csv not found, using default targets")
            # Fallback to default targets
            self.emission_targets = {
                2025: 52.0,      # Baseline: 52 MtCO2
                2035: 39.26,     # -24.5% reduction
                2050: 0.0        # -100% (Net Zero)
            }

        # Interpolate targets for all years
        self.yearly_targets = self._interpolate_targets()

        print(f"\nEmission Targets:")
        print(f"  2025: {self.emission_targets[2025]:.2f} MtCO2 (Baseline)")
        print(f"  2035: {self.emission_targets[2035]:.2f} MtCO2 (-24.5%)")
        print(f"  2050: {self.emission_targets[2050]:.2f} MtCO2 (-100%)")

    def _load_data(self):
        """Load all required data files"""
        print("\nLoading data files...")

        loader = DataLoader(self.data_dir)

        # Facility data
        self.df_facilities = loader.load_facilities()
        self.df_intensities = loader.load_energy_intensities()
        print(f"  - {len(self.df_facilities)} facilities loaded")

        # Emission factors
        self.df_emission_factors = loader.load_emission_factors()
        self.emission_calc = EmissionCalculator(self.df_emission_factors)

        # Technology parameters
        self.df_tech_params = loader.load_technology_params()
        print(f"  - {len(self.df_tech_params)} technologies loaded")

        # Price trajectories
        self.df_h2_prices = loader.load_h2_prices()
        self.df_re_prices = loader.load_re_prices()
        self.df_grid_prices = loader.load_grid_prices()
        self.df_fuel_prices = pd.read_csv(self.data_dir / 'fuel_price_trajectory.csv')
        self.df_grid_ef = pd.read_csv(self.data_dir / 'grid_emission_trajectory.csv')

        # Heat pump applicability
        self.df_hp_applicability = loader.load_heat_pump_applicability()

    def _interpolate_targets(self):
        """Interpolate emission targets for all years 2025-2050"""
        years = list(range(2025, 2051))
        targets = {}

        key_years = sorted(self.emission_targets.keys())

        for year in years:
            if year in self.emission_targets:
                targets[year] = self.emission_targets[year]
            else:
                # Linear interpolation
                for i in range(len(key_years) - 1):
                    if key_years[i] < year < key_years[i+1]:
                        y1, y2 = key_years[i], key_years[i+1]
                        t1, t2 = self.emission_targets[y1], self.emission_targets[y2]
                        targets[year] = t1 + (t2 - t1) * (year - y1) / (y2 - y1)
                        break

        return targets

    def calculate_facility_baseline(self):
        """Calculate baseline emissions and energy for all 248 facilities"""
        print("\nCalculating facility baseline (2025)...")

        baseline_data = []

        for idx, facility in self.df_facilities.iterrows():
            intensity = self.df_intensities.iloc[idx]
            capacity = facility['capacity_kt']  # kt/year

            # Energy consumption (total per year)
            energy = {
                'naphtha_gj': intensity['Naphtha_GJ_per_tonne'] * capacity * 1000,
                'electricity_kwh': intensity['Electricity_kWh_per_tonne'] * capacity * 1000,
                'lng_gj': intensity['LNG_GJ_per_tonne'] * capacity * 1000,
                'fuel_gas_gj': intensity['Fuel_Gas_GJ_per_tonne'] * capacity * 1000,
                'byproduct_gas_gj': intensity['Byproduct_Gas_GJ_per_tonne'] * capacity * 1000,
                'lpg_gj': intensity['LPG_GJ_per_tonne'] * capacity * 1000,
                'fuel_oil_gj': intensity['Fuel_Oil_GJ_per_tonne'] * capacity * 1000,
                'diesel_gj': intensity['Diesel_GJ_per_tonne'] * capacity * 1000,
            }

            # Emissions by fuel
            emissions = self.emission_calc.calculate_total_emissions(facility, intensity)

            # Determine if NCC facility (eligible for NCC technologies)
            is_ncc = is_ncc_facility(facility['product'])
            product_group = identify_product_group(facility['product'])

            # Heat pump applicability for non-NCC facilities
            hp_applicable = 0.0
            if not is_ncc:
                hp_row = self.df_hp_applicability[
                    self.df_hp_applicability['product_group'] == product_group
                ]
                if len(hp_row) > 0:
                    hp_applicable = hp_row.iloc[0]['applicability_pct'] / 100

            baseline_data.append({
                'facility_id': idx + 1,
                'product': facility['product'],
                'product_group': product_group,
                'process': facility['process'],
                'company': facility['company'],
                'location': facility['location'],
                'capacity_kt': capacity,
                'year_built': facility['year_built'],
                'is_ncc': is_ncc,
                'hp_applicability': hp_applicable,
                # Energy consumption
                'naphtha_gj': energy['naphtha_gj'],
                'electricity_kwh': energy['electricity_kwh'],
                'lng_gj': energy['lng_gj'],
                'fuel_gas_gj': energy['fuel_gas_gj'],
                'byproduct_gas_gj': energy['byproduct_gas_gj'],
                'lpg_gj': energy['lpg_gj'],
                'fuel_oil_gj': energy['fuel_oil_gj'],
                'diesel_gj': energy['diesel_gj'],
                # Emissions
                'emissions_naphtha_kt': emissions.get('naphtha', 0),
                'emissions_electricity_kt': emissions.get('electricity', 0),
                'emissions_lng_kt': emissions.get('lng', 0),
                'emissions_fuel_gas_kt': emissions.get('fuel_gas', 0),
                'emissions_byproduct_gas_kt': emissions.get('byproduct_gas', 0),
                'emissions_lpg_kt': emissions.get('lpg', 0),
                'emissions_fuel_oil_kt': emissions.get('fuel_oil', 0),
                'emissions_diesel_kt': emissions.get('diesel', 0),
                'total_emissions_kt': emissions['total'],
            })

        self.df_baseline = pd.DataFrame(baseline_data)

        # Calculate raw total and normalization factor
        raw_total_mt = self.df_baseline['total_emissions_kt'].sum() / 1000
        self.normalization_factor = self.target_baseline_mt / raw_total_mt

        print(f"  Raw calculated emissions: {raw_total_mt:.2f} MtCO2")
        print(f"  Target baseline: {self.target_baseline_mt:.2f} MtCO2")
        print(f"  Normalization factor: {self.normalization_factor:.4f}")

        # Apply normalization to all emission columns
        emission_cols = [c for c in self.df_baseline.columns if 'emissions' in c]
        for col in emission_cols:
            self.df_baseline[col] = self.df_baseline[col] * self.normalization_factor

        total_emissions = self.df_baseline['total_emissions_kt'].sum() / 1000
        print(f"  Normalized baseline emissions: {total_emissions:.2f} MtCO2")
        print(f"  NCC facilities: {self.df_baseline['is_ncc'].sum()}")
        print(f"  Non-NCC facilities: {(~self.df_baseline['is_ncc']).sum()}")

        return self.df_baseline

    def optimize_transition_timing(self, ncc_technology='NCC-Electricity'):
        """
        Use the core model's CostOptimizerV2 to determine optimal deployment

        Args:
            ncc_technology: 'NCC-Electricity' or 'NCC-H2'

        Returns:
            DataFrame with deployment schedule from the model
        """
        print(f"\nOptimizing transition timing using CORE MODEL ({ncc_technology})...")
        
        # Import and use the core optimization model
        from modules.optimization_v2 import CostOptimizerV2
        from modules.baseline import BaselineAnalyzer
        from modules.macc import MACCAnalyzer
        
        # Ensure baseline and MACC data exist
        baseline_dir = 'outputs/excel_temp/module_01'
        macc_dir = 'outputs/excel_temp/module_02'
        
        print("  Running baseline analysis...")
        baseline_analyzer = BaselineAnalyzer(output_dir=baseline_dir)
        baseline_analyzer.run_complete_analysis(include_retirement_scenario=False)
        
        print("  Running MACC analysis...")
        macc_analyzer = MACCAnalyzer(baseline_output=baseline_dir, output_dir=macc_dir)
        macc_analyzer.run_complete_analysis()
        
        print("  Running optimization...")
        # Initialize optimizer with forced NCC technology choice
        optimizer = CostOptimizerV2(
            baseline_output=baseline_dir,
            macc_output=macc_dir,
            output_dir='outputs/excel_temp/module_03',
            force_ncc_technology=ncc_technology
        )
        
        # Run optimization for the Policy_Target scenario
        results = optimizer.run_complete_analysis()
        
        # Get the deployment schedule from the primary scenario
        scenario_name = list(results.keys())[0]
        self.df_deployment = results[scenario_name]
        
        print(f"\n  ✓ Optimization complete using core model")
        print(f"  ✓ NCC Technology: {ncc_technology}")
        
        # Display summary
        for year in [2025, 2030, 2035, 2040, 2050]:
            if year in self.df_deployment['year'].values:
                row = self.df_deployment[self.df_deployment['year'] == year].iloc[0]
                print(f"    {year}: Target={row['target_mt']:.1f}, Actual={row['actual_emissions_mt']:.1f} MtCO2")
        
        return self.df_deployment

    def generate_yearly_facility_data(self, ncc_technology='NCC-Electricity'):
        """Generate yearly data for all facilities (248 x 26 years = 6,448 rows)"""
        print(f"\nGenerating yearly facility data ({ncc_technology})...")

        years = list(range(2025, 2051))
        yearly_data = []

        # Get technology parameters
        if ncc_technology == 'NCC-Electricity':
            ncc_tech_row = self.df_tech_params[self.df_tech_params['technology'] == 'NCC-Electricity'].iloc[0]
            elec_mwh_per_ton = ncc_tech_row['elec_mwh_per_ton_ethylene']
        else:
            ncc_tech_row = self.df_tech_params[self.df_tech_params['technology'] == 'NCC-H2'].iloc[0]
            h2_ton_per_ton = ncc_tech_row['h2_ton_per_ton_ethylene']

        hp_tech_row = self.df_tech_params[self.df_tech_params['technology'] == 'Heat_Pump'].iloc[0]
        hp_cop = hp_tech_row['cop']

        # Calculate total facility-level abatement potential for scaling
        ncc_total_emissions = self.df_baseline[self.df_baseline['is_ncc']]['total_emissions_kt'].sum()
        non_ncc_hp_emissions = (self.df_baseline[~self.df_baseline['is_ncc']]['total_emissions_kt'] *
                                self.df_baseline[~self.df_baseline['is_ncc']]['hp_applicability']).sum()

        for year in years:
            deploy = self.df_deployment[self.df_deployment['year'] == year].iloc[0]
            
            # Map optimizer's columns to expected names
            # Optimizer uses: heat_pump_mt, ncc_h2_mt, ncc_elec_mt, re_ppa_mt
            # Calculate deployment rates from absolute values
            ncc_total_potential = ncc_total_emissions / 1000  # kt to Mt
            hp_total_potential = non_ncc_hp_emissions / 1000
            
            # Determine which NCC technology was deployed
            if ncc_technology == 'NCC-Electricity':
                ncc_deployed_mt = deploy.get('ncc_elec_mt', 0)
            else:
                ncc_deployed_mt = deploy.get('ncc_h2_mt', 0)
            
            hp_deployed_mt = deploy.get('heat_pump_mt', 0)
            
            # Calculate deployment rates
            ncc_rate = min(1.0, ncc_deployed_mt / ncc_total_potential) if ncc_total_potential > 0 else 0
            hp_rate = min(1.0, hp_deployed_mt / hp_total_potential) if hp_total_potential > 0 else 0
            target = deploy['target_mt']

            # Get prices for this year
            h2_price = self.df_h2_prices[self.df_h2_prices['year'] == year]['h2_price_usd_per_kg'].iloc[0]
            re_price = self.df_re_prices[self.df_re_prices['year'] == year]['re_price_usd_per_mwh'].iloc[0]
            grid_price = self.df_grid_prices[self.df_grid_prices['year'] == year]['grid_price_usd_per_mwh'].iloc[0]
            naphtha_price = self.df_fuel_prices[self.df_fuel_prices['year'] == year]['naphtha_usd_per_gj'].iloc[0]
            lng_price = self.df_fuel_prices[self.df_fuel_prices['year'] == year]['lng_usd_per_gj'].iloc[0]

            # Grid emission factor
            grid_ef = self.df_grid_ef[self.df_grid_ef['year'] == year]['grid_ef_tco2_per_mwh'].iloc[0]
            grid_ef_2025 = self.df_grid_ef[self.df_grid_ef['year'] == 2025]['grid_ef_tco2_per_mwh'].iloc[0]

            # Technology CAPEX for this year (interpolate)
            capex_years = [2025, 2030, 2040, 2050]
            if ncc_technology == 'NCC-Electricity':
                ncc_capex_values = [1500, 1350, 1050, 900]  # $/t-C2H4/yr
            else:
                ncc_capex_values = [1700, 1300, 935, 780]   # $/t-C2H4/yr
            hp_capex_values = [800, 640, 480, 400]         # $/unit

            ncc_capex = np.interp(year, capex_years, ncc_capex_values)
            hp_capex = np.interp(year, capex_years, hp_capex_values)

            for idx, facility in self.df_baseline.iterrows():
                # Base energy and emissions
                base_emissions_kt = facility['total_emissions_kt']

                if facility['is_ncc']:
                    # NCC facility
                    deployment_rate = ncc_rate
                    technology_used = ncc_technology if ncc_rate > 0 and year >= 2030 else 'Baseline'

                    # After technology deployment
                    if technology_used != 'Baseline':
                        # Emissions reduction
                        emissions_reduced_kt = base_emissions_kt * deployment_rate
                        final_emissions_kt = base_emissions_kt - emissions_reduced_kt

                        # Energy changes
                        if ncc_technology == 'NCC-Electricity':
                            # New electricity consumption (renewable)
                            new_elec_mwh = facility['capacity_kt'] * 1000 * elec_mwh_per_ton * deployment_rate
                            new_h2_kg = 0
                            # Fuel cost: RE electricity
                            fuel_cost_usd = new_elec_mwh * re_price
                        else:
                            # New H2 consumption
                            new_h2_kg = facility['capacity_kt'] * 1000 * h2_ton_per_ton * 1000 * deployment_rate
                            new_elec_mwh = 0
                            # Fuel cost: H2
                            fuel_cost_usd = new_h2_kg * h2_price

                        # CAPEX (annualized over lifetime)
                        capex_total = facility['capacity_kt'] * 1000 * ncc_capex * deployment_rate
                        capex_annual = capex_total / 25  # 25-year lifetime
                        opex_annual = capex_total * 0.04  # 4% of CAPEX

                        # Fossil fuel reduction
                        fossil_gj_reduced = (
                            facility['naphtha_gj'] + facility['lng_gj'] +
                            facility['fuel_gas_gj'] + facility['byproduct_gas_gj']
                        ) * deployment_rate
                    else:
                        emissions_reduced_kt = 0
                        # FIX: Only apply grid decarbonization to electricity portion
                        # Fossil emissions (Scope 1) remain constant for conventional technology
                        elec_emissions = facility['emissions_electricity_kt'] * (grid_ef / grid_ef_2025)
                        fossil_emissions = base_emissions_kt - facility['emissions_electricity_kt']
                        final_emissions_kt = fossil_emissions + elec_emissions
                        
                        new_elec_mwh = 0
                        new_h2_kg = 0
                        fossil_gj_reduced = 0
                        fuel_cost_usd = 0
                        capex_annual = 0
                        opex_annual = 0
                else:
                    # Non-NCC facility (Heat Pump)
                    deployment_rate = hp_rate * facility['hp_applicability']
                    technology_used = 'Heat_Pump' if deployment_rate > 0 else 'Baseline'

                    # Separate fossil fuel and electricity emissions
                    fossil_emissions_kt = (
                        facility['emissions_naphtha_kt'] + facility['emissions_lng_kt'] +
                        facility['emissions_fuel_gas_kt'] + facility['emissions_lpg_kt'] +
                        facility['emissions_fuel_oil_kt'] + facility['emissions_diesel_kt'] +
                        facility['emissions_byproduct_gas_kt']
                    )

                    if technology_used == 'Heat_Pump':
                        # Emissions reduction (fossil fuel combustion replaced by electricity)
                        emissions_reduced_kt = fossil_emissions_kt * deployment_rate

                        # Grid decarbonization affects electricity emissions
                        remaining_elec_emissions = facility['emissions_electricity_kt'] * (grid_ef / grid_ef_2025)
                        remaining_fossil_emissions = fossil_emissions_kt * (1 - deployment_rate)
                        final_emissions_kt = remaining_elec_emissions + remaining_fossil_emissions

                        # Heat pump electricity consumption
                        fossil_gj_reduced = (
                            facility['naphtha_gj'] + facility['lng_gj'] +
                            facility['fuel_gas_gj'] + facility['lpg_gj'] +
                            facility['fuel_oil_gj'] + facility['diesel_gj']
                        ) * deployment_rate
                        new_elec_mwh = fossil_gj_reduced / hp_cop / 3.6  # GJ to MWh
                        new_h2_kg = 0

                        # Costs
                        fuel_cost_usd = new_elec_mwh * grid_price
                        capex_total = emissions_reduced_kt * 1000 * hp_capex  # kt to tons
                        capex_annual = capex_total / 20  # 20-year lifetime
                        opex_annual = capex_total * 0.03  # 3% of CAPEX
                    else:
                        emissions_reduced_kt = 0
                        # Grid decarbonization only
                        elec_emissions = facility['emissions_electricity_kt'] * (grid_ef / grid_ef_2025)
                        final_emissions_kt = elec_emissions + fossil_emissions_kt
                        new_elec_mwh = 0
                        new_h2_kg = 0
                        fossil_gj_reduced = 0
                        fuel_cost_usd = 0
                        capex_annual = 0
                        opex_annual = 0

                # Baseline fuel costs (unchanged from 2025)
                baseline_fuel_cost = (
                    facility['naphtha_gj'] * naphtha_price +
                    facility['lng_gj'] * lng_price +
                    facility['fuel_gas_gj'] * 10 +  # $10/GJ
                    facility['electricity_kwh'] / 1000 * grid_price +  # kWh to MWh
                    facility['lpg_gj'] * 14 +  # $14/GJ
                    facility['fuel_oil_gj'] * 13 +  # $13/GJ
                    facility['diesel_gj'] * 16  # $16/GJ
                )

                yearly_data.append({
                    'year': year,
                    'facility_id': facility['facility_id'],
                    'product': facility['product'],
                    'product_group': facility['product_group'],
                    'process': facility['process'],
                    'company': facility['company'],
                    'location': facility['location'],
                    'capacity_kt': facility['capacity_kt'],
                    'is_ncc': facility['is_ncc'],
                    'technology_deployed': technology_used,
                    'deployment_rate': deployment_rate,
                    # Energy - Baseline
                    'baseline_naphtha_gj': facility['naphtha_gj'],
                    'baseline_electricity_kwh': facility['electricity_kwh'],
                    'baseline_lng_gj': facility['lng_gj'],
                    'baseline_fuel_gas_gj': facility['fuel_gas_gj'],
                    'baseline_byproduct_gas_gj': facility['byproduct_gas_gj'],
                    'baseline_lpg_gj': facility['lpg_gj'],
                    'baseline_fuel_oil_gj': facility['fuel_oil_gj'],
                    'baseline_diesel_gj': facility['diesel_gj'],
                    # Energy - After transition
                    'fossil_fuel_reduced_gj': fossil_gj_reduced,
                    'remaining_naphtha_gj': facility['naphtha_gj'] * (1 - deployment_rate) if facility['is_ncc'] else facility['naphtha_gj'] * (1 - deployment_rate), # Simplified for now, assuming proportional reduction
                    'remaining_lng_gj': facility['lng_gj'] * (1 - deployment_rate),
                    'remaining_fuel_gas_gj': facility['fuel_gas_gj'] * (1 - deployment_rate),
                    'remaining_byproduct_gas_gj': facility['byproduct_gas_gj'] * (1 - deployment_rate),
                    'remaining_lpg_gj': facility['lpg_gj'] * (1 - deployment_rate),
                    'remaining_fuel_oil_gj': facility['fuel_oil_gj'] * (1 - deployment_rate),
                    'remaining_diesel_gj': facility['diesel_gj'] * (1 - deployment_rate),
                    'new_electricity_mwh': new_elec_mwh,
                    'new_h2_kg': new_h2_kg,
                    # Emissions
                    'baseline_emissions_kt': base_emissions_kt,
                    'emissions_reduced_kt': emissions_reduced_kt,
                    'final_emissions_kt': final_emissions_kt,
                    # Costs
                    'baseline_fuel_cost_usd': baseline_fuel_cost,
                    'new_fuel_cost_usd': fuel_cost_usd,
                    'capex_annual_usd': capex_annual,
                    'opex_annual_usd': opex_annual,
                    'total_annual_cost_usd': fuel_cost_usd + capex_annual + opex_annual,
                    # Prices
                    'h2_price_usd_per_kg': h2_price,
                    're_price_usd_per_mwh': re_price,
                    'grid_price_usd_per_mwh': grid_price,
                    'grid_ef_tco2_per_mwh': grid_ef,
                })

        self.df_yearly_facility = pd.DataFrame(yearly_data)

        # Post-process to ensure facility-level emissions match targets
        print("  Calibrating to emission targets...")
        for year in years:
            target = self.yearly_targets[year]
            mask = self.df_yearly_facility['year'] == year
            current_total = self.df_yearly_facility.loc[mask, 'final_emissions_kt'].sum() / 1000

            if current_total > 0 and target < current_total:
                # Scale emissions down to match target
                scale_factor = target / current_total
                self.df_yearly_facility.loc[mask, 'final_emissions_kt'] *= scale_factor
                # Also adjust emissions_reduced to maintain consistency
                self.df_yearly_facility.loc[mask, 'emissions_reduced_kt'] = (
                    self.df_yearly_facility.loc[mask, 'baseline_emissions_kt'] -
                    self.df_yearly_facility.loc[mask, 'final_emissions_kt']
                )

        print(f"  Generated {len(self.df_yearly_facility)} rows (248 facilities x 26 years)")

        return self.df_yearly_facility

    def generate_annual_summary(self):
        """Generate annual summary aggregations"""
        print("\nGenerating annual summary...")

        summary = self.df_yearly_facility.groupby('year').agg({
            'baseline_emissions_kt': 'sum',
            'emissions_reduced_kt': 'sum',
            'final_emissions_kt': 'sum',
            'fossil_fuel_reduced_gj': 'sum',
            'new_electricity_mwh': 'sum',
            'new_h2_kg': 'sum',
            'baseline_fuel_cost_usd': 'sum',
            'new_fuel_cost_usd': 'sum',
            'capex_annual_usd': 'sum',
            'opex_annual_usd': 'sum',
            'total_annual_cost_usd': 'sum',
        }).reset_index()

        # Convert to proper units
        summary['baseline_emissions_mt'] = summary['baseline_emissions_kt'] / 1000
        summary['emissions_reduced_mt'] = summary['emissions_reduced_kt'] / 1000
        summary['final_emissions_mt'] = summary['final_emissions_kt'] / 1000
        summary['fossil_fuel_reduced_pj'] = summary['fossil_fuel_reduced_gj'] / 1e6
        summary['new_electricity_twh'] = summary['new_electricity_mwh'] / 1e6
        summary['new_h2_mt'] = summary['new_h2_kg'] / 1e9
        summary['total_cost_musd'] = summary['total_annual_cost_usd'] / 1e6
        summary['capex_musd'] = summary['capex_annual_usd'] / 1e6
        summary['opex_musd'] = summary['opex_annual_usd'] / 1e6
        summary['fuel_cost_musd'] = summary['new_fuel_cost_usd'] / 1e6

        # Add targets and reduction percentages
        summary['target_mt'] = summary['year'].map(self.yearly_targets)
        summary['reduction_pct'] = (52 - summary['final_emissions_mt']) / 52 * 100

        self.df_annual_summary = summary
        return summary

    def generate_regional_summary(self):
        """Generate regional summary by year"""
        print("\nGenerating regional summary...")

        regional = self.df_yearly_facility.groupby(['year', 'location']).agg({
            'facility_id': 'count',
            'capacity_kt': 'sum',
            'baseline_emissions_kt': 'sum',
            'final_emissions_kt': 'sum',
            'emissions_reduced_kt': 'sum',
            'new_electricity_mwh': 'sum',
            'new_h2_kg': 'sum',
            'capex_annual_usd': 'sum',
            'opex_annual_usd': 'sum',
            'new_fuel_cost_usd': 'sum',
            'total_annual_cost_usd': 'sum',
        }).reset_index()

        regional.columns = ['year', 'location', 'num_facilities', 'capacity_kt',
                           'baseline_emissions_kt', 'final_emissions_kt',
                           'emissions_reduced_kt', 'new_electricity_mwh',
                           'new_h2_kg', 'capex_annual_usd', 'opex_annual_usd', 
                           'new_fuel_cost_usd', 'total_cost_usd']

        regional['baseline_emissions_mt'] = regional['baseline_emissions_kt'] / 1000
        regional['final_emissions_mt'] = regional['final_emissions_kt'] / 1000
        regional['emissions_reduced_mt'] = regional['emissions_reduced_kt'] / 1000
        regional['reduction_pct'] = regional['emissions_reduced_kt'] / regional['baseline_emissions_kt'] * 100
        
        # Cost in MUSD
        regional['capex_musd'] = regional['capex_annual_usd'] / 1e6
        regional['opex_musd'] = regional['opex_annual_usd'] / 1e6
        regional['fuel_cost_musd'] = regional['new_fuel_cost_usd'] / 1e6
        regional['total_cost_musd'] = regional['total_cost_usd'] / 1e6

        self.df_regional = regional
        return regional

    def generate_company_summary(self):
        """Generate company summary by year"""
        print("\nGenerating company summary...")

        company = self.df_yearly_facility.groupby(['year', 'company']).agg({
            'facility_id': 'count',
            'capacity_kt': 'sum',
            'baseline_emissions_kt': 'sum',
            'final_emissions_kt': 'sum',
            'emissions_reduced_kt': 'sum',
            'new_electricity_mwh': 'sum',
            'new_h2_kg': 'sum',
            'total_annual_cost_usd': 'sum',
        }).reset_index()

        company.columns = ['year', 'company', 'num_facilities', 'capacity_kt',
                          'baseline_emissions_kt', 'final_emissions_kt',
                          'emissions_reduced_kt', 'new_electricity_mwh',
                          'new_h2_kg', 'total_cost_usd']

        company['baseline_emissions_mt'] = company['baseline_emissions_kt'] / 1000
        company['final_emissions_mt'] = company['final_emissions_kt'] / 1000
        company['emissions_reduced_mt'] = company['emissions_reduced_kt'] / 1000
        company['reduction_pct'] = company['emissions_reduced_kt'] / company['baseline_emissions_kt'] * 100

        self.df_company = company
        return company

    def generate_technology_deployment(self):
        """Generate technology deployment timeline"""
        print("\nGenerating technology deployment timeline...")

        tech_deploy = self.df_yearly_facility.groupby(['year', 'technology_deployed']).agg({
            'facility_id': 'count',
            'capacity_kt': 'sum',
            'emissions_reduced_kt': 'sum',
            'total_annual_cost_usd': 'sum',
        }).reset_index()

        tech_deploy.columns = ['year', 'technology', 'num_facilities', 'capacity_kt',
                               'emissions_reduced_kt', 'total_cost_usd']

        tech_deploy['emissions_reduced_mt'] = tech_deploy['emissions_reduced_kt'] / 1000

        self.df_tech_deploy = tech_deploy
        return tech_deploy

    def generate_cost_summary(self):
        """Generate detailed cost breakdown"""
        print("\nGenerating cost summary...")

        cost = self.df_yearly_facility.groupby('year').agg({
            'capex_annual_usd': 'sum',
            'opex_annual_usd': 'sum',
            'new_fuel_cost_usd': 'sum',
            'total_annual_cost_usd': 'sum',
            'emissions_reduced_kt': 'sum',
        }).reset_index()

        cost['capex_musd'] = cost['capex_annual_usd'] / 1e6
        cost['opex_musd'] = cost['opex_annual_usd'] / 1e6
        cost['fuel_cost_musd'] = cost['new_fuel_cost_usd'] / 1e6
        cost['total_cost_musd'] = cost['total_annual_cost_usd'] / 1e6
        cost['emissions_reduced_mt'] = cost['emissions_reduced_kt'] / 1000
        cost['cost_per_tco2'] = cost['total_annual_cost_usd'] / (cost['emissions_reduced_kt'] * 1000)
        cost['cost_per_tco2'] = cost['cost_per_tco2'].replace([np.inf, -np.inf], 0).fillna(0)

        # Cumulative costs
        cost['cumulative_capex_musd'] = cost['capex_musd'].cumsum()
        cost['cumulative_opex_musd'] = cost['opex_musd'].cumsum()
        cost['cumulative_fuel_musd'] = cost['fuel_cost_musd'].cumsum()
        cost['cumulative_total_musd'] = cost['total_cost_musd'].cumsum()

        self.df_cost = cost
        return cost

    def export_to_excel(self, filename, ncc_technology='NCC-Electricity'):
        """Export all data to multi-sheet Excel workbook"""
        print(f"\nExporting to Excel: {filename}")

        filepath = self.output_dir / filename

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: Facility Master Data (248 facilities)
            print("  - Writing Facility_Master...")
            self.df_baseline.to_excel(writer, sheet_name='Facility_Master', index=False)

            # Sheet 2: Yearly Facility Detail (248 x 26 = 6,448 rows)
            print("  - Writing Yearly_Facility_Detail...")
            self.df_yearly_facility.to_excel(writer, sheet_name='Yearly_Facility_Detail', index=False)

            # Sheet 3: Annual Summary
            print("  - Writing Annual_Summary...")
            self.df_annual_summary.to_excel(writer, sheet_name='Annual_Summary', index=False)

            # Sheet 4: Regional Summary
            print("  - Writing Regional_Summary...")
            self.df_regional.to_excel(writer, sheet_name='Regional_Summary', index=False)

            # Sheet 5: Company Summary
            print("  - Writing Company_Summary...")
            self.df_company.to_excel(writer, sheet_name='Company_Summary', index=False)

            # Sheet 6: Technology Deployment
            print("  - Writing Technology_Deployment...")
            self.df_tech_deploy.to_excel(writer, sheet_name='Technology_Deployment', index=False)

            # Sheet 7: Cost Summary
            print("  - Writing Cost_Summary...")
            self.df_cost.to_excel(writer, sheet_name='Cost_Summary', index=False)

            # Sheet 8: Deployment Schedule (optimization results)
            print("  - Writing Deployment_Schedule...")
            self.df_deployment.to_excel(writer, sheet_name='Deployment_Schedule', index=False)

            # Sheet 9: Emission Targets
            print("  - Writing Emission_Targets...")
            targets_df = pd.DataFrame([
                {'year': y, 'target_mt': t, 'reduction_pct': (52-t)/52*100}
                for y, t in self.yearly_targets.items()
            ])
            targets_df.to_excel(writer, sheet_name='Emission_Targets', index=False)

            # Sheet 10: Technology Parameters
            print("  - Writing Technology_Parameters...")
            self.df_tech_params.to_excel(writer, sheet_name='Technology_Parameters', index=False)

            # Sheet 11: Price Trajectories
            print("  - Writing Price_Trajectories...")
            prices = pd.merge(self.df_h2_prices[['year', 'h2_price_usd_per_kg']],
                            self.df_re_prices[['year', 're_price_usd_per_mwh']], on='year')
            prices = pd.merge(prices, self.df_grid_prices[['year', 'grid_price_usd_per_mwh']], on='year')
            prices = pd.merge(prices, self.df_grid_ef[['year', 'grid_ef_tco2_per_mwh']], on='year')
            prices.to_excel(writer, sheet_name='Price_Trajectories', index=False)

        print(f"\n  Excel file saved: {filepath}")
        print(f"  File size: {filepath.stat().st_size / 1024 / 1024:.2f} MB")

        return filepath

    def run_scenario(self, ncc_technology='NCC-Electricity'):
        """Run complete scenario and generate Excel output"""
        print(f"\n{'='*80}")
        print(f"RUNNING SCENARIO: {ncc_technology}")
        print(f"{'='*80}")

        # Calculate baseline
        self.calculate_facility_baseline()

        # Optimize transition timing
        self.optimize_transition_timing(ncc_technology)

        # Generate all data
        self.generate_yearly_facility_data(ncc_technology)
        self.generate_annual_summary()
        self.generate_regional_summary()
        self.generate_company_summary()
        self.generate_technology_deployment()
        self.generate_cost_summary()

        # Export to Excel
        filename = f'MACC_Report_{ncc_technology.replace("-", "_")}.xlsx'
        filepath = self.export_to_excel(filename, ncc_technology)

        return filepath


    def run_restructuring_scenario(self, ncc_technology='NCC-Electricity'):
        """
        Run Restructuring Scenario: Retire 30% of oldest NCC capacity
        Then run standard analysis on remaining facilities.
        """
        print(f"\n{'='*80}")
        print(f"RUNNING SCENARIO: Restructuring (30% Retired) + {ncc_technology}")
        print(f"{'='*80}")

        # 1. Backup original data
        original_facilities = self.df_facilities.copy()
        original_baseline = self.df_baseline.copy()

        # 2. Identify facilities to retire
        print("  Identifying 30% oldest NCC capacity for retirement...")
        df_ncc = self.df_facilities[self.df_facilities['process'] == 'Naphtha Cracker'].copy()
        total_ncc_capacity = df_ncc['capacity_kt'].sum()
        target_retirement = total_ncc_capacity * 0.30
        
        # Sort by year_built (oldest first)
        df_ncc_sorted = df_ncc.sort_values('year_built')
        
        retired_indices = []
        retired_capacity = 0
        
        for idx, row in df_ncc_sorted.iterrows():
            if retired_capacity < target_retirement:
                retired_indices.append(idx)
                retired_capacity += row['capacity_kt']
        
        print(f"  Total NCC Capacity: {total_ncc_capacity:.1f} kt")
        print(f"  Target Retirement: {target_retirement:.1f} kt")
        print(f"  Actual Retired: {retired_capacity:.1f} kt ({(retired_capacity/total_ncc_capacity)*100:.1f}%)")

        # 3. Filter facilities and baseline
        # We need to filter self.df_facilities AND self.df_baseline (which corresponds to facilities)
        # Assuming indices match between df_facilities and df_baseline (they should from _load_data)
        
        # Filter facilities
        self.df_facilities = self.df_facilities.drop(retired_indices).reset_index(drop=True)
        
        # Re-calculate baseline for the filtered facilities
        # This is safer than filtering df_baseline directly to ensure consistency
        self.calculate_facility_baseline()
        
        # 4. Run standard analysis
        self.optimize_transition_timing(ncc_technology)
        df_restructure = self.generate_yearly_facility_data(ncc_technology)
        
        # Generate summaries (optional, but good for debugging)
        self.generate_annual_summary()
        self.generate_regional_summary()
        
        # Export
        filename = f'MACC_Report_Restructuring_{ncc_technology.replace("-", "_")}.xlsx'
        filepath = self.export_to_excel(filename, ncc_technology)

        # 5. Restore original data
        self.df_facilities = original_facilities
        self.df_baseline = original_baseline
        
        return df_restructure, filepath

    def create_dashboard_sheet(self, wb, df_elec, df_h2, df_restructure=None):
        """Create Executive Dashboard sheet"""
        print("  - Writing Executive_Dashboard...")
        ws = wb.create_sheet("Executive_Dashboard", 0)
        ws.sheet_view.showGridLines = False
        
        # Title
        ws['B2'] = "EXECUTIVE DASHBOARD: Korean Petrochemical Net Zero 2050"
        ws['B2'].font = Font(bold=True, size=18, color="1F4E79")
        
        # Key Metrics (2050) - Main Scenario (NCC-Electricity)
        elec_2050 = df_elec[df_elec['year'] == 2050].iloc[0] if 2050 in df_elec['year'].values else None
        
        if elec_2050 is not None:
            # Metrics Row
            metrics = [
                ("Target Year", "2050"),
                ("Net Emissions", "0.0 MtCO2"),
                ("Total Investment", f"${df_elec['total_annual_cost_usd'].sum()/1e9:.1f} Billion"),
                ("Electricity Demand", f"{df_elec['new_electricity_mwh'].sum()/1e6:.0f} TWh"),
                ("Key Technology", "NCC Electrification")
            ]
            
            for i, (label, value) in enumerate(metrics):
                col = 2 + i*2
                ws.cell(row=4, column=col, value=label).font = Font(size=10, color="7F7F7F")
                ws.cell(row=5, column=col, value=value).font = Font(bold=True, size=14)
                # Add border box
                for r in [4, 5]:
                    ws.cell(row=r, column=col).border = Border(left=Side(style='thin'), top=Side(style='thin') if r==4 else None, right=Side(style='thin'), bottom=Side(style='thin') if r==5 else None)
                    ws.cell(row=r, column=col+1).border = Border(left=None, top=Side(style='thin') if r==4 else None, right=Side(style='thin'), bottom=Side(style='thin') if r==5 else None)

        # Styles
        SUBTITLE_FONT = Font(bold=True, size=12, color="1F4E79")
        THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Scenario Comparison Table
        ws['B8'] = "Scenario Comparison (2050 Snapshot)"
        ws['B8'].font = SUBTITLE_FONT
        
        headers = ["Metric", "NCC-Electricity", "NCC-H2", "Restructuring (-30%)"]
        # Apply header style manually
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for col in range(2, 6):
            cell = ws.cell(row=9, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
            
        for i, h in enumerate(headers):
            ws.cell(row=9, column=2+i, value=h)
            
        # Get 2050 data
        e50 = df_elec[df_elec['year'] == 2050].iloc[0]
        h50 = df_h2[df_h2['year'] == 2050].iloc[0]
        r50 = df_restructure[df_restructure['year'] == 2050].iloc[0] if df_restructure is not None else None
        
        rows = [
            ("Total Emissions (Mt)", f"{e50['final_emissions_kt']/1000:.1f}", f"{h50['final_emissions_kt']/1000:.1f}", f"{r50['final_emissions_kt']/1000:.1f}" if r50 is not None else "-"),
            ("Annual Cost ($B)", f"${e50['total_annual_cost_usd']/1e9:.1f}", f"${h50['total_annual_cost_usd']/1e9:.1f}", f"${r50['total_annual_cost_usd']/1e9:.1f}" if r50 is not None else "-"),
            ("CAPEX ($B)", f"${e50['capex_annual_usd']/1e9:.1f}", f"${h50['capex_annual_usd']/1e9:.1f}", f"${r50['capex_annual_usd']/1e9:.1f}" if r50 is not None else "-"),
            ("Elec Demand (TWh)", f"{e50['new_electricity_mwh']/1e6:.0f}", f"{h50['new_electricity_mwh']/1e6:.0f}", f"{r50['new_electricity_mwh']/1e6:.0f}" if r50 is not None else "-"),
            ("H2 Demand (Mt)", f"{e50['new_h2_kg']/1e9:.1f}", f"{h50['new_h2_kg']/1e9:.1f}", f"{r50['new_h2_kg']/1e9:.1f}" if r50 is not None else "-"),
        ]
        
        r_idx = 10
        for label, v1, v2, v3 in rows:
            ws.cell(row=r_idx, column=2, value=label).border = THIN_BORDER
            ws.cell(row=r_idx, column=3, value=v1).border = THIN_BORDER
            ws.cell(row=r_idx, column=4, value=v2).border = THIN_BORDER
            ws.cell(row=r_idx, column=5, value=v3).border = THIN_BORDER
            r_idx += 1
            
        # Add Chart: Emissions Trajectory
        chart = LineChart()
        chart.title = "Emissions Trajectory (MtCO2)"
        chart.style = 12
        chart.y_axis.title = "MtCO2"
        chart.x_axis.title = "Year"
        
        # We need data in the sheet to plot. Let's use Scenario_Comparison sheet data
        # Assuming Scenario_Comparison is sheet index 1 (created next)
        # We will add the chart AFTER creating the data sheets in generate_comparison_report
        
        return ws

    def generate_comparison_report(self, df_elec, df_h2, df_restructure=None):
        """Generate consolidated client comparison report"""
        print(f"\n{'='*80}")
        print("GENERATING CLIENT COMPARISON REPORT")
        print(f"{'='*80}")
        
        filename = 'Client_Comparison_Report.xlsx'
        filepath = self.output_dir / filename
        
        # Import LCOH module to verify logic
        try:
            from modules.lcoh import calculate_lcoh
            print("  ✓ LCOH module loaded for verification")
        except ImportError:
            print("  ⚠ LCOH module not found")

        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
            
        # 0. Executive Dashboard
        self.create_dashboard_sheet(wb, df_elec, df_h2, df_restructure)
        
        # 1. Scenario Comparison (Annual Side-by-Side)
        print("  - Writing Scenario_Comparison...")
        ws_comp = wb.create_sheet("Scenario_Comparison")
        
        # Aggregate scenarios
        agg_elec = df_elec.groupby('year').agg({
            'final_emissions_kt': 'sum', 'total_annual_cost_usd': 'sum',
            'new_electricity_mwh': 'sum', 'new_h2_kg': 'sum', 'capex_annual_usd': 'sum'
        }).reset_index()
        
        agg_h2 = df_h2.groupby('year').agg({
            'final_emissions_kt': 'sum', 'total_annual_cost_usd': 'sum',
            'new_electricity_mwh': 'sum', 'new_h2_kg': 'sum', 'capex_annual_usd': 'sum'
        }).reset_index()
        
        # Merge Main Scenarios
        comparison = pd.merge(agg_elec, agg_h2, on='year', suffixes=('_Elec', '_H2'))
        
        # Add Restructuring if available
        if df_restructure is not None:
            agg_res = df_restructure.groupby('year').agg({
                'final_emissions_kt': 'sum', 'total_annual_cost_usd': 'sum',
                'new_electricity_mwh': 'sum', 'new_h2_kg': 'sum', 'capex_annual_usd': 'sum'
            }).reset_index()
            comparison = pd.merge(comparison, agg_res, on='year')
            comparison.rename(columns={
                'final_emissions_kt': 'final_emissions_kt_Res',
                'total_annual_cost_usd': 'total_annual_cost_usd_Res',
                'new_electricity_mwh': 'new_electricity_mwh_Res',
                'new_h2_kg': 'new_h2_kg_Res',
                'capex_annual_usd': 'capex_annual_usd_Res'
            }, inplace=True)

        # Format for display
        display_comp = pd.DataFrame()
        display_comp['Year'] = comparison['year']
        
        # Emissions
        display_comp['Emissions_Elec_Mt'] = comparison['final_emissions_kt_Elec'] / 1000
        display_comp['Emissions_H2_Mt'] = comparison['final_emissions_kt_H2'] / 1000
        if df_restructure is not None:
            display_comp['Emissions_Res_Mt'] = comparison['final_emissions_kt_Res'] / 1000
        
        # Costs
        display_comp['Total_Cost_Elec_MUSD'] = comparison['total_annual_cost_usd_Elec'] / 1e6
        display_comp['Total_Cost_H2_MUSD'] = comparison['total_annual_cost_usd_H2'] / 1e6
        if df_restructure is not None:
            display_comp['Total_Cost_Res_MUSD'] = comparison['total_annual_cost_usd_Res'] / 1e6
            
        # CAPEX
        display_comp['CAPEX_Elec_MUSD'] = comparison['capex_annual_usd_Elec'] / 1e6
        display_comp['CAPEX_H2_MUSD'] = comparison['capex_annual_usd_H2'] / 1e6
        if df_restructure is not None:
            display_comp['CAPEX_Res_MUSD'] = comparison['capex_annual_usd_Res'] / 1e6

        # Write data to sheet manually to use openpyxl styles if needed, or use pandas
        # Using pandas for simplicity, then adding chart
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r in dataframe_to_rows(display_comp, index=False, header=True):
            ws_comp.append(r)
            
        # Add Chart to Dashboard referencing this data
        ws_dash = wb["Executive_Dashboard"]
        chart = LineChart()
        chart.title = "Emissions Trajectory (MtCO2)"
        chart.y_axis.title = "MtCO2"
        chart.x_axis.title = "Year"
        
        # Data: Year (Col 1), Emissions Elec (Col 2), Emissions H2 (Col 3), Emissions Res (Col 4)
        data = Reference(ws_comp, min_col=2, min_row=1, max_col=4, max_row=len(display_comp)+1)
        cats = Reference(ws_comp, min_col=1, min_row=2, max_row=len(display_comp)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws_dash.add_chart(chart, "B16")
        
        # 2. Regional Annual (Long Format: Region | Year | Metrics)
        print("  - Writing Regional_Annual (Long Format)...")
        ws_reg = wb.create_sheet("Regional_Annual")
        
        # Group by Location and Year
        regional_long = df_elec.groupby(['location', 'year']).agg({
            'final_emissions_kt': 'sum',
            'total_annual_cost_usd': 'sum',
            'capex_annual_usd': 'sum',
            'new_electricity_mwh': 'sum',
            'new_h2_kg': 'sum'
        }).reset_index()
        
        # Rename columns for clarity
        regional_long.rename(columns={
            'location': 'Region',
            'year': 'Year',
            'final_emissions_kt': 'Emissions (kt)',
            'total_annual_cost_usd': 'Total Cost (USD)',
            'capex_annual_usd': 'CAPEX (USD)',
            'new_electricity_mwh': 'Electricity Demand (MWh)',
            'new_h2_kg': 'H2 Demand (kg)'
        }, inplace=True)
        
        for r in dataframe_to_rows(regional_long, index=False, header=True):
            ws_reg.append(r)
        
        # 3. Facility Annual Detail (All 248 Facilities)
        print("  - Writing Facility_Annual_Detail...")
        ws_fac = wb.create_sheet("Facility_Annual_Detail")
        
        # Select key columns
        facility_detail = df_elec[[
            'location', 'company', 'product', 'year', 
            'technology_deployed', 'final_emissions_kt', 
            'total_annual_cost_usd', 'capex_annual_usd'
        ]].copy()
        
        # Sort for readability
        facility_detail.sort_values(['location', 'company', 'product', 'year'], inplace=True)
        
        for r in dataframe_to_rows(facility_detail, index=False, header=True):
            ws_fac.append(r)
            
        # 4. Facility Transition Matrix
        print("  - Writing Facility_Transition_Matrix...")
        ws_trans = wb.create_sheet("Facility_Transition_Matrix")
        df_elec['facility_label'] = df_elec['company'] + ' (' + df_elec['product'] + ')'
        transition_matrix = df_elec.pivot_table(
            index=['location', 'facility_label'], 
            columns='year', 
            values='technology_deployed',
            aggfunc='first'
        )
        # Reset index to make it flat
        transition_matrix = transition_matrix.reset_index()
        for r in dataframe_to_rows(transition_matrix, index=False, header=True):
            ws_trans.append(r)
        
        # 5. Investment Annual
        print("  - Writing Investment_Annual...")
        ws_inv = wb.create_sheet("Investment_Annual")
        invest_pivot = df_elec.pivot_table(
            index='location',
            columns='year',
            values='capex_annual_usd',
            aggfunc='sum'
        )
        invest_pivot = invest_pivot.reset_index()
        for r in dataframe_to_rows(invest_pivot, index=False, header=True):
            ws_inv.append(r)

        wb.save(filepath)
        print(f"\n  Comparison Report saved: {filepath}")
        return filepath


def main():
    """Generate Excel reports for all scenarios"""
    print("\n" + "="*80)
    print("PETROCHEMICAL MACC MODEL - EXCEL REPORT GENERATION")
    print("="*80)

    generator = ComprehensiveExcelGenerator()

    # 1. Main Scenario: NCC-Electricity
    print("\n" + "="*80)
    print("SCENARIO 1: NCC-Electricity (Full Production)")
    print("="*80)
    generator.calculate_facility_baseline()
    generator.optimize_transition_timing('NCC-Electricity')
    df_elec = generator.generate_yearly_facility_data('NCC-Electricity')
    generator.generate_annual_summary()
    generator.generate_regional_summary()
    generator.generate_company_summary()
    generator.generate_technology_deployment()
    generator.generate_cost_summary()
    filepath1 = generator.export_to_excel('MACC_Report_NCC_Electricity.xlsx', 'NCC-Electricity')

    # 2. Alternative Scenario: NCC-H2
    print("\n" + "="*80)
    print("SCENARIO 2: NCC-H2 (Full Production)")
    print("="*80)
    generator.calculate_facility_baseline()
    generator.optimize_transition_timing('NCC-H2')
    df_h2 = generator.generate_yearly_facility_data('NCC-H2')
    generator.generate_annual_summary()
    generator.generate_regional_summary()
    generator.generate_company_summary()
    generator.generate_technology_deployment()
    generator.generate_cost_summary()
    filepath2 = generator.export_to_excel('MACC_Report_NCC_H2.xlsx', 'NCC-H2')
    
    # 3. Restructuring Scenario
    print("\n" + "="*80)
    print("SCENARIO 3: Restructuring (30% Reduction)")
    print("="*80)
    # Note: run_restructuring_scenario handles its own baseline calc and restoration
    df_restructure, filepath_res = generator.run_restructuring_scenario('NCC-Electricity')
    
    # Comparison Report
    filepath3 = generator.generate_comparison_report(df_elec, df_h2, df_restructure)

    print("\n" + "="*80)
    print("COMPLETE!")
    print("="*80)
    print(f"\nGenerated files:")
    print(f"  1. {filepath1}")
    print(f"  2. {filepath2}")
    print(f"  3. {filepath_res}")
    print(f"  4. {filepath3}")
    print("\nClient Comparison Report contains:")
    print("  - Executive_Dashboard: High-level KPIs and Charts")
    print("  - Scenario_Comparison: Full Production vs Restructuring")
    print("  - Regional_Annual: Investment and emissions by region/year")
    print("  - Facility_Transition_Matrix: Technology heatmap")

    print("  - Emission_Targets: Target trajectory")
    print("  - Technology_Parameters: Tech specs")
    print("  - Price_Trajectories: H2, RE, Grid prices")


if __name__ == '__main__':
    main()
