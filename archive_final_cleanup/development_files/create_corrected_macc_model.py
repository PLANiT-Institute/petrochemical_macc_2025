#!/usr/bin/env python3
"""
Create Corrected MACC Model with Naphtha Integration
Based on existing korean_petrochemical_macc_enhanced.xlsx structure
Fundamentally a cost optimization model
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import os

class CorrectedMACCModel:
    """Create corrected MACC model with naphtha integration"""

    def __init__(self):
        # Load existing Excel model
        self.excel_path = "/Users/jinsupark/jinsu-coding/petrochemical_macc_2025/organized_analysis/data/korean_petrochemical_macc_enhanced.xlsx"
        self.existing_sheets = pd.read_excel(self.excel_path, sheet_name=None)

        # Naphtha parameters from your analysis
        self.naphtha_external_ghg = 0.90  # tCO2/t naphtha (your share breakdown)
        self.naphtha_heating_value = 43.5  # GJ/t naphtha
        self.naphtha_emission_factor = self.naphtha_external_ghg / self.naphtha_heating_value  # 0.0207 tCO2/GJ
        self.bio_naphtha_emission_factor = self.naphtha_emission_factor * 0.15  # 85% reduction

        print(f"🛢️  Naphtha integration parameters:")
        print(f"   External GHG factor: {self.naphtha_external_ghg:.2f} tCO₂/t naphtha")
        print(f"   Emission factor: {self.naphtha_emission_factor:.4f} tCO₂/GJ naphtha")
        print(f"   Bio-naphtha factor: {self.bio_naphtha_emission_factor:.4f} tCO₂/GJ naphtha")

    def create_corrected_ci_matrix(self):
        """Add naphtha consumption columns to existing CI matrix"""

        print("\n📊 Creating Corrected CI Matrix with Naphtha")
        print("=" * 60)

        # Load existing CI matrix
        ci_df = self.existing_sheets['CI'].copy()

        print(f"Original CI matrix: {ci_df.shape}")
        print(f"Original columns: {list(ci_df.columns)}")

        # Translate Korean column names for clarity
        column_translation = {
            '제품': 'Product',
            '공정': 'Process',
            'LNG(GJ/t)': 'LNG_GJ_per_t',
            '부생가스(GJ/t)': 'Byproduct_Gas_GJ_per_t',
            'LPG-프로판(GJ/t)': 'LPG_Propane_GJ_per_t',
            'LPG-부탄(GJ/t)': 'LPG_Butane_GJ_per_t',
            '연료가스(Fuel gas mix)(GJ/t)': 'Fuel_Gas_Mix_GJ_per_t',
            '중유(Fuel oil)(GJ/t)': 'Fuel_Oil_GJ_per_t',
            '디젤(GJ/t)': 'Diesel_GJ_per_t',
            '전력(Baseline)(kWh/t)': 'Electricity_kWh_per_t'
        }

        ci_corrected = ci_df.rename(columns=column_translation)

        # Add naphtha consumption columns based on process type and product
        print(f"\n➕ Adding naphtha consumption columns...")

        # Initialize naphtha columns
        ci_corrected['Naphtha_Feedstock_GJ_per_t'] = 0.0
        ci_corrected['Naphtha_Thermal_GJ_per_t'] = 0.0
        ci_corrected['Bio_Naphtha_Feedstock_GJ_per_t'] = 0.0  # Alternative
        ci_corrected['Bio_Naphtha_Thermal_GJ_per_t'] = 0.0    # Alternative

        # Estimate naphtha consumption based on process type
        for i, row in ci_corrected.iterrows():
            process = row['Process']
            product = row['Product']

            if 'Naphtha Cracker' in str(process):
                # NCC processes - high naphtha consumption
                if 'Ethylene' in str(product):
                    ci_corrected.loc[i, 'Naphtha_Feedstock_GJ_per_t'] = 134.85  # 3.1 t/t × 43.5 GJ/t
                    ci_corrected.loc[i, 'Naphtha_Thermal_GJ_per_t'] = 18.5     # Furnace heating
                elif 'Propylene' in str(product):
                    ci_corrected.loc[i, 'Naphtha_Feedstock_GJ_per_t'] = 121.8   # 2.8 t/t × 43.5 GJ/t
                    ci_corrected.loc[i, 'Naphtha_Thermal_GJ_per_t'] = 16.2
                elif 'Butadiene' in str(product):
                    ci_corrected.loc[i, 'Naphtha_Feedstock_GJ_per_t'] = 134.85  # Similar to ethylene
                    ci_corrected.loc[i, 'Naphtha_Thermal_GJ_per_t'] = 18.5
                else:
                    # Other NCC products
                    ci_corrected.loc[i, 'Naphtha_Feedstock_GJ_per_t'] = 100.0   # Average
                    ci_corrected.loc[i, 'Naphtha_Thermal_GJ_per_t'] = 12.0

            elif 'BTX Plant' in str(process):
                # BTX processes - moderate naphtha consumption
                ci_corrected.loc[i, 'Naphtha_Feedstock_GJ_per_t'] = 60.9      # 1.4 t/t × 43.5 GJ/t
                ci_corrected.loc[i, 'Naphtha_Thermal_GJ_per_t'] = 0.0        # Uses electricity mainly

            elif 'Utility' in str(process):
                # Utility processes - minimal naphtha
                ci_corrected.loc[i, 'Naphtha_Feedstock_GJ_per_t'] = 2.2       # Minimal feedstock
                ci_corrected.loc[i, 'Naphtha_Thermal_GJ_per_t'] = 0.0        # Uses gas/oil

            # Bio-naphtha alternatives (initially zero, for optimization)
            ci_corrected.loc[i, 'Bio_Naphtha_Feedstock_GJ_per_t'] = 0.0
            ci_corrected.loc[i, 'Bio_Naphtha_Thermal_GJ_per_t'] = 0.0

        # Calculate total naphtha consumption for validation
        total_naphtha_feedstock = (ci_corrected['Naphtha_Feedstock_GJ_per_t'] / self.naphtha_heating_value).sum()
        total_naphtha_thermal = (ci_corrected['Naphtha_Thermal_GJ_per_t'] / self.naphtha_heating_value).sum()

        print(f"✅ Naphtha consumption added:")
        print(f"   Total feedstock naphtha: {total_naphtha_feedstock:.1f} t/t product")
        print(f"   Total thermal naphtha: {total_naphtha_thermal:.1f} t/t product")
        print(f"   New matrix shape: {ci_corrected.shape}")

        return ci_corrected

    def create_corrected_ci2_matrix(self):
        """Add naphtha emission factors to existing CI2 matrix"""

        print("\n🌍 Creating Corrected CI2 Matrix with Naphtha Emission Factors")
        print("=" * 60)

        # Load existing CI2 matrix
        ci2_df = self.existing_sheets['CI2'].copy()

        print(f"Original CI2 matrix: {ci2_df.shape}")
        print(f"Original columns: {list(ci2_df.columns)}")

        # Translate Korean column names
        ci2_translation = {
            'LNG( tCO₂/GJ )': 'LNG_tCO2_per_GJ',
            '부생가스( tCO₂/GJ )': 'Byproduct_Gas_tCO2_per_GJ',
            'LPG-프로판( tCO₂/GJ )': 'LPG_Propane_tCO2_per_GJ',
            'LPG-부탄( tCO₂/GJ )': 'LPG_Butane_tCO2_per_GJ',
            '연료가스(Fuel gas mix)( tCO₂/GJ )': 'Fuel_Gas_Mix_tCO2_per_GJ',
            '중유(Fuel oil)( tCO₂/GJ )': 'Fuel_Oil_tCO2_per_GJ',
            '디젤( tCO₂/GJ )': 'Diesel_tCO2_per_GJ',
            '전력(Baseline)( tCO₂/kWh )': 'Electricity_tCO2_per_kWh'
        }

        ci2_corrected = ci2_df.rename(columns=ci2_translation)

        # Add naphtha emission factor columns
        print(f"\n➕ Adding naphtha emission factors...")

        ci2_corrected['Naphtha_Feedstock_tCO2_per_GJ'] = self.naphtha_emission_factor
        ci2_corrected['Naphtha_Thermal_tCO2_per_GJ'] = 0.0535 + self.naphtha_emission_factor  # Combustion + external GHG
        ci2_corrected['Bio_Naphtha_Feedstock_tCO2_per_GJ'] = self.bio_naphtha_emission_factor
        ci2_corrected['Bio_Naphtha_Thermal_tCO2_per_GJ'] = self.bio_naphtha_emission_factor  # Carbon neutral combustion

        print(f"✅ Naphtha emission factors added:")
        print(f"   Conventional feedstock: {self.naphtha_emission_factor:.4f} tCO₂/GJ")
        print(f"   Conventional thermal: {0.0535 + self.naphtha_emission_factor:.4f} tCO₂/GJ")
        print(f"   Bio-naphtha feedstock: {self.bio_naphtha_emission_factor:.4f} tCO₂/GJ")
        print(f"   Bio-naphtha thermal: {self.bio_naphtha_emission_factor:.4f} tCO₂/GJ")
        print(f"   New matrix shape: {ci2_corrected.shape}")

        return ci2_corrected

    def create_corrected_macc_matrix(self):
        """Add bio-naphtha technology to existing MACC matrix"""

        print("\n💰 Creating Corrected MACC Matrix with Bio-Naphtha Technology")
        print("=" * 60)

        # Load existing MACC matrix
        macc_df = self.existing_sheets['MACC_Template'].copy()

        print(f"Original MACC matrix: {macc_df.shape}")
        print(f"Existing technologies: {len(macc_df)}")

        # Add bio-naphtha substitution technology
        bio_naphtha_tech = {
            'TechID': 'TECH_024',
            'TechName': 'Bio-Naphtha Feedstock Substitution',
            'Cost_USD_per_tCO2': 272,  # From your analysis: $272/tCO2e
            'Abatement_Potential_MtCO2_per_year': 24.8,  # 85% of 29.2 MtCO2e naphtha emissions
            'TRL': 5,  # Technology readiness level
            'Commercial_Year': 2027,
            'Notes': 'Bio-naphtha substitution for conventional naphtha feedstock - 85% emission reduction',
            'Cumulative_Abatement_MtCO2_per_year': 0  # Will be calculated
        }

        # Add to MACC matrix
        macc_corrected = pd.concat([macc_df, pd.DataFrame([bio_naphtha_tech])], ignore_index=True)

        # Recalculate cumulative abatement (sort by cost first)
        macc_corrected = macc_corrected.sort_values('Cost_USD_per_tCO2').reset_index(drop=True)
        macc_corrected['Cumulative_Abatement_MtCO2_per_year'] = macc_corrected['Abatement_Potential_MtCO2_per_year'].cumsum()

        print(f"✅ Bio-naphtha technology added:")
        print(f"   Technology: {bio_naphtha_tech['TechName']}")
        print(f"   Cost: ${bio_naphtha_tech['Cost_USD_per_tCO2']}/tCO₂")
        print(f"   Abatement potential: {bio_naphtha_tech['Abatement_Potential_MtCO2_per_year']:.1f} MtCO₂/year")
        print(f"   New matrix shape: {macc_corrected.shape}")

        return macc_corrected

    def create_corrected_techoptions_matrix(self):
        """Add bio-naphtha to TechOptions sheet"""

        print("\n⚙️  Creating Corrected TechOptions Matrix")
        print("=" * 60)

        # Load existing TechOptions
        techoptions_df = self.existing_sheets['TechOptions'].copy()

        # Add bio-naphtha technology option
        bio_naphtha_option = {
            'TechID': 'TECH_024',
            'TechName': 'Bio-Naphtha Feedstock Substitution',
            'Product': 'Naphtha Feedstock',
            'ProcessType': 'Alternative',
            'EnergyCarrier': 'Bio-Naphtha',
            'MaxPenetration': 0.40,  # 40% max by 2050
            'RampUpSharePerYear': 0.03,  # 3% per year ramp-up
            'StartYear_Commercial': 2027,
            'Notes': 'Bio-naphtha substitution - supply constrained, 85% emission reduction'
        }

        techoptions_corrected = pd.concat([techoptions_df, pd.DataFrame([bio_naphtha_option])], ignore_index=True)

        print(f"✅ Bio-naphtha option added to TechOptions")
        print(f"   Max penetration: {bio_naphtha_option['MaxPenetration']:.0%}")
        print(f"   Ramp-up rate: {bio_naphtha_option['RampUpSharePerYear']:.1%} per year")

        return techoptions_corrected

    def create_naphtha_breakdown_sheet(self):
        """Create detailed naphtha emission breakdown sheet"""

        print("\n📊 Creating Naphtha Emission Breakdown Sheet")
        print("=" * 60)

        # Your naphtha emission share breakdown
        breakdown_data = [
            {
                'Emission_Source': 'Extraction_and_Production',
                'Description': 'Upstream oil extraction, refining, naphtha production',
                'Emission_Factor_tCO2_per_tonne': 0.45,
                'Share_of_Total_%': 50.0,
                'Source_Reference': 'IEA Oil & Gas Methane Tracker 2023',
                'Applicable_to': 'All naphtha applications'
            },
            {
                'Emission_Source': 'Indirect_Emissions',
                'Description': 'Indirect emissions from energy use in production',
                'Emission_Factor_tCO2_per_tonne': 0.25,
                'Share_of_Total_%': 27.8,
                'Source_Reference': 'IPCC AR6 Working Group 3',
                'Applicable_to': 'All naphtha applications'
            },
            {
                'Emission_Source': 'Methane_Leaks',
                'Description': 'Upstream methane leakage (GWP100 basis)',
                'Emission_Factor_tCO2_per_tonne': 0.12,
                'Share_of_Total_%': 13.3,
                'Source_Reference': 'EPA GHG Inventory 2023',
                'Applicable_to': 'All naphtha applications'
            },
            {
                'Emission_Source': 'Transportation',
                'Description': 'Transportation via pipeline, tanker, truck',
                'Emission_Factor_tCO2_per_tonne': 0.08,
                'Share_of_Total_%': 8.9,
                'Source_Reference': 'IMO Fourth GHG Study 2020',
                'Applicable_to': 'All naphtha applications'
            },
            {
                'Emission_Source': 'TOTAL_EXTERNAL_GHG',
                'Description': 'Total external GHG factor for conventional naphtha',
                'Emission_Factor_tCO2_per_tonne': 0.90,
                'Share_of_Total_%': 100.0,
                'Source_Reference': 'Comprehensive lifecycle analysis',
                'Applicable_to': 'Conventional naphtha baseline'
            }
        ]

        breakdown_df = pd.DataFrame(breakdown_data)

        print(f"✅ Naphtha breakdown sheet created")
        print(f"   Total external GHG factor: {self.naphtha_external_ghg:.2f} tCO₂/t naphtha")

        return breakdown_df

    def generate_corrected_excel_model(self):
        """Generate the complete corrected Excel model"""

        print("🛢️  GENERATING CORRECTED MACC EXCEL MODEL")
        print("=" * 80)
        print("Based on existing korean_petrochemical_macc_enhanced.xlsx")
        print("Adding naphtha integration to cost optimization framework")
        print("=" * 80)

        # Create corrected matrices
        ci_corrected = self.create_corrected_ci_matrix()
        ci2_corrected = self.create_corrected_ci2_matrix()
        macc_corrected = self.create_corrected_macc_matrix()
        techoptions_corrected = self.create_corrected_techoptions_matrix()
        naphtha_breakdown = self.create_naphtha_breakdown_sheet()

        # Create new workbook based on existing structure
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add corrected sheets
        print(f"\n📊 Creating Excel sheets...")

        # 1. Corrected CI sheet
        ws_ci = wb.create_sheet("CI_Corrected")
        for r in dataframe_to_rows(ci_corrected, index=False, header=True):
            ws_ci.append(r)

        # 2. Corrected CI2 sheet
        ws_ci2 = wb.create_sheet("CI2_Corrected")
        for r in dataframe_to_rows(ci2_corrected, index=False, header=True):
            ws_ci2.append(r)

        # 3. Corrected MACC template
        ws_macc = wb.create_sheet("MACC_Template_Corrected")
        for r in dataframe_to_rows(macc_corrected, index=False, header=True):
            ws_macc.append(r)

        # 4. Corrected TechOptions
        ws_tech = wb.create_sheet("TechOptions_Corrected")
        for r in dataframe_to_rows(techoptions_corrected, index=False, header=True):
            ws_tech.append(r)

        # 5. Naphtha breakdown
        ws_breakdown = wb.create_sheet("Naphtha_Emission_Breakdown")
        for r in dataframe_to_rows(naphtha_breakdown, index=False, header=True):
            ws_breakdown.append(r)

        # 6. Copy existing sheets that don't need modification
        for sheet_name in ['source', 'utilities', 'Emissions_Target', 'Baseline_Assumptions', 'TechLinks']:
            if sheet_name in self.existing_sheets:
                ws_copy = wb.create_sheet(f"{sheet_name}_Original")
                for r in dataframe_to_rows(self.existing_sheets[sheet_name], index=False, header=True):
                    ws_copy.append(r)

        # Format headers
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for sheet in wb.worksheets:
            # Format header row
            if sheet.max_row > 0:
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Save corrected model
        filename = "Korean_Petrochemical_MACC_Model_Corrected_with_Naphtha.xlsx"
        wb.save(filename)

        # Calculate baseline correction impact
        print(f"\n📊 COST OPTIMIZATION MODEL CORRECTION SUMMARY")
        print("=" * 80)

        # Estimate total naphtha consumption from CI matrix
        naphtha_consumption_per_product = (
            ci_corrected['Naphtha_Feedstock_GJ_per_t'] + ci_corrected['Naphtha_Thermal_GJ_per_t']
        ) / self.naphtha_heating_value

        estimated_total_naphtha = naphtha_consumption_per_product.sum() * 1000  # Rough estimate
        estimated_naphtha_emissions = estimated_total_naphtha * self.naphtha_external_ghg / 1000  # MtCO2e

        print(f"✅ CORRECTED MODEL GENERATED: {filename}")
        print(f"\n🎯 THREE CORE MATRICES CORRECTED:")
        print(f"   1. CI Matrix: {ci_corrected.shape[0]} products × {ci_corrected.shape[1]} consumption types")
        print(f"   2. CI2 Matrix: {ci2_corrected.shape[0]} emission scenarios × {ci2_corrected.shape[1]} fuel types")
        print(f"   3. MACC Matrix: {macc_corrected.shape[0]} technologies × costs/abatement")

        print(f"\n🛢️  NAPHTHA INTEGRATION:")
        print(f"   • External GHG factor: {self.naphtha_external_ghg:.2f} tCO₂/t naphtha")
        print(f"   • Emission factor: {self.naphtha_emission_factor:.4f} tCO₂/GJ naphtha")
        print(f"   • Bio-naphtha reduction: 85% emission reduction")
        print(f"   • Estimated total naphtha emissions: {estimated_naphtha_emissions:.1f} MtCO₂e/year")

        print(f"\n🔄 COST OPTIMIZATION LOGIC:")
        print(f"   • Baseline = Facility_Capacity × CI_consumption × CI2_emission_factors")
        print(f"   • Technologies = MACC costs and abatement potentials")
        print(f"   • Objective: Minimize cost to achieve emission targets")
        print(f"   • Bio-naphtha: $272/tCO₂, 24.8 MtCO₂/year potential")

        return filename

def main():
    """Main execution function"""

    print("🏭 KOREAN PETROCHEMICAL MACC MODEL")
    print("CORRECTED COST OPTIMIZATION MODEL WITH NAPHTHA")
    print("=" * 80)

    # Create corrected model
    model = CorrectedMACCModel()

    # Generate corrected Excel model
    filename = model.generate_corrected_excel_model()

    print(f"\n🚀 READY FOR COST OPTIMIZATION:")
    print(f"   • Open {filename} in Excel")
    print(f"   • Use CI_Corrected for consumption data")
    print(f"   • Use CI2_Corrected for emission factors")
    print(f"   • Use MACC_Template_Corrected for technology costs")
    print(f"   • Naphtha emissions fully internalized in optimization")

if __name__ == "__main__":
    main()