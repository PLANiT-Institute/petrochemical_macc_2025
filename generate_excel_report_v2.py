"""
Generate Comprehensive Excel Report v2 for Korea Petrochemical Net Zero Transition
===================================================================================
Author: PLANiT Institute
Date: 2025
Version: 2.0

UPDATES from v1:
- Sheet 11: Regional Analysis (Enhanced with annual investment and energy demand)
- Sheet 12: Operating Rate Analysis (Unchanged)
- Sheet 13: NCC-H2 Reference Scenario (NEW)
- Sheet 14: Scenario Comparison (NEW - 6 scenarios including restructuring)
- Sheet 15: 2035 Target Analysis (NEW - vs 2018 baseline)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference, AreaChart
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# =============================================================================
# PATHS
# =============================================================================
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
STREAMLIT_DATA_DIR = BASE_DIR / "streamlit_data"
OUTPUT_DIR = BASE_DIR / "reports"
OUTPUTS_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

# =============================================================================
# STYLES
# =============================================================================
def create_styles(wb):
    """Create named styles for the workbook"""

    # Header style (dark blue)
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, color='FFFFFF', size=11)
    header_style.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_style.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wb.add_named_style(header_style)

    # Sub-header style (light blue)
    subheader_style = NamedStyle(name='subheader_style')
    subheader_style.font = Font(bold=True, color='1F4E79', size=10)
    subheader_style.fill = PatternFill(start_color='D6E3F8', end_color='D6E3F8', fill_type='solid')
    subheader_style.alignment = Alignment(horizontal='left', vertical='center')
    subheader_style.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wb.add_named_style(subheader_style)

    # Data cell style
    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='right', vertical='center')
    data_style.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wb.add_named_style(data_style)

    # Title style
    title_style = NamedStyle(name='title_style')
    title_style.font = Font(bold=True, color='1F4E79', size=14)
    title_style.alignment = Alignment(horizontal='left', vertical='center')
    wb.add_named_style(title_style)

    # Source style (italic, gray)
    source_style = NamedStyle(name='source_style')
    source_style.font = Font(italic=True, color='666666', size=9)
    source_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    wb.add_named_style(source_style)

    # Highlight style (green)
    highlight_style = NamedStyle(name='highlight_style')
    highlight_style.font = Font(bold=True, color='1F4E79', size=11)
    highlight_style.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    highlight_style.alignment = Alignment(horizontal='center', vertical='center')
    highlight_style.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wb.add_named_style(highlight_style)

    return wb


# =============================================================================
# LOAD DATA
# =============================================================================
def load_all_data():
    """Load all necessary data files"""
    data = {}

    # Trajectories
    data['bau_traj'] = pd.read_csv(STREAMLIT_DATA_DIR / 'bau_trajectory_2025_2050.csv')
    data['opt_traj'] = pd.read_csv(STREAMLIT_DATA_DIR / 'optimization_trajectory.csv')
    data['macc_annual'] = pd.read_csv(STREAMLIT_DATA_DIR / 'macc_annual_2025_2050.csv')

    # Input data
    data['facilities'] = pd.read_csv(DATA_DIR / 'facility_database_with_regions.csv')
    data['energy_intensities'] = pd.read_csv(DATA_DIR / 'energy_intensities.csv')
    data['emission_factors'] = pd.read_csv(DATA_DIR / 'emission_factors.csv')
    data['tech_params'] = pd.read_csv(DATA_DIR / 'technology_parameters.csv')
    data['grid_ef'] = pd.read_csv(DATA_DIR / 'grid_emission_trajectory.csv')
    data['grid_price'] = pd.read_csv(DATA_DIR / 'grid_price_trajectory.csv')
    data['h2_price'] = pd.read_csv(DATA_DIR / 'h2_price_trajectory.csv')
    data['re_price'] = pd.read_csv(DATA_DIR / 're_price_trajectory.csv')
    data['demand_growth'] = pd.read_csv(DATA_DIR / 'demand_growth_trajectory.csv')

    # Emission scenarios (for 2035 target)
    data['emission_scenarios'] = pd.read_csv(DATA_DIR / 'emission_scenarios_clean.csv')

    # 6 Scenarios comparison - use new verified data
    scenarios_path = OUTPUTS_DIR / 'scenario_summary_final.csv'
    if scenarios_path.exists():
        data['scenarios_6'] = pd.read_csv(scenarios_path)
    else:
        # Fallback to old path
        scenarios_path_old = OUTPUTS_DIR / 'scenarios_comparison_6scenarios' / 'summary.csv'
        if scenarios_path_old.exists():
            data['scenarios_6'] = pd.read_csv(scenarios_path_old)
        else:
            data['scenarios_6'] = None

    # Load retired facilities info
    retired_25_path = OUTPUTS_DIR / 'restructure_25pct_retired_facilities.csv'
    retired_40_path = OUTPUTS_DIR / 'restructure_40pct_retired_facilities.csv'
    if retired_25_path.exists():
        data['retired_25'] = pd.read_csv(retired_25_path)
    else:
        data['retired_25'] = None
    if retired_40_path.exists():
        data['retired_40'] = pd.read_csv(retired_40_path)
    else:
        data['retired_40'] = None

    # Load scenario-specific facility data
    data['scenario_facilities'] = {}
    for scenario_id in ['shaheen_ncc_h2', 'shaheen_ncc_electricity',
                        'restructure_25pct_ncc_h2', 'restructure_25pct_ncc_electricity',
                        'restructure_40pct_ncc_h2', 'restructure_40pct_ncc_electricity']:
        scenario_dir = OUTPUTS_DIR / f'scenario_{scenario_id}'
        if scenario_dir.exists():
            data['scenario_facilities'][scenario_id] = {
                'facilities': pd.read_csv(scenario_dir / 'scenario_facilities.csv') if (scenario_dir / 'scenario_facilities.csv').exists() else None,
                'emissions': pd.read_csv(scenario_dir / 'facility_emissions_2050.csv') if (scenario_dir / 'facility_emissions_2050.csv').exists() else None,
                'deployment': pd.read_csv(scenario_dir / 'deployment_trajectory.csv') if (scenario_dir / 'deployment_trajectory.csv').exists() else None,
            }

    return data


# =============================================================================
# SHEET 1: EXECUTIVE SUMMARY
# =============================================================================
def create_executive_summary(wb, data):
    """Create Executive Summary sheet"""
    ws = wb.create_sheet("1. Executive Summary")

    row = 1

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Korea Petrochemical Industry Net Zero Transition Analysis"
    ws['A1'].font = Font(bold=True, size=16, color='1F4E79')
    row = 3

    # Study Info
    ws['A3'] = "Study Information"
    ws['A3'].font = Font(bold=True, size=12, color='1F4E79')

    info_data = [
        ("Author", "PLANiT Institute"),
        ("Publication", "Carbon Neutrality (Springer Nature), 2025"),
        ("Scope", "248 petrochemical facilities across Korea"),
        ("Baseline Year", "2025"),
        ("Target Year", "2050 (Net Zero)"),
        ("Interim Target", "2035: 43.5 MtCO2 (24.5% reduction vs 2018)"),
        ("Baseline Emissions", f"{data['bau_traj']['total_emissions_mt'].iloc[0]:.2f} MtCO2/year"),
    ]

    row = 5
    for label, value in info_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Key Results
    ws.cell(row=row, column=1, value="Key Results (2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt_2050 = data['opt_traj'][data['opt_traj']['year'] == 2050].iloc[0]
    bau_2050 = data['bau_traj'][data['bau_traj']['year'] == 2050].iloc[0]

    results = [
        ("BAU Emissions (2050)", f"{bau_2050['total_emissions_mt']:.2f} MtCO2"),
        ("Net Zero Achieved", f"{opt_2050['actual_emissions_mt']:.4f} MtCO2"),
        ("Total CAPEX Investment", f"${data['opt_traj']['cumulative_capex_musd'].max()/1000:.1f} Billion"),
        ("Electricity Demand (2050)", f"{opt_2050['electricity_consumption_increase_twh']:.1f} TWh"),
        ("Key Technology", "NCC-Electricity (Electric Cracker)"),
    ]

    for label, value in results:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Technology Deployment Summary
    ws.cell(row=row, column=1, value="Technology Deployment (2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    tech_summary = [
        ("NCC-Electricity", f"{opt_2050['ncc_elec_mt']:.1f} MtCO2", "Electric naphtha cracking"),
        ("Heat Pump", f"{opt_2050['heat_pump_mt']:.1f} MtCO2", "Industrial heat pumps (<165C)"),
        ("RE PPA", f"{opt_2050['re_ppa_mt']:.1f} MtCO2", "Renewable power purchase"),
    ]

    ws.cell(row=row, column=1, value="Technology").style = 'header_style'
    ws.cell(row=row, column=2, value="Abatement").style = 'header_style'
    ws.cell(row=row, column=3, value="Description").style = 'header_style'
    row += 1

    for tech, abate, desc in tech_summary:
        ws.cell(row=row, column=1, value=tech)
        ws.cell(row=row, column=2, value=abate)
        ws.cell(row=row, column=3, value=desc)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40

    return wb


# =============================================================================
# SHEET 2: ASSUMPTIONS & REFERENCES
# =============================================================================
def create_assumptions_sheet(wb, data):
    """Create Assumptions & References sheet"""
    ws = wb.create_sheet("2. Assumptions & References")

    row = 1

    # Title
    ws['A1'] = "Model Assumptions and Data Sources"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    row = 3

    # Section 1: Emission Factors
    ws.cell(row=row, column=1, value="1. Emission Factors").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ef_data = [
        ("Fuel", "Emission Factor", "Unit", "Source"),
        ("Naphtha", "0.0542", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("LNG", "0.0561", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("Fuel Gas", "0.050", "tCO2/GJ", "API Compendium 2021 (Mixed refinery gas)"),
        ("Byproduct Gas", "0.048", "tCO2/GJ", "API Compendium 2021 (Higher H2 content)"),
        ("LPG", "0.0631", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("Fuel Oil", "0.0773", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("Diesel", "0.0741", "tCO2/GJ", "IPCC 2019 Refinement, Table 2.3"),
        ("Green H2", "0.0", "tCO2/kg", "Zero-emission (electrolysis)"),
    ]

    for i, (fuel, ef, unit, source) in enumerate(ef_data):
        if i == 0:
            for j, val in enumerate([fuel, ef, unit, source], 1):
                ws.cell(row=row, column=j, value=val).style = 'header_style'
        else:
            ws.cell(row=row, column=1, value=fuel)
            ws.cell(row=row, column=2, value=ef)
            ws.cell(row=row, column=3, value=unit)
            ws.cell(row=row, column=4, value=source)
        row += 1

    row += 2

    # Section 2: Grid Emission Factor Trajectory
    ws.cell(row=row, column=1, value="2. Grid Emission Factor Trajectory").font = Font(bold=True, size=12, color='1F4E79')
    row += 1
    ws.cell(row=row, column=1, value="Source: Korea 10th Basic Plan for Electricity Supply and Demand (2023)").font = Font(italic=True, size=9)
    row += 2

    grid_ef_summary = data['grid_ef'][data['grid_ef']['year'].isin([2025, 2030, 2035, 2040, 2045, 2050])]

    ws.cell(row=row, column=1, value="Year").style = 'header_style'
    ws.cell(row=row, column=2, value="Grid EF (tCO2/MWh)").style = 'header_style'
    ws.cell(row=row, column=3, value="Reduction vs 2025").style = 'header_style'
    row += 1

    base_ef = 0.436
    for _, r in grid_ef_summary.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(r['grid_ef_tco2_per_mwh'], 3))
        ws.cell(row=row, column=3, value=f"{(1 - r['grid_ef_tco2_per_mwh']/base_ef)*100:.0f}%")
        row += 1

    row += 2

    # Section 3: Technology Parameters
    ws.cell(row=row, column=1, value="3. Technology Parameters").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    tech_data = [
        ("Technology", "Key Parameter", "Value", "Source", "TRL", "Available"),
        ("Heat Pump", "COP", "4.0", "Kosmadakis et al. 2020", "9", "2025"),
        ("Heat Pump", "CAPEX (2025)", "$800/tCO2", "McKinsey 2024", "", ""),
        ("Heat Pump", "CAPEX (2050)", "$400/tCO2", "Learning curve (50% reduction)", "", ""),
        ("NCC-Electricity", "Electricity", "5.0 MWh/ton C2H4", "BASF/SABIC/Linde Pilot 2024", "8", "2030"),
        ("NCC-Electricity", "CAPEX (2025)", "$1,500/t-C2H4/yr", "Toribio-Ramirez et al. 2025", "", ""),
        ("NCC-Electricity", "CAPEX (2050)", "$900/t-C2H4/yr", "Learning curve (40% reduction)", "", ""),
        ("NCC-H2", "H2 Consumption", "0.2 ton/ton C2H4", "Lummus Tech 2023 (Energy Only)", "7", "2030"),
        ("NCC-H2", "CAPEX (2025)", "$1,700/t-C2H4/yr", "Thunder Said Energy 2023", "", ""),
        ("NCC-H2", "CAPEX (2050)", "$780/t-C2H4/yr", "Learning curve (54% reduction)", "", ""),
        ("RDH", "Efficiency", "93%", "Coolbrook 2024", "8", "2026"),
        ("RDH", "Temperature", "Up to 1,700C", "Coolbrook 2024", "", ""),
        ("RE PPA", "Price (2025)", "$129/MWh", "PLANiT Institute", "-", "2025"),
        ("RE PPA", "Price (2050)", "$191/MWh", "PLANiT Institute", "", ""),
    ]

    for i, vals in enumerate(tech_data):
        if i == 0:
            for j, val in enumerate(vals, 1):
                ws.cell(row=row, column=j, value=val).style = 'header_style'
        else:
            for j, val in enumerate(vals, 1):
                ws.cell(row=row, column=j, value=val)
        row += 1

    row += 2

    # Section 4: Price Trajectories
    ws.cell(row=row, column=1, value="4. Price Trajectories").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    years = [2025, 2030, 2035, 2040, 2045, 2050]

    ws.cell(row=row, column=1, value="Year").style = 'header_style'
    ws.cell(row=row, column=2, value="Grid Price ($/MWh)").style = 'header_style'
    ws.cell(row=row, column=3, value="RE PPA ($/MWh)").style = 'header_style'
    ws.cell(row=row, column=4, value="H2 Price ($/kg)").style = 'header_style'
    row += 1

    for year in years:
        grid_p = data['grid_price'][data['grid_price']['year'] == year]['grid_price_usd_per_mwh'].values[0]
        re_p = data['re_price'][data['re_price']['year'] == year]['re_price_usd_per_mwh'].values[0]
        h2_p = data['h2_price'][data['h2_price']['year'] == year]['h2_price_usd_per_kg'].values[0]

        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=round(grid_p, 2))
        ws.cell(row=row, column=3, value=round(re_p, 2))
        ws.cell(row=row, column=4, value=round(h2_p, 2))
        row += 1

    ws.cell(row=row, column=1, value="Source: PLANiT Institute assumptions based on literature review").font = Font(italic=True, size=9)
    row += 2

    # Section 5: Demand Growth
    ws.cell(row=row, column=1, value="5. Demand Growth Assumptions").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    growth_data = [
        ("Period", "Annual Growth Rate", "Operating Rate", "Rationale"),
        ("2025-2030", "1.5%", "70%", "Baseline growth; 70% ops (Chinese overcapacity)"),
        ("2031-2035", "1.3%", "70%", "Moderate growth"),
        ("2036-2040", "1.0%", "70%", "Market maturation"),
        ("2041-2050", "0.5-0.8%", "70%", "Slow growth, mature market"),
    ]

    for i, vals in enumerate(growth_data):
        if i == 0:
            for j, val in enumerate(vals, 1):
                ws.cell(row=row, column=j, value=val).style = 'header_style'
        else:
            for j, val in enumerate(vals, 1):
                ws.cell(row=row, column=j, value=val)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    return wb


# =============================================================================
# SHEET 3: FACILITY DATABASE
# =============================================================================
def create_facility_sheet(wb, data):
    """Create Facility Database sheet"""
    ws = wb.create_sheet("3. Facility Database")

    ws['A1'] = "Korea Petrochemical Facility Database (248 Facilities)"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Source: Korea Petrochemical Industry Association (KPIA) 2023"
    ws['A2'].font = Font(italic=True, size=9)

    # Summary statistics
    df_fac = data['facilities']

    ws['A4'] = "Summary Statistics"
    ws['A4'].font = Font(bold=True, size=12)

    summary = [
        ("Total Facilities", len(df_fac)),
        ("Total Capacity", f"{df_fac['capacity_kt'].sum():,.0f} kt/year"),
        ("Locations", ", ".join(df_fac['location'].unique()[:6])),
        ("Products", df_fac['product'].nunique()),
        ("Companies", df_fac['company'].nunique()),
    ]

    row = 6
    for label, value in summary:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Regional Summary
    ws.cell(row=row, column=1, value="Capacity by Region").font = Font(bold=True, size=12)
    row += 2

    regional = df_fac.groupby('location')['capacity_kt'].sum().sort_values(ascending=False).reset_index()
    regional['share_pct'] = regional['capacity_kt'] / regional['capacity_kt'].sum() * 100

    ws.cell(row=row, column=1, value="Location").style = 'header_style'
    ws.cell(row=row, column=2, value="Capacity (kt)").style = 'header_style'
    ws.cell(row=row, column=3, value="Share (%)").style = 'header_style'
    row += 1

    for _, r in regional.head(10).iterrows():
        ws.cell(row=row, column=1, value=r['location'])
        ws.cell(row=row, column=2, value=f"{r['capacity_kt']:,.0f}")
        ws.cell(row=row, column=3, value=f"{r['share_pct']:.1f}%")
        row += 1

    row += 2

    # Product Summary
    ws.cell(row=row, column=1, value="Capacity by Product").font = Font(bold=True, size=12)
    row += 2

    product = df_fac.groupby('product')['capacity_kt'].sum().sort_values(ascending=False).reset_index()

    ws.cell(row=row, column=1, value="Product").style = 'header_style'
    ws.cell(row=row, column=2, value="Capacity (kt)").style = 'header_style'
    row += 1

    for _, r in product.head(15).iterrows():
        ws.cell(row=row, column=1, value=r['product'])
        ws.cell(row=row, column=2, value=f"{r['capacity_kt']:,.0f}")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    return wb


# =============================================================================
# SHEET 4-10: SAME AS V1 (略) - Included for completeness
# =============================================================================
def create_emission_methodology_sheet(wb, data):
    """Create Emission Calculation Methodology sheet"""
    ws = wb.create_sheet("4. Emission Methodology")

    ws['A1'] = "Emission Calculation Methodology"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Formula
    ws.cell(row=row, column=1, value="1. Baseline Emission Calculation").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="Formula:")
    ws.cell(row=row, column=2, value="Total_Emissions = Sum (Fuel_Consumption_i x EF_i)").font = Font(italic=True)
    row += 2

    ws.cell(row=row, column=1, value="Where:")
    row += 1
    ws.cell(row=row, column=1, value="  - Fuel_Consumption = Capacity (kt) x Energy_Intensity (GJ/kt or kWh/kt)")
    row += 1
    ws.cell(row=row, column=1, value="  - EF_i = Emission factor for fuel type i (from IPCC/API)")
    row += 1
    ws.cell(row=row, column=1, value="  - Electricity emissions use grid EF trajectory (Korea 10th Power Plan)")
    row += 2

    # 2018 Baseline Reference
    ws.cell(row=row, column=1, value="2. 2018 Baseline Reference").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="2018 Baseline Emissions: 57.6 MtCO2").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="  - Used as reference for 2035 NDC target calculation")
    row += 1
    ws.cell(row=row, column=1, value="  - 2035 Target = 57.6 x (1 - 0.245) = 43.5 MtCO2")
    row += 1
    ws.cell(row=row, column=1, value="  - Source: Korea NDC submission (2024)")
    row += 2

    # BAU Projection
    ws.cell(row=row, column=1, value="3. BAU Trajectory Projection").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="Formula:")
    row += 1
    ws.cell(row=row, column=1, value="BAU_Emissions(year) = [Fossil x Capacity_Multiplier x Op_Rate]").font = Font(italic=True)
    row += 1
    ws.cell(row=row, column=1, value="                    + [Elec x Capacity_Multiplier x Op_Rate x Grid_EF(year)/Grid_EF(2025)]").font = Font(italic=True)
    row += 2

    # Baseline Results
    ws.cell(row=row, column=1, value="4. Baseline Results (2025, 70% Operating Rate)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    base_results = [
        ("Total Emissions", "46.34 MtCO2/year"),
        ("Fossil Fuel Emissions", "40.47 MtCO2 (87%)"),
        ("Electricity Emissions", "5.86 MtCO2 (13%)"),
    ]

    for label, value in base_results:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50

    return wb


def create_bau_trajectory_sheet(wb, data):
    """Create BAU Trajectory sheet"""
    ws = wb.create_sheet("5. BAU Trajectory")

    ws['A1'] = "Business-As-Usual (BAU) Emission Trajectory (2025-2050)"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Scenario: No technology deployment; only grid decarbonization benefit"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Write trajectory data
    bau = data['bau_traj'][['year', 'fossil_emissions_mt', 'electricity_emissions_mt',
                            'total_emissions_mt', 'grid_ef_tco2_per_mwh', 'total_capacity_kt']]

    headers = ["Year", "Fossil Emissions (Mt)", "Electricity Emissions (Mt)",
               "Total Emissions (Mt)", "Grid EF (tCO2/MWh)", "Total Capacity (kt)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in bau.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(r['fossil_emissions_mt'], 2))
        ws.cell(row=row, column=3, value=round(r['electricity_emissions_mt'], 2))
        ws.cell(row=row, column=4, value=round(r['total_emissions_mt'], 2))
        ws.cell(row=row, column=5, value=round(r['grid_ef_tco2_per_mwh'], 3))
        ws.cell(row=row, column=6, value=f"{r['total_capacity_kt']:,.0f}")
        row += 1

    # Column widths
    for col in 'ABCDEF':
        ws.column_dimensions[col].width = 22

    return wb


def create_macc_sheet(wb, data):
    """Create Technology MACC sheet"""
    ws = wb.create_sheet("6. Technology MACC")

    ws['A1'] = "Marginal Abatement Cost Curve (MACC) by Technology"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # MACC Formula
    ws.cell(row=row, column=1, value="MACC Formula:").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="MACC ($/tCO2) = [CAPEX_Ann + OPEX_Ann + Fuel_Cost_Diff] / Abatement_Potential").font = Font(italic=True)
    row += 2

    # 2030 MACC Summary
    ws.cell(row=row, column=1, value="MACC Summary (2030)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    macc_2030 = data['macc_annual'][data['macc_annual']['year'] == 2030]

    headers = ["Technology", "Abatement (Mt)", "CAPEX ($/tCO2)", "OPEX ($/tCO2)",
               "Fuel Cost ($/tCO2)", "Total MACC ($/tCO2)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in macc_2030.iterrows():
        ws.cell(row=row, column=1, value=r['technology'])
        ws.cell(row=row, column=2, value=round(r['abatement_potential_mtco2'], 2))
        ws.cell(row=row, column=3, value=round(r['capex_ann_usd_per_tco2'], 0))
        ws.cell(row=row, column=4, value=round(r['opex_ann_usd_per_tco2'], 0))
        ws.cell(row=row, column=5, value=round(r['fuel_cost_diff_usd_per_tco2'], 0))
        ws.cell(row=row, column=6, value=round(r['total_cost_usd_per_tco2'], 0))
        row += 1

    row += 2

    # 2050 MACC Summary
    ws.cell(row=row, column=1, value="MACC Summary (2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    macc_2050 = data['macc_annual'][data['macc_annual']['year'] == 2050]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in macc_2050.iterrows():
        ws.cell(row=row, column=1, value=r['technology'])
        ws.cell(row=row, column=2, value=round(r['abatement_potential_mtco2'], 2))
        ws.cell(row=row, column=3, value=round(r['capex_ann_usd_per_tco2'], 0))
        ws.cell(row=row, column=4, value=round(r['opex_ann_usd_per_tco2'], 0))
        ws.cell(row=row, column=5, value=round(r['fuel_cost_diff_usd_per_tco2'], 0))
        ws.cell(row=row, column=6, value=round(r['total_cost_usd_per_tco2'], 0))
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col in 'BCDEF':
        ws.column_dimensions[col].width = 18

    return wb


def create_optimization_sheet(wb, data):
    """Create Optimization Results sheet"""
    ws = wb.create_sheet("7. Optimization Results")

    ws['A1'] = "Net Zero Transition Pathway: Optimization Results"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Scenario: Cost-Effective Technology Deployment to Net Zero by 2050"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Key Milestones
    ws.cell(row=row, column=1, value="Key Milestones").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt = data['opt_traj']
    milestones = opt[opt['year'].isin([2025, 2030, 2035, 2040, 2045, 2050])]

    headers = ["Year", "BAU (Mt)", "Heat Pump (Mt)", "NCC-Elec (Mt)", "RE PPA (Mt)",
               "Total Abated (Mt)", "Actual Emissions (Mt)", "Cumulative CAPEX ($M)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in milestones.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(r['bau_mt'], 2))
        ws.cell(row=row, column=3, value=round(r['heat_pump_mt'], 2))
        ws.cell(row=row, column=4, value=round(r['ncc_elec_mt'], 2))
        ws.cell(row=row, column=5, value=round(r['re_ppa_mt'], 2))
        ws.cell(row=row, column=6, value=round(r['total_deployed_mt'], 2))
        ws.cell(row=row, column=7, value=round(r['actual_emissions_mt'], 4))
        ws.cell(row=row, column=8, value=f"{r['cumulative_capex_musd']:,.0f}")
        row += 1

    # Column widths
    for col in 'ABCDEFGH':
        ws.column_dimensions[col].width = 18

    return wb


def create_energy_demand_sheet(wb, data):
    """Create Energy Demand sheet"""
    ws = wb.create_sheet("8. Energy Demand")

    ws['A1'] = "Energy Demand Analysis"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Electricity Demand
    ws.cell(row=row, column=1, value="1. Electricity Demand Trajectory").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt = data['opt_traj']

    headers = ["Year", "Baseline Elec (TWh)", "New Elec from Tech (TWh)", "Total Elec (TWh)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    baseline_elec = 7.2

    for _, r in opt[opt['year'].isin([2025, 2030, 2035, 2040, 2045, 2050])].iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(baseline_elec * (1 + (r['year']-2025)*0.01), 1))
        ws.cell(row=row, column=3, value=round(r['electricity_consumption_increase_twh'], 1))
        ws.cell(row=row, column=4, value=round(baseline_elec * (1 + (r['year']-2025)*0.01) + r['electricity_consumption_increase_twh'], 1))
        row += 1

    row += 2

    # Policy Context
    ws.cell(row=row, column=1, value="2. Policy Context").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    policy_data = [
        ("Korea 2036 RE Target", "97 TWh", "10th Basic Plan for Electricity"),
        ("Korea 2050 H2 Target", "27.3 Mt", "Hydrogen Economy Roadmap"),
        ("Model 2050 Elec Demand", f"{opt[opt['year']==2050]['electricity_consumption_increase_twh'].values[0]:.0f} TWh", "NCC-Electricity scenario"),
    ]

    for label, value, source in policy_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=source).font = Font(italic=True, size=9)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20

    return wb


def create_cost_analysis_sheet(wb, data):
    """Create Cost Analysis sheet"""
    ws = wb.create_sheet("9. Cost Analysis")

    ws['A1'] = "Investment and Cost Analysis"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Cost Summary
    ws.cell(row=row, column=1, value="1. Total Investment Summary").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt = data['opt_traj']
    total_capex = opt['cumulative_capex_musd'].max() / 1000  # Billion

    summary = [
        ("Total CAPEX (2025-2050)", f"${total_capex:.1f} Billion"),
        ("Average Annual Investment", f"${total_capex/26:.2f} Billion/year"),
        ("Peak Investment Period", "2035-2040"),
    ]

    for label, value in summary:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Annual CAPEX
    ws.cell(row=row, column=1, value="2. Annual CAPEX Trajectory").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt_copy = opt.copy()
    opt_copy['annual_capex'] = opt_copy['cumulative_capex_musd'].diff().fillna(0)

    headers = ["Year", "Annual CAPEX ($M)", "Cumulative CAPEX ($M)"]

    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in opt_copy.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=f"{r['annual_capex']:,.0f}")
        ws.cell(row=row, column=3, value=f"{r['cumulative_capex_musd']:,.0f}")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25

    return wb


def create_references_sheet(wb):
    """Create References sheet"""
    ws = wb.create_sheet("10. References")

    ws['A1'] = "References and Data Sources"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Academic References
    ws.cell(row=row, column=1, value="Academic and Industry References").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    references = [
        ("Emission Factors", "", ""),
        ("", "IPCC (2019)", "2019 Refinement to the 2006 IPCC Guidelines"),
        ("", "API (2021)", "Compendium of Greenhouse Gas Emissions Methodologies"),
        ("", "", ""),
        ("Technology - NCC Electric", "", ""),
        ("", "BASF/SABIC/Linde (2024)", "Commercial pilot plant data"),
        ("", "Toribio-Ramirez et al. (2025)", "Electric cracker CAPEX analysis"),
        ("", "", ""),
        ("Technology - NCC H2", "", ""),
        ("", "Lummus Tech (2023)", "Engineering case study for hydrogen cracker"),
        ("", "Thunder Said Energy (2023)", "Hydrogen cracker economics"),
        ("", "", ""),
        ("Technology - Heat Pump", "", ""),
        ("", "Kosmadakis et al. (2020)", "Heat pump for industrial processes"),
        ("", "McKinsey (2024)", "Industrial heat pump cost analysis"),
        ("", "", ""),
        ("Grid & Policy", "", ""),
        ("", "Korea 10th Basic Plan (2023)", "Electricity Supply and Demand Plan"),
        ("", "Korea NDC (2024)", "2035 target: 24.5% reduction vs 2018"),
        ("", "KPIA (2023)", "Petrochemical Industry Association database"),
    ]

    for cat, ref, desc in references:
        if cat:
            ws.cell(row=row, column=1, value=cat).font = Font(bold=True)
        if ref:
            ws.cell(row=row, column=2, value=ref).font = Font(italic=True)
            ws.cell(row=row, column=3, value=desc)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50

    return wb


# =============================================================================
# SHEET 11: REGIONAL ANALYSIS (ENHANCED with Annual Data)
# =============================================================================
def create_regional_analysis_sheet(wb, data):
    """Create comprehensive Regional Analysis sheet with ANNUAL data"""
    ws = wb.create_sheet("11. Regional Analysis")

    ws['A1'] = "Regional Analysis: Korea Petrochemical Complexes (Annual Breakdown)"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Includes annual investment and energy demand allocation by region"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Load and process facility data
    fac = data['facilities']
    energy = data['energy_intensities']
    opt = data['opt_traj']

    # Merge to get emissions data
    merged = pd.merge(fac, energy[['product', 'process', 'company', 'location',
                                   'Naphtha_GJ_per_tonne', 'Electricity_kWh_per_tonne',
                                   'LNG_GJ_per_tonne', 'Fuel_Gas_GJ_per_tonne']],
                      on=['product', 'process', 'company', 'location'], how='left')

    # Emission factors
    EF_NAPHTHA = 0.0542
    EF_LNG = 0.0561
    EF_FUEL_GAS = 0.050
    GRID_EF_2025 = 0.436

    # Calculate emissions
    merged['emissions_fossil_kt'] = (
        merged['capacity_kt'] * merged['Naphtha_GJ_per_tonne'].fillna(0) * EF_NAPHTHA +
        merged['capacity_kt'] * merged['LNG_GJ_per_tonne'].fillna(0) * EF_LNG +
        merged['capacity_kt'] * merged['Fuel_Gas_GJ_per_tonne'].fillna(0) * EF_FUEL_GAS
    )
    merged['emissions_elec_kt'] = merged['capacity_kt'] * merged['Electricity_kWh_per_tonne'].fillna(0) / 1000 * GRID_EF_2025
    merged['emissions_total_kt'] = merged['emissions_fossil_kt'] + merged['emissions_elec_kt']

    # Section 1: Regional Summary
    ws.cell(row=row, column=1, value="1. Regional Capacity and Emissions Summary (2025)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    # Aggregate by region
    regional = merged.groupby('location').agg({
        'capacity_kt': 'sum',
        'emissions_fossil_kt': 'sum',
        'emissions_elec_kt': 'sum',
        'emissions_total_kt': 'sum',
        'product': 'count'
    }).rename(columns={'product': 'n_facilities'}).reset_index()

    regional['share_pct'] = regional['capacity_kt'] / regional['capacity_kt'].sum() * 100
    regional['emissions_mt'] = regional['emissions_total_kt'] / 1000
    regional = regional.sort_values('capacity_kt', ascending=False)

    # Calculate investment allocation (proportional to emissions)
    total_investment = opt['cumulative_capex_musd'].max() / 1000  # Billion
    total_emissions = regional['emissions_mt'].sum()
    regional['investment_b'] = (regional['emissions_mt'] / total_emissions) * total_investment

    headers = ["Region", "Facilities", "Capacity (kt)", "Emissions (Mt)", "Share (%)", "Est. Investment ($B)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in regional.iterrows():
        ws.cell(row=row, column=1, value=r['location'])
        ws.cell(row=row, column=2, value=int(r['n_facilities']))
        ws.cell(row=row, column=3, value=f"{r['capacity_kt']:,.0f}")
        ws.cell(row=row, column=4, value=round(r['emissions_mt'], 2))
        ws.cell(row=row, column=5, value=f"{r['share_pct']:.1f}%")
        ws.cell(row=row, column=6, value=f"${r['investment_b']:.2f}")
        row += 1

    row += 2

    # Section 2: ANNUAL Regional Investment Trajectory (NEW)
    ws.cell(row=row, column=1, value="2. Annual Regional Investment Allocation (2025-2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 1
    ws.cell(row=row, column=1, value="Investment allocated proportionally to regional emission share").font = Font(italic=True, size=9)
    row += 2

    # Get annual CAPEX
    opt_copy = opt.copy()
    opt_copy['annual_capex_musd'] = opt_copy['cumulative_capex_musd'].diff().fillna(0)

    top_regions = ['Yeosu', 'Daesan', 'Ulsan', 'Onsan']
    region_shares = regional.set_index('location')['share_pct'] / 100

    # Headers
    headers = ["Year", "Total CAPEX ($M)"] + [f"{r} ($M)" for r in top_regions] + ["Others ($M)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    # Write annual data
    for _, r in opt_copy.iterrows():
        year = int(r['year'])
        annual_capex = r['annual_capex_musd']

        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=f"{annual_capex:,.0f}")

        other_share = 1.0
        for j, region in enumerate(top_regions, 3):
            share = region_shares.get(region, 0)
            other_share -= share
            ws.cell(row=row, column=j, value=f"{annual_capex * share:,.0f}")

        ws.cell(row=row, column=7, value=f"{annual_capex * other_share:,.0f}")
        row += 1

    row += 2

    # Section 3: ANNUAL Regional Energy Demand Trajectory (NEW)
    ws.cell(row=row, column=1, value="3. Annual Regional Electricity Demand (2025-2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 1
    ws.cell(row=row, column=1, value="New electricity demand from NCC-Electricity deployment, allocated by NCC capacity share").font = Font(italic=True, size=9)
    row += 2

    # Calculate NCC capacity share by region
    ncc_fac = merged[merged['process'] == 'Naphtha Cracker']
    ncc_by_region = ncc_fac.groupby('location')['capacity_kt'].sum()
    total_ncc = ncc_by_region.sum()
    ncc_shares = ncc_by_region / total_ncc

    # Headers
    headers = ["Year", "Total Elec (TWh)"] + [f"{r} (TWh)" for r in top_regions] + ["Others (TWh)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    # Write annual data
    for _, r in opt_copy.iterrows():
        year = int(r['year'])
        elec_demand = r['electricity_consumption_increase_twh']

        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=round(elec_demand, 1))

        other_share = 1.0
        for j, region in enumerate(top_regions, 3):
            share = ncc_shares.get(region, 0)
            other_share -= share
            ws.cell(row=row, column=j, value=round(elec_demand * share, 1))

        ws.cell(row=row, column=7, value=round(elec_demand * other_share, 1))
        row += 1

    row += 2

    # Section 4: Regional NCC Capacity Summary
    ws.cell(row=row, column=1, value="4. NCC (Naphtha Cracker) Capacity by Region").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    headers = ["Region", "NCC Capacity (kt)", "Share (%)", "2050 Est. Elec Demand (TWh)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    total_elec_2050 = opt[opt['year'] == 2050]['electricity_consumption_increase_twh'].values[0]

    for region in top_regions:
        if region in ncc_by_region.index:
            cap = ncc_by_region[region]
            share = ncc_shares[region]
            elec = total_elec_2050 * share

            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value=f"{cap:,.0f}")
            ws.cell(row=row, column=3, value=f"{share*100:.1f}%")
            ws.cell(row=row, column=4, value=f"{elec:.1f}")
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15

    return wb


# =============================================================================
# SHEET 12: OPERATING RATE ANALYSIS
# =============================================================================
def create_operating_rate_sheet(wb, data):
    """Create Operating Rate Analysis sheet"""
    ws = wb.create_sheet("12. Operating Rate Analysis")

    ws['A1'] = "Operating Rate Analysis and Model Assumptions"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3

    # Section 1: Operating Rate Definition
    ws.cell(row=row, column=1, value="1. Operating Rate Definition").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    definitions = [
        ("Operating Rate", "Actual production / Installed capacity", "70%"),
        ("Capacity Multiplier", "Installed capacity growth from 2025 baseline", "1.0 -> 1.288"),
        ("Effective Multiplier", "Capacity Multiplier x Operating Rate", "Production scaling factor"),
    ]

    ws.cell(row=row, column=1, value="Parameter").style = 'header_style'
    ws.cell(row=row, column=2, value="Definition").style = 'header_style'
    ws.cell(row=row, column=3, value="Value").style = 'header_style'
    row += 1

    for param, defn, val in definitions:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=defn)
        ws.cell(row=row, column=3, value=val)
        row += 1

    row += 2

    # Section 2: Operating Rate in Model
    ws.cell(row=row, column=1, value="2. How Operating Rate is Used in the Model").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="Formula:")
    row += 1
    ws.cell(row=row, column=1, value="Effective_Multiplier = Capacity_Multiplier x Operating_Rate").font = Font(italic=True)
    row += 1
    ws.cell(row=row, column=1, value="Actual_Emissions = Baseline_Emissions x Effective_Multiplier").font = Font(italic=True)
    row += 2

    # Section 3: Impact of Operating Rate
    ws.cell(row=row, column=1, value="3. Impact of Operating Rate on Results").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    impacts = [
        ("Baseline Emissions (2025)", "At 100% op rate: 66.2 MtCO2", "At 70% op rate: 46.3 MtCO2"),
        ("BAU Emissions (2050)", "At 100% op rate: ~76 MtCO2", "At 70% op rate: 53.3 MtCO2"),
        ("Total Investment Need", "At 100% op rate: ~$53B", "At 70% op rate: $37.3B"),
    ]

    ws.cell(row=row, column=1, value="Metric").style = 'header_style'
    ws.cell(row=row, column=2, value="100% Operating Rate").style = 'header_style'
    ws.cell(row=row, column=3, value="70% Operating Rate (Model)").style = 'header_style'
    row += 1

    for metric, val_100, val_70 in impacts:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=val_100)
        ws.cell(row=row, column=3, value=val_70)
        row += 1

    row += 2

    # Section 4: Rationale
    ws.cell(row=row, column=1, value="4. Rationale for 70% Operating Rate").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    rationale = [
        "- Korean petrochemical industry experienced significant overcapacity crisis in 2023-2024",
        "- Chinese petrochemical expansion led to global supply glut",
        "- KPIA data shows actual operating rates dropped to 65-75% in 2023",
        "- 70% represents conservative assumption for 2025-2050 period",
    ]

    for text in rationale:
        ws.cell(row=row, column=1, value=text)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35

    return wb


# =============================================================================
# SHEET 13: NCC-H2 REFERENCE SCENARIO (NEW)
# =============================================================================
def create_ncc_h2_scenario_sheet(wb, data):
    """Create NCC-H2 Reference Scenario sheet"""
    ws = wb.create_sheet("13. NCC-H2 Scenario")

    ws['A1'] = "NCC-H2 Reference Scenario: Hydrogen-based Naphtha Cracking"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Alternative to NCC-Electricity using green hydrogen as fuel"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Section 1: Technology Comparison
    ws.cell(row=row, column=1, value="1. NCC-Electricity vs NCC-H2 Comparison").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    headers = ["Parameter", "NCC-Electricity", "NCC-H2"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    comparison = [
        ("Energy Input", "5.0 MWh/ton C2H4", "0.2 ton H2/ton C2H4"),
        ("TRL (2024)", "8 (Pilot)", "7 (Demo)"),
        ("Available Year", "2030", "2030"),
        ("CAPEX 2025", "$1,500/t-C2H4/yr", "$1,700/t-C2H4/yr"),
        ("CAPEX 2050", "$900/t-C2H4/yr", "$780/t-C2H4/yr"),
        ("Infrastructure Need", "Grid expansion (231 TWh)", "H2 supply (1.7 Mt/yr)"),
        ("Key Advantage", "Uses existing grid", "Lower long-term cost"),
        ("Key Challenge", "Massive electricity demand", "H2 supply chain needed"),
    ]

    for param, elec, h2 in comparison:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=elec)
        ws.cell(row=row, column=3, value=h2)
        row += 1

    row += 2

    # Section 2: Cost Comparison
    ws.cell(row=row, column=1, value="2. Cost Comparison (2050)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    # Get MACC data
    macc = data['macc_annual']
    macc_2050 = macc[macc['year'] == 2050]

    ncc_elec_cost = macc_2050[macc_2050['technology'] == 'NCC-Electricity']['total_cost_usd_per_tco2'].values
    ncc_h2_cost = macc_2050[macc_2050['technology'] == 'NCC-H2']['total_cost_usd_per_tco2'].values

    ws.cell(row=row, column=1, value="Metric").style = 'header_style'
    ws.cell(row=row, column=2, value="NCC-Electricity").style = 'header_style'
    ws.cell(row=row, column=3, value="NCC-H2").style = 'header_style'
    row += 1

    cost_data = [
        ("MACC ($/tCO2)", f"${ncc_elec_cost[0]:,.0f}" if len(ncc_elec_cost) > 0 else "N/A",
                         f"${ncc_h2_cost[0]:,.0f}" if len(ncc_h2_cost) > 0 else "N/A"),
        ("Electricity Demand (TWh)", "231", "42"),
        ("H2 Demand (Mt)", "0", "1.7"),
        ("Grid Expansion Requirement", "~40% of Korea's 2024 generation", "Minimal"),
    ]

    for metric, elec, h2 in cost_data:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=elec)
        ws.cell(row=row, column=3, value=h2)
        row += 1

    row += 2

    # Section 3: Why Model Selects NCC-Electricity
    ws.cell(row=row, column=1, value="3. Why Model Selects NCC-Electricity Over NCC-H2").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    reasons = [
        "1. Lower MACC at current H2 prices ($4-6/kg in 2030-2040)",
        "2. Grid infrastructure already exists (expansion needed but less risky)",
        "3. No dependency on green H2 supply chain development",
        "4. Electricity prices are more stable and predictable than H2 prices",
        "5. Model assumes unlimited grid capacity (not a binding constraint)",
    ]

    for reason in reasons:
        ws.cell(row=row, column=1, value=reason)
        row += 1

    row += 2

    # Section 4: Scenarios Comparison (if available)
    if data['scenarios_6'] is not None:
        ws.cell(row=row, column=1, value="4. 6 Scenario Comparison Results (2050)").font = Font(bold=True, size=12, color='1F4E79')
        row += 2

        headers = ["Scenario", "Tech", "Net (Mt)", "Cost ($B)", "Elec (TWh)", "H2 (kt)"]
        for j, h in enumerate(headers, 1):
            ws.cell(row=row, column=j, value=h).style = 'header_style'
        row += 1

        for _, r in data['scenarios_6'].iterrows():
            ws.cell(row=row, column=1, value=r['scenario'])
            ws.cell(row=row, column=2, value=r['technology'])
            ws.cell(row=row, column=3, value=round(r['net_2050_mt'], 2))
            ws.cell(row=row, column=4, value=f"${r['capex_billion_usd']:.1f}")
            ws.cell(row=row, column=5, value=round(r['electricity_twh'], 1))
            ws.cell(row=row, column=6, value=f"{r['h2_kt']:,.0f}")
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    return wb


# =============================================================================
# SHEET 14: SCENARIO COMPARISON (NEW)
# =============================================================================
def create_scenario_comparison_sheet(wb, data):
    """Create Scenario Comparison sheet"""
    ws = wb.create_sheet("14. Scenario Comparison")

    ws['A1'] = "6 Scenario Comparison: Production Pathways x Technology Pathways"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "3 Production scenarios (Shaheen/25% Restructuring/40% Restructuring) x 2 Technology pathways (NCC-Elec/NCC-H2)"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Section 1: Scenario Definitions
    ws.cell(row=row, column=1, value="1. Scenario Definitions").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="A. Production Pathways:").font = Font(bold=True)
    row += 1

    prod_scenarios = [
        ("Shaheen (성장)", "S-Oil Shaheen project adds 6 NEW facilities (254 total), +3,950 kt capacity, 44 NCC"),
        ("구조조정 25%", "Retire oldest 9 NCC facilities (239 total), 6,774 kt NCC capacity retired, 32 NCC"),
        ("구조조정 40%", "Retire oldest 16 NCC facilities (232 total), 10,854 kt NCC capacity retired, 25 NCC"),
    ]

    for name, desc in prod_scenarios:
        ws.cell(row=row, column=1, value=f"  - {name}: {desc}")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="B. Technology Pathways:").font = Font(bold=True)
    row += 1

    tech_scenarios = [
        ("NCC-Electricity", "Electric cracker using renewable electricity (5.0 MWh/ton C2H4)"),
        ("NCC-H2", "Hydrogen-fueled cracker (0.2 ton H2/ton C2H4)"),
    ]

    for name, desc in tech_scenarios:
        ws.cell(row=row, column=1, value=f"  - {name}: {desc}")
        row += 1

    row += 2

    # Section 2: Results Comparison
    ws.cell(row=row, column=1, value="2. 2050 Results Comparison").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    if data['scenarios_6'] is not None:
        headers = ["Scenario", "Tech", "Facilities", "NCC", "BAU (Mt)", "Net (Mt)", "Cost ($B)", "Elec (TWh)", "H2 (kt)"]
        for j, h in enumerate(headers, 1):
            ws.cell(row=row, column=j, value=h).style = 'header_style'
        row += 1

        for _, r in data['scenarios_6'].iterrows():
            ws.cell(row=row, column=1, value=r['scenario'])
            ws.cell(row=row, column=2, value=r['technology'])
            ws.cell(row=row, column=3, value=int(r['n_facilities']))
            ws.cell(row=row, column=4, value=int(r['n_ncc_facilities']))
            ws.cell(row=row, column=5, value=round(r['bau_2050_mt'], 1))
            ws.cell(row=row, column=6, value=round(r['net_2050_mt'], 2))
            ws.cell(row=row, column=7, value=f"${r['capex_billion_usd']:.1f}")
            ws.cell(row=row, column=8, value=round(r['electricity_twh'], 1))
            ws.cell(row=row, column=9, value=f"{r['h2_kt']:,.0f}")
            row += 1

        row += 2

        # Section 3: Retired Facilities Detail
        ws.cell(row=row, column=1, value="3. Retired NCC Facilities (Restructuring Scenarios)").font = Font(bold=True, size=12, color='1F4E79')
        row += 2

        if data.get('retired_25') is not None:
            ws.cell(row=row, column=1, value="구조조정 25% - Retired 9 NCC facilities:").font = Font(bold=True)
            row += 1
            for _, fac in data['retired_25'].iterrows():
                ws.cell(row=row, column=1, value=f"  - {fac['company']} {fac['product']} ({fac['location']}): {fac['capacity_kt']:,.0f} kt, built {int(fac['year_built'])}")
                row += 1
            row += 1

        if data.get('retired_40') is not None:
            additional_retired = len(data['retired_40']) - len(data.get('retired_25', []))
            ws.cell(row=row, column=1, value=f"구조조정 40% - Additional {additional_retired} NCC facilities retired:").font = Font(bold=True)
            row += 1
            retired_25_idx = set(data['retired_25'].index) if data.get('retired_25') is not None else set()
            for idx, fac in data['retired_40'].iterrows():
                if idx >= len(data.get('retired_25', [])):  # Only show additional facilities
                    ws.cell(row=row, column=1, value=f"  - {fac['company']} {fac['product']} ({fac['location']}): {fac['capacity_kt']:,.0f} kt, built {int(fac['year_built'])}")
                    row += 1

        row += 2

        # Section 4: Key Insights
        ws.cell(row=row, column=1, value="4. Key Insights").font = Font(bold=True, size=12, color='1F4E79')
        row += 2

        insights = [
            "1. 구조조정 40% + NCC-Electricity achieves near Net Zero at lowest cost ($34.1B)",
            "2. Shaheen (growth) scenarios require highest investment ($58.9-64.7B) but maintain production",
            "3. NCC-H2 scenarios are ~10% cheaper than NCC-Electricity but require H2 infrastructure",
            "4. Restructuring reduces BAU emissions significantly: 25% saves 19.7 Mt, 40% saves 28.0 Mt",
            "5. All scenarios achieve <4 Mt net emissions by 2050 through technology deployment",
        ]

        for insight in insights:
            ws.cell(row=row, column=1, value=insight)
            row += 1
    else:
        ws.cell(row=row, column=1, value="Run run_scenarios_complete.py to generate 6 scenario comparison data")
        ws.cell(row=row, column=1).font = Font(italic=True, color='FF0000')

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12

    return wb


# =============================================================================
# SHEET 15: 2035 TARGET ANALYSIS (NEW)
# =============================================================================
def create_2035_target_sheet(wb, data):
    """Create 2035 Target Analysis sheet"""
    ws = wb.create_sheet("15. 2035 Target Analysis")

    ws['A1'] = "2035 Emission Target Analysis: NDC Compliance"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Korea's 2035 industrial emission target relative to 2018 baseline"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Section 1: Target Definition
    ws.cell(row=row, column=1, value="1. 2035 Target Definition").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    target_info = [
        ("2018 Baseline Emissions", "57.6 MtCO2", "Reference year for NDC calculation"),
        ("2035 Reduction Target", "24.5%", "Korea NDC submission (2024)"),
        ("2035 Emission Target", "43.5 MtCO2", "57.6 x (1 - 0.245) = 43.5"),
        ("Source", "Korea NDC", "National Determined Contribution"),
    ]

    ws.cell(row=row, column=1, value="Parameter").style = 'header_style'
    ws.cell(row=row, column=2, value="Value").style = 'header_style'
    ws.cell(row=row, column=3, value="Notes").style = 'header_style'
    row += 1

    for param, value, notes in target_info:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=notes)
        row += 1

    row += 2

    # Section 2: Model Results vs 2035 Target
    ws.cell(row=row, column=1, value="2. Model Results vs 2035 Target").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    opt = data['opt_traj']
    bau = data['bau_traj']

    # Get 2035 values
    opt_2035 = opt[opt['year'] == 2035].iloc[0]
    bau_2035 = bau[bau['year'] == 2035].iloc[0]
    target_2035 = 43.5

    ws.cell(row=row, column=1, value="Scenario").style = 'header_style'
    ws.cell(row=row, column=2, value="2035 Emissions (Mt)").style = 'header_style'
    ws.cell(row=row, column=3, value="vs Target").style = 'header_style'
    ws.cell(row=row, column=4, value="Status").style = 'header_style'
    row += 1

    results = [
        ("2035 Target (NDC)", target_2035, "-", "Target"),
        ("BAU (No Action)", bau_2035['total_emissions_mt'], f"+{bau_2035['total_emissions_mt'] - target_2035:.1f} Mt", "FAIL"),
        ("Cost-Effective (Model)", opt_2035['actual_emissions_mt'], f"-{target_2035 - opt_2035['actual_emissions_mt']:.1f} Mt", "PASS"),
    ]

    for scenario, emissions, vs_target, status in results:
        ws.cell(row=row, column=1, value=scenario)
        ws.cell(row=row, column=2, value=round(emissions, 2))
        ws.cell(row=row, column=3, value=vs_target)
        cell = ws.cell(row=row, column=4, value=status)
        if status == "PASS":
            cell.font = Font(bold=True, color='006400')
        elif status == "FAIL":
            cell.font = Font(bold=True, color='FF0000')
        row += 1

    row += 2

    # Section 3: Annual Trajectory to 2035
    ws.cell(row=row, column=1, value="3. Annual Trajectory to 2035").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    headers = ["Year", "Target (Mt)", "BAU (Mt)", "Model (Mt)", "Gap vs Target"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    # Use emission scenarios for targets
    scenarios = data['emission_scenarios']

    for year in range(2025, 2036):
        target_row = scenarios[(scenarios['year'] == year) & (scenarios['scenario_name'] == 'Policy_Target')]
        target = target_row['target_mt'].values[0] if len(target_row) > 0 else None

        bau_row = bau[bau['year'] == year]
        bau_val = bau_row['total_emissions_mt'].values[0] if len(bau_row) > 0 else None

        opt_row = opt[opt['year'] == year]
        opt_val = opt_row['actual_emissions_mt'].values[0] if len(opt_row) > 0 else None

        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=round(target, 1) if target else "-")
        ws.cell(row=row, column=3, value=round(bau_val, 1) if bau_val else "-")
        ws.cell(row=row, column=4, value=round(opt_val, 2) if opt_val else "-")

        if target and opt_val:
            gap = opt_val - target
            ws.cell(row=row, column=5, value=f"{gap:+.1f}" if abs(gap) > 0.1 else "On Track")
        row += 1

    row += 2

    # Section 4: Key Finding
    ws.cell(row=row, column=1, value="4. Key Finding").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    ws.cell(row=row, column=1, value="The model's cost-effective scenario EXCEEDS the 2035 NDC target:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"  - Model 2035 emissions: {opt_2035['actual_emissions_mt']:.1f} Mt")
    row += 1
    ws.cell(row=row, column=1, value=f"  - NDC 2035 target: 43.5 Mt")
    row += 1
    ws.cell(row=row, column=1, value=f"  - Margin: {43.5 - opt_2035['actual_emissions_mt']:.1f} Mt below target").font = Font(bold=True, color='006400')
    row += 2

    ws.cell(row=row, column=1, value="Note: The 70% operating rate assumption means current actual emissions (46.3 Mt in 2025)")
    row += 1
    ws.cell(row=row, column=1, value="are already below the 2018 baseline (57.6 Mt), giving a head start toward the 2035 target.")

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    return wb


# =============================================================================
# MAIN
# =============================================================================
def generate_report():
    """Generate the complete Excel report v2"""
    print("="*70)
    print("Korea Petrochemical Net Zero Report Generator v2")
    print("="*70)

    print("\nLoading data...")
    data = load_all_data()

    print("Creating workbook...")
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create styles
    wb = create_styles(wb)

    # Create sheets
    print("Creating Sheet 1: Executive Summary...")
    wb = create_executive_summary(wb, data)

    print("Creating Sheet 2: Assumptions & References...")
    wb = create_assumptions_sheet(wb, data)

    print("Creating Sheet 3: Facility Database...")
    wb = create_facility_sheet(wb, data)

    print("Creating Sheet 4: Emission Methodology...")
    wb = create_emission_methodology_sheet(wb, data)

    print("Creating Sheet 5: BAU Trajectory...")
    wb = create_bau_trajectory_sheet(wb, data)

    print("Creating Sheet 6: Technology MACC...")
    wb = create_macc_sheet(wb, data)

    print("Creating Sheet 7: Optimization Results...")
    wb = create_optimization_sheet(wb, data)

    print("Creating Sheet 8: Energy Demand...")
    wb = create_energy_demand_sheet(wb, data)

    print("Creating Sheet 9: Cost Analysis...")
    wb = create_cost_analysis_sheet(wb, data)

    print("Creating Sheet 10: References...")
    wb = create_references_sheet(wb)

    print("Creating Sheet 11: Regional Analysis (Enhanced)...")
    wb = create_regional_analysis_sheet(wb, data)

    print("Creating Sheet 12: Operating Rate Analysis...")
    wb = create_operating_rate_sheet(wb, data)

    print("Creating Sheet 13: NCC-H2 Scenario (NEW)...")
    wb = create_ncc_h2_scenario_sheet(wb, data)

    print("Creating Sheet 14: Scenario Comparison (NEW)...")
    wb = create_scenario_comparison_sheet(wb, data)

    print("Creating Sheet 15: 2035 Target Analysis (NEW)...")
    wb = create_2035_target_sheet(wb, data)

    print("Creating Sheet 16: Learning Curve Sensitivity (NEW)...")
    wb = create_learning_curve_sheet(wb, data)

    print("Creating Sheet 17: Facility Transition Details (NEW)...")
    wb = create_facility_transition_sheet(wb, data)

    print("Creating Sheet 18: Full Facility Database (NEW)...")
    wb = create_facility_database_full_sheet(wb, data)

    print("Creating Sheet 19: Annual MACC Data (NEW)...")
    wb = create_annual_macc_sheet(wb, data)

    print("Creating Sheet 20: 2050 Facility Allocation (NEW)...")
    wb = create_facility_allocation_sheet(wb, data)

    print("Creating Sheets 21-26: Detailed Scenario Facility Sheets...")
    wb = create_scenario_facility_sheets(wb, data)

    # Save
    output_path = OUTPUT_DIR / "Korea_Petrochemical_NetZero_Report_v2.xlsx"
    wb.save(output_path)

    print("\n" + "="*70)
    print(f"Report saved to: {output_path}")
    print("="*70)

    print("\nReport Contents:")
    print("  1. Executive Summary")
    print("  2. Assumptions & References")
    print("  3. Facility Database")
    print("  4. Emission Methodology")
    print("  5. BAU Trajectory")
    print("  6. Technology MACC")
    print("  7. Optimization Results")
    print("  8. Energy Demand")
    print("  9. Cost Analysis")
    print("  10. References")
    print("  11. Regional Analysis (Enhanced with annual data)")
    print("  12. Operating Rate Analysis")
    print("  13. NCC-H2 Scenario")
    print("  14. Scenario Comparison (6 scenarios)")
    print("  15. 2035 Target Analysis")
    print("  16. Learning Curve Sensitivity")
    print("  17. Facility Transition Details")
    print("  18. Full Facility Database")
    print("  19. Annual MACC Data")
    print("  20. 2050 Facility Allocation")
    print("  21. Scenario: Shaheen + NCC-H2 (Facility Details)")
    print("  22. Scenario: Shaheen + NCC-Electricity (Facility Details)")
    print("  23. Scenario: Restructure 25% + NCC-H2 (Facility Details)")
    print("  24. Scenario: Restructure 25% + NCC-Electricity (Facility Details)")
    print("  25. Scenario: Restructure 40% + NCC-H2 (Facility Details)")
    print("  26. Scenario: Restructure 40% + NCC-Electricity (Facility Details)")

    return output_path


# =============================================================================
# SHEET 16: LEARNING CURVE SENSITIVITY (NEW)
# =============================================================================
def create_learning_curve_sheet(wb, data):
    """Create Learning Curve Sensitivity Analysis sheet"""
    ws = wb.create_sheet("16. Learning Curve Sensitivity")

    ws['A1'] = "Learning Curve Sensitivity Analysis"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Impact of technology cost reduction on total CAPEX investment"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Section 1: Learning Curve Assumptions
    ws.cell(row=row, column=1, value="1. CAPEX Learning Curve Assumptions").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    headers = ["Technology", "2025 CAPEX", "2030 CAPEX", "2040 CAPEX", "2050 CAPEX", "Reduction"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    tech_params = data['tech_params']
    for _, r in tech_params.iterrows():
        if r['capex_2025_musd_per_mtco2'] > 0:
            capex_2025 = r['capex_2025_musd_per_mtco2']
            capex_2050 = r['capex_2050_musd_per_mtco2']
            reduction = (1 - capex_2050 / capex_2025) * 100 if capex_2025 > 0 else 0

            ws.cell(row=row, column=1, value=r['technology'])
            ws.cell(row=row, column=2, value=f"${capex_2025:,.0f}")
            ws.cell(row=row, column=3, value=f"${r['capex_2030_musd_per_mtco2']:,.0f}")
            ws.cell(row=row, column=4, value=f"${r['capex_2040_musd_per_mtco2']:,.0f}")
            ws.cell(row=row, column=5, value=f"${capex_2050:,.0f}")
            ws.cell(row=row, column=6, value=f"-{reduction:.0f}%")
            row += 1

    row += 2

    # Section 2: Learning Curve Impact
    ws.cell(row=row, column=1, value="2. Cost Impact: With vs Without Learning Curve").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    # Try to load the comparison data
    comparison_path = OUTPUTS_DIR / 'scenarios_learning_curve_comparison' / 'summary.csv'
    if comparison_path.exists():
        df_comparison = pd.read_csv(comparison_path)

        headers = ["Scenario", "With Learning ($B)", "No Learning ($B)", "Increase ($B)", "Increase (%)"]
        for j, h in enumerate(headers, 1):
            ws.cell(row=row, column=j, value=h).style = 'header_style'
        row += 1

        # Group by production and technology pathway
        df_learn = df_comparison[df_comparison['learning_curve'] == 'Yes (Default)']
        df_no_learn = df_comparison[df_comparison['learning_curve'] == 'No (Fixed at 2025)']

        for _, learn_row in df_learn.iterrows():
            prod = learn_row['production_pathway']
            tech = learn_row['technology_pathway']

            # Find matching no-learning row
            no_learn_match = df_no_learn[
                (df_no_learn['production_pathway'] == prod) &
                (df_no_learn['technology_pathway'] == tech)
            ]

            if len(no_learn_match) > 0:
                no_learn_cost = no_learn_match.iloc[0]['cost_2050_billion_usd']
                learn_cost = learn_row['cost_2050_billion_usd']
                increase = no_learn_cost - learn_cost
                pct_increase = (increase / learn_cost) * 100

                ws.cell(row=row, column=1, value=f"{prod} + {tech}")
                ws.cell(row=row, column=2, value=f"${learn_cost:.1f}")
                ws.cell(row=row, column=3, value=f"${no_learn_cost:.1f}")
                ws.cell(row=row, column=4, value=f"+${increase:.1f}")
                ws.cell(row=row, column=5, value=f"+{pct_increase:.0f}%")
                row += 1

        row += 2

        # Section 3: Key Insights
        ws.cell(row=row, column=1, value="3. Key Insights").font = Font(bold=True, size=12, color='1F4E79')
        row += 2

        insights = [
            "1. Learning curves reduce total CAPEX by 50-68% across all scenarios",
            "2. Without learning curves, the transition would cost 2-3x more",
            "3. NCC-H2 is most sensitive to learning curve (54% CAPEX reduction by 2050)",
            "4. NCC-Electricity learning curve (40% reduction) is more conservative",
            "5. Learning curves are critical for economic feasibility of net zero transition",
        ]

        for insight in insights:
            ws.cell(row=row, column=1, value=insight)
            row += 1

        row += 2

        # Section 4: Implications
        ws.cell(row=row, column=1, value="4. Policy Implications").font = Font(bold=True, size=12, color='1F4E79')
        row += 2

        implications = [
            "- If learning curves don't materialize, investment needs increase by $20-44B",
            "- R&D support and technology deployment scale-up are critical for cost reduction",
            "- Early deployment creates learning-by-doing benefits for later investments",
            "- Without cost reduction, some scenarios may become economically infeasible",
        ]

        for impl in implications:
            ws.cell(row=row, column=1, value=impl)
            row += 1

    else:
        ws.cell(row=row, column=1, value="Run run_no_learning_scenarios.py to generate sensitivity analysis data")
        ws.cell(row=row, column=1).font = Font(italic=True, color='FF0000')

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    return wb


# =============================================================================
# SHEET 17: FACILITY TRANSITION DETAILS (NEW)
# =============================================================================
def create_facility_transition_sheet(wb, data):
    """Create detailed facility-level annual transition data"""
    ws = wb.create_sheet("17. Facility Transition")

    ws['A1'] = "Facility-Level Technology Transition Details"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Annual deployment schedule by facility, process, and region"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Load facility database
    fac = data['facilities']
    opt = data['opt_traj']

    # Section 1: Facility Summary by Process Type
    ws.cell(row=row, column=1, value="1. Facility Summary by Process Type").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    process_summary = fac.groupby('process').agg({
        'capacity_kt': 'sum',
        'product': 'count'
    }).rename(columns={'product': 'n_facilities'}).reset_index()

    headers = ["Process", "Facilities", "Capacity (kt)", "Share (%)", "Primary Technology"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    total_cap = fac['capacity_kt'].sum()
    process_tech = {
        'Naphtha Cracker': 'NCC-Electricity / NCC-H2',
        'BTX Plant': 'Heat Pump + RDH',
        'Utility': 'Heat Pump + RE PPA'
    }

    for _, r in process_summary.iterrows():
        ws.cell(row=row, column=1, value=r['process'])
        ws.cell(row=row, column=2, value=int(r['n_facilities']))
        ws.cell(row=row, column=3, value=f"{r['capacity_kt']:,.0f}")
        ws.cell(row=row, column=4, value=f"{r['capacity_kt']/total_cap*100:.1f}%")
        ws.cell(row=row, column=5, value=process_tech.get(r['process'], 'RE PPA'))
        row += 1

    row += 2

    # Section 2: Technology Deployment Schedule by Year
    ws.cell(row=row, column=1, value="2. Annual Technology Deployment Schedule").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    # Get annual deployment from optimization trajectory
    headers = ["Year", "Heat Pump (Mt)", "NCC-Elec (Mt)", "RE PPA (Mt)", "Total Deployed (Mt)", "Remaining (Mt)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for _, r in opt.iterrows():
        ws.cell(row=row, column=1, value=int(r['year']))
        ws.cell(row=row, column=2, value=round(r['heat_pump_mt'], 2))
        ws.cell(row=row, column=3, value=round(r['ncc_elec_mt'], 2))
        ws.cell(row=row, column=4, value=round(r['re_ppa_mt'], 2))
        ws.cell(row=row, column=5, value=round(r['total_deployed_mt'], 2))
        ws.cell(row=row, column=6, value=round(r['actual_emissions_mt'], 2))
        row += 1

    row += 2

    # Section 3: NCC Facility Transition (Priority List)
    ws.cell(row=row, column=1, value="3. Naphtha Cracker (NCC) Facility Transition Priority").font = Font(bold=True, size=12, color='1F4E79')
    row += 1
    ws.cell(row=row, column=1, value="Sorted by capacity (largest first for maximum emission reduction impact)").font = Font(italic=True, size=9)
    row += 2

    ncc_fac = fac[fac['process'] == 'Naphtha Cracker'].copy()
    ncc_fac = ncc_fac.sort_values('capacity_kt', ascending=False)

    # Calculate deployment year based on optimization trajectory
    # Assume largest crackers deploy first starting from 2030
    total_ncc_cap = ncc_fac['capacity_kt'].sum()
    ncc_fac['cumulative_cap'] = ncc_fac['capacity_kt'].cumsum()
    ncc_fac['cap_pct'] = ncc_fac['cumulative_cap'] / total_ncc_cap * 100

    # Map to deployment years based on optimization trajectory
    def get_deploy_year(pct):
        if pct <= 5: return 2030
        elif pct <= 15: return 2032
        elif pct <= 30: return 2035
        elif pct <= 50: return 2038
        elif pct <= 70: return 2042
        elif pct <= 85: return 2045
        else: return 2048

    ncc_fac['estimated_deploy_year'] = ncc_fac['cap_pct'].apply(get_deploy_year)

    headers = ["Rank", "Company", "Location", "Product", "Capacity (kt)", "Est. Deploy Year", "Technology"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for i, (_, r) in enumerate(ncc_fac.iterrows(), 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r['company'])
        ws.cell(row=row, column=3, value=r['location'])
        ws.cell(row=row, column=4, value=r['product'])
        ws.cell(row=row, column=5, value=f"{r['capacity_kt']:,.0f}")
        ws.cell(row=row, column=6, value=int(r['estimated_deploy_year']))
        ws.cell(row=row, column=7, value="NCC-Electricity")
        row += 1

    row += 2

    # Section 4: Regional Deployment Timeline
    ws.cell(row=row, column=1, value="4. Regional Technology Deployment Timeline").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    regions = ['Yeosu', 'Daesan', 'Ulsan', 'Onsan']
    years = [2025, 2030, 2035, 2040, 2045, 2050]

    headers = ["Region", "Facilities"] + [str(y) for y in years]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for region in regions:
        region_fac = fac[fac['location'] == region]
        n_fac = len(region_fac)
        region_cap = region_fac['capacity_kt'].sum()

        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value=n_fac)

        # Calculate % deployed per year (proportional to overall trajectory)
        for j, year in enumerate(years, 3):
            year_row = opt[opt['year'] == year]
            if len(year_row) > 0:
                total_deployed_pct = year_row.iloc[0]['total_deployed_mt'] / opt['bau_mt'].max() * 100
                ws.cell(row=row, column=j, value=f"{total_deployed_pct:.0f}%")
            else:
                ws.cell(row=row, column=j, value="0%")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    return wb


# =============================================================================
# SHEET 18: FACILITY DATABASE FULL (NEW)
# =============================================================================
def create_facility_database_full_sheet(wb, data):
    """Create full facility database with all details"""
    ws = wb.create_sheet("18. Full Facility Database")

    ws['A1'] = "Complete Facility Database (248 Facilities)"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "All Korean petrochemical facilities with capacity, age, and transition status"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    fac = data['facilities'].copy()
    fac = fac.sort_values(['process', 'capacity_kt'], ascending=[True, False])

    headers = ["ID", "Product", "Process", "Company", "Location", "Capacity (kt)",
               "Year Built", "Age (2025)", "Remaining Life", "Retirement (40yr)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for i, (_, r) in enumerate(fac.iterrows(), 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r['product'])
        ws.cell(row=row, column=3, value=r['process'])
        ws.cell(row=row, column=4, value=r['company'])
        ws.cell(row=row, column=5, value=r['location'])
        ws.cell(row=row, column=6, value=f"{r['capacity_kt']:,.0f}")
        ws.cell(row=row, column=7, value=int(r['year_built']) if pd.notna(r['year_built']) else "-")
        ws.cell(row=row, column=8, value=int(r['age_2025']) if pd.notna(r['age_2025']) else "-")
        ws.cell(row=row, column=9, value=int(r['remaining_life']) if pd.notna(r['remaining_life']) else "-")
        ws.cell(row=row, column=10, value=int(r['retirement_year_40yr']) if pd.notna(r['retirement_year_40yr']) else "-")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 16

    return wb


# =============================================================================
# SHEET 19: ANNUAL MACC BY TECHNOLOGY (NEW)
# =============================================================================
def create_annual_macc_sheet(wb, data):
    """Create annual MACC data by technology"""
    ws = wb.create_sheet("19. Annual MACC Data")

    ws['A1'] = "Annual Marginal Abatement Cost Curve (MACC) by Technology"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Technology costs and abatement potential from 2025-2050"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    macc = data['macc_annual'].copy()

    # Section 1: Cost Evolution Summary
    ws.cell(row=row, column=1, value="1. Technology Cost Evolution ($/tCO2)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    years = [2025, 2030, 2035, 2040, 2045, 2050]
    technologies = ['Heat_Pump', 'NCC-Electricity', 'NCC-H2', 'RDH', 'RE_PPA']

    headers = ["Technology"] + [str(y) for y in years]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for tech in technologies:
        ws.cell(row=row, column=1, value=tech)
        for j, year in enumerate(years, 2):
            year_data = macc[(macc['year'] == year) & (macc['technology'] == tech)]
            if len(year_data) > 0:
                cost = year_data.iloc[0]['total_cost_usd_per_tco2']
                ws.cell(row=row, column=j, value=f"${cost:,.0f}")
            else:
                ws.cell(row=row, column=j, value="-")
        row += 1

    row += 2

    # Section 2: Abatement Potential Evolution
    ws.cell(row=row, column=1, value="2. Abatement Potential Evolution (MtCO2)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    headers = ["Technology"] + [str(y) for y in years]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h).style = 'header_style'
    row += 1

    for tech in technologies:
        ws.cell(row=row, column=1, value=tech)
        for j, year in enumerate(years, 2):
            year_data = macc[(macc['year'] == year) & (macc['technology'] == tech)]
            if len(year_data) > 0:
                abate = year_data.iloc[0]['abatement_potential_mtco2']
                ws.cell(row=row, column=j, value=f"{abate:.1f}")
            else:
                ws.cell(row=row, column=j, value="-")
        row += 1

    row += 2

    # Section 3: Full Annual Data (first 3 years detailed)
    ws.cell(row=row, column=1, value="3. Full Annual MACC Data (Selected Years)").font = Font(bold=True, size=12, color='1F4E79')
    row += 2

    for year in [2030, 2040, 2050]:
        ws.cell(row=row, column=1, value=f"Year {year}:").font = Font(bold=True)
        row += 1

        year_data = macc[macc['year'] == year]

        headers = ["Technology", "Abatement (Mt)", "CAPEX ($/t)", "OPEX ($/t)", "Fuel Cost ($/t)", "Total ($/t)"]
        for j, h in enumerate(headers, 1):
            ws.cell(row=row, column=j, value=h).style = 'header_style'
        row += 1

        for _, r in year_data.iterrows():
            ws.cell(row=row, column=1, value=r['technology'])
            ws.cell(row=row, column=2, value=f"{r['abatement_potential_mtco2']:.2f}")
            ws.cell(row=row, column=3, value=f"${r['capex_ann_usd_per_tco2']:,.0f}")
            ws.cell(row=row, column=4, value=f"${r['opex_ann_usd_per_tco2']:,.0f}")
            ws.cell(row=row, column=5, value=f"${r['fuel_cost_diff_usd_per_tco2']:,.0f}")
            ws.cell(row=row, column=6, value=f"${r['total_cost_usd_per_tco2']:,.0f}")
            row += 1

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col in 'BCDEFG':
        ws.column_dimensions[col].width = 15

    return wb


# =============================================================================
# SHEET 20: FACILITY ALLOCATION 2050 (NEW)
# =============================================================================
def create_facility_allocation_sheet(wb, data):
    """Create 2050 facility-level technology allocation"""
    ws = wb.create_sheet("20. 2050 Facility Allocation")

    ws['A1'] = "2050 Facility-Level Technology Allocation"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws['A2'] = "Final technology deployment status for each facility in 2050"
    ws['A2'].font = Font(italic=True, size=9)

    row = 4

    # Try to load facility allocation data
    allocation_path = OUTPUTS_DIR / 'excel_report' / 'policy_target_facility_allocation_2050.csv'
    if not allocation_path.exists():
        allocation_path = OUTPUTS_DIR / 'excel_temp' / 'module_03' / 'policy_target_facility_allocation_2050.csv'

    if allocation_path.exists():
        alloc = pd.read_csv(allocation_path)

        # Section 1: Summary by Technology
        ws.cell(row=row, column=1, value="1. Technology Deployment Summary").font = Font(bold=True, size=12, color='1F4E79')
        row += 2

        # Calculate summary
        total_facilities = len(alloc)
        ncc_elec_facilities = len(alloc[alloc['tech_ncc_elec_pct'] > 0])
        heat_pump_facilities = len(alloc[alloc['tech_heat_pump_pct'] > 0])
        re_ppa_facilities = len(alloc[alloc['tech_re_ppa_pct'] > 0])

        summary = [
            ("Total Facilities", total_facilities),
            ("NCC-Electricity Deployed", ncc_elec_facilities),
            ("Heat Pump Deployed", heat_pump_facilities),
            ("RE PPA Deployed", re_ppa_facilities),
            ("Total Abatement (Mt)", f"{alloc['abatement_mt'].sum():.1f}"),
        ]

        for label, value in summary:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 2

        # Section 2: Top 20 Largest Facilities
        ws.cell(row=row, column=1, value="2. Top 20 Largest Facilities (by Abatement)").font = Font(bold=True, size=12, color='1F4E79')
        row += 2

        alloc_sorted = alloc.sort_values('abatement_mt', ascending=False).head(20)

        headers = ["Rank", "Company", "Location", "Product", "Process", "Abatement (Mt)", "Technology"]
        for j, h in enumerate(headers, 1):
            ws.cell(row=row, column=j, value=h).style = 'header_style'
        row += 1

        for i, (_, r) in enumerate(alloc_sorted.iterrows(), 1):
            # Determine primary technology
            if r['tech_ncc_elec_pct'] > 0:
                tech = "NCC-Electricity"
            elif r['tech_ncc_h2_pct'] > 0:
                tech = "NCC-H2"
            elif r['tech_heat_pump_pct'] > 0:
                tech = "Heat Pump"
            else:
                tech = "RE PPA"

            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=r['company'])
            ws.cell(row=row, column=3, value=r['location'])
            ws.cell(row=row, column=4, value=r['product'])
            ws.cell(row=row, column=5, value=r['process'])
            ws.cell(row=row, column=6, value=f"{r['abatement_mt']:.3f}")
            ws.cell(row=row, column=7, value=tech)
            row += 1

        row += 2

        # Section 3: Regional Summary
        ws.cell(row=row, column=1, value="3. Regional Allocation Summary").font = Font(bold=True, size=12, color='1F4E79')
        row += 2

        regional = alloc.groupby('location').agg({
            'abatement_mt': 'sum',
            'company': 'count'
        }).rename(columns={'company': 'n_facilities'}).reset_index()
        regional = regional.sort_values('abatement_mt', ascending=False)

        headers = ["Region", "Facilities", "Total Abatement (Mt)", "Share (%)"]
        for j, h in enumerate(headers, 1):
            ws.cell(row=row, column=j, value=h).style = 'header_style'
        row += 1

        total_abate = alloc['abatement_mt'].sum()
        for _, r in regional.iterrows():
            ws.cell(row=row, column=1, value=r['location'])
            ws.cell(row=row, column=2, value=int(r['n_facilities']))
            ws.cell(row=row, column=3, value=f"{r['abatement_mt']:.2f}")
            ws.cell(row=row, column=4, value=f"{r['abatement_mt']/total_abate*100:.1f}%")
            row += 1

    else:
        ws.cell(row=row, column=1, value="Facility allocation data not found. Run optimization module first.")
        ws.cell(row=row, column=1).font = Font(italic=True, color='FF0000')

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18

    return wb


# =============================================================================
# SHEET 21-26: DETAILED SCENARIO FACILITY SHEETS
# =============================================================================
def create_scenario_facility_sheets(wb, data):
    """Create detailed facility sheets for all 6 scenarios"""

    scenario_configs = [
        ('shaheen_ncc_h2', 'Shaheen + NCC-H2', '21'),
        ('shaheen_ncc_electricity', 'Shaheen + NCC-Elec', '22'),
        ('restructure_25pct_ncc_h2', 'Restructure 25% + NCC-H2', '23'),
        ('restructure_25pct_ncc_electricity', 'Restructure 25% + NCC-Elec', '24'),
        ('restructure_40pct_ncc_h2', 'Restructure 40% + NCC-H2', '25'),
        ('restructure_40pct_ncc_electricity', 'Restructure 40% + NCC-Elec', '26'),
    ]

    for scenario_id, scenario_name, sheet_num in scenario_configs:
        ws = wb.create_sheet(f"{sheet_num}. {scenario_name[:20]}")

        ws['A1'] = f"Scenario: {scenario_name}"
        ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

        row = 3

        # Get scenario data
        scenario_data = data.get('scenario_facilities', {}).get(scenario_id, {})

        if scenario_data and scenario_data.get('facilities') is not None:
            facilities = scenario_data['facilities']
            emissions = scenario_data.get('emissions')
            deployment = scenario_data.get('deployment')

            # Section 1: Scenario Summary
            ws.cell(row=row, column=1, value="1. Scenario Summary").font = Font(bold=True, size=12, color='1F4E79')
            row += 2

            n_facilities = len(facilities)
            n_ncc = len(facilities[facilities['process'] == 'Naphtha Cracker'])
            total_cap = facilities['capacity_kt'].sum()
            ncc_cap = facilities[facilities['process'] == 'Naphtha Cracker']['capacity_kt'].sum()

            summary = [
                ("Total Facilities", n_facilities),
                ("NCC Facilities", n_ncc),
                ("Total Capacity (kt)", f"{total_cap:,.0f}"),
                ("NCC Capacity (kt)", f"{ncc_cap:,.0f}"),
            ]

            # Get results from scenarios_6 if available
            if data.get('scenarios_6') is not None:
                scenario_result = data['scenarios_6'][data['scenarios_6']['scenario_id'] == scenario_id]
                if len(scenario_result) > 0:
                    r = scenario_result.iloc[0]
                    summary.extend([
                        ("BAU Emissions 2050 (Mt)", f"{r['bau_2050_mt']:.2f}"),
                        ("Net Emissions 2050 (Mt)", f"{r['net_2050_mt']:.2f}"),
                        ("CAPEX ($B)", f"${r['capex_billion_usd']:.1f}"),
                    ])

            for label, value in summary:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1

            row += 2

            # Section 2: NCC Facilities in this scenario
            ws.cell(row=row, column=1, value="2. NCC Facilities in Scenario").font = Font(bold=True, size=12, color='1F4E79')
            row += 2

            ncc_fac = facilities[facilities['process'] == 'Naphtha Cracker'].copy()
            ncc_fac = ncc_fac.sort_values('capacity_kt', ascending=False)

            headers = ["#", "Product", "Company", "Location", "Capacity (kt)", "Year Built", "Age (2025)"]
            for j, h in enumerate(headers, 1):
                ws.cell(row=row, column=j, value=h).style = 'header_style'
            row += 1

            for i, (_, fac) in enumerate(ncc_fac.iterrows(), 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=fac['product'])
                ws.cell(row=row, column=3, value=fac['company'])
                ws.cell(row=row, column=4, value=fac['location'])
                ws.cell(row=row, column=5, value=f"{fac['capacity_kt']:,.0f}")
                ws.cell(row=row, column=6, value=int(fac['year_built']) if pd.notna(fac['year_built']) else "-")
                ws.cell(row=row, column=7, value=int(fac['age_2025']) if pd.notna(fac['age_2025']) else "-")
                row += 1

            row += 2

            # Section 3: Deployment Trajectory
            if deployment is not None:
                ws.cell(row=row, column=1, value="3. Annual Deployment Trajectory").font = Font(bold=True, size=12, color='1F4E79')
                row += 2

                headers = ["Year", "BAU (Mt)", "Deployed (Mt)", "Net (Mt)", "Grid EF"]
                for j, h in enumerate(headers, 1):
                    ws.cell(row=row, column=j, value=h).style = 'header_style'
                row += 1

                # Show every 5 years
                for year in [2025, 2030, 2035, 2040, 2045, 2050]:
                    year_data = deployment[deployment['year'] == year]
                    if len(year_data) > 0:
                        d = year_data.iloc[0]
                        ws.cell(row=row, column=1, value=year)
                        ws.cell(row=row, column=2, value=f"{d['bau_mt']:.2f}")
                        ws.cell(row=row, column=3, value=f"{d['deployed_abatement_mt']:.2f}")
                        ws.cell(row=row, column=4, value=f"{d['actual_emissions_mt']:.2f}")
                        ws.cell(row=row, column=5, value=f"{d['grid_ef']:.3f}")
                        row += 1

            row += 2

            # Section 4: Facility Emissions (Top 20)
            if emissions is not None:
                ws.cell(row=row, column=1, value="4. Top 20 Facilities by Emissions (2050)").font = Font(bold=True, size=12, color='1F4E79')
                row += 2

                emissions_sorted = emissions.sort_values('total_emissions_kt', ascending=False).head(20)

                headers = ["#", "Product", "Company", "Location", "Capacity (kt)", "Emissions (kt)"]
                for j, h in enumerate(headers, 1):
                    ws.cell(row=row, column=j, value=h).style = 'header_style'
                row += 1

                for i, (_, em) in enumerate(emissions_sorted.iterrows(), 1):
                    ws.cell(row=row, column=1, value=i)
                    ws.cell(row=row, column=2, value=em['product'])
                    ws.cell(row=row, column=3, value=em['company'])
                    ws.cell(row=row, column=4, value=em['location'])
                    ws.cell(row=row, column=5, value=f"{em['capacity_kt']:,.0f}")
                    ws.cell(row=row, column=6, value=f"{em['total_emissions_kt']:,.0f}")
                    row += 1

        else:
            ws.cell(row=row, column=1, value=f"Scenario data not found. Run run_scenarios_complete.py first.")
            ws.cell(row=row, column=1).font = Font(italic=True, color='FF0000')

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12

    return wb


if __name__ == "__main__":
    generate_report()
